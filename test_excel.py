import openpyxl
import os

print("🚀 Probando lectura de Excel...")

# Probar con el primer archivo
archivo = "INVENTARIO ALMACEN SEPTIEMBRE-  2025 .xlsx"

if os.path.exists(archivo):
    print(f"✅ Archivo encontrado: {archivo}")
    try:
        workbook = openpyxl.load_workbook(archivo)
        print(f"✅ Archivo cargado exitosamente")
        print(f"📋 Hojas: {workbook.sheetnames}")
        
        # Leer la primera hoja
        sheet = workbook.active
        print(f"📏 Dimensiones: {sheet.max_row} filas x {sheet.max_column} columnas")
        
        # Leer primera fila
        print("📝 Primera fila:")
        for col in range(1, min(6, sheet.max_column + 1)):
            cell = sheet.cell(row=1, column=col)
            print(f"   Columna {col}: {cell.value}")
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ Error: {e}")
else:
    print(f"❌ Archivo no encontrado: {archivo}")
