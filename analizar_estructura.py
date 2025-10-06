import openpyxl
import os

def analizar_archivo(nombre_archivo):
    print(f"\n📊 ANALIZANDO: {nombre_archivo}")
    print("=" * 60)
    
    try:
        workbook = openpyxl.load_workbook(nombre_archivo)
        sheet = workbook.active
        
        print(f"📋 Hoja activa: {sheet.title}")
        print(f"📏 Dimensiones: {sheet.max_row} filas x {sheet.max_column} columnas")
        
        # Buscar la fila de encabezados (buscar fila con más texto)
        mejor_fila_headers = 1
        max_campos = 0
        
        for row in range(1, min(20, sheet.max_row + 1)):
            campos_no_vacios = 0
            for col in range(1, min(20, sheet.max_column + 1)):
                cell = sheet.cell(row=row, column=col)
                if cell.value and str(cell.value).strip():
                    campos_no_vacios += 1
            
            if campos_no_vacios > max_campos:
                max_campos = campos_no_vacios
                mejor_fila_headers = row
        
        print(f"🎯 Fila de encabezados detectada: {mejor_fila_headers}")
        
        # Mostrar encabezados
        print(f"\n📝 ENCABEZADOS:")
        headers = []
        for col in range(1, min(20, sheet.max_column + 1)):
            cell = sheet.cell(row=mejor_fila_headers, column=col)
            if cell.value:
                header = str(cell.value).strip()
                headers.append(header)
                print(f"   Columna {col}: {header}")
        
        # Mostrar algunas filas de datos
        print(f"\n📄 MUESTRA DE DATOS (filas {mejor_fila_headers + 1} a {mejor_fila_headers + 5}):")
        for row in range(mejor_fila_headers + 1, min(mejor_fila_headers + 6, sheet.max_row + 1)):
            print(f"\n   Fila {row}:")
            for col in range(1, min(len(headers) + 1, 15)):  # Solo primeras 15 columnas
                cell = sheet.cell(row=row, column=col)
                if cell.value:
                    valor = str(cell.value)[:50]  # Limitar a 50 caracteres
                    print(f"      {headers[col-1] if col-1 < len(headers) else f'Col{col}'}: {valor}")
        
        workbook.close()
        return headers
        
    except Exception as e:
        print(f"❌ Error: {e}")
        return []

# Analizar los 3 archivos
archivos = [
    "INVENTARIO ALMACEN SEPTIEMBRE-  2025 .xlsx",
    "INV QUIMICOS SEPTIEMBRE - 2025 (1) .xlsx", 
    "SALDOS POSCOSECHA  SEPTIEMBRE - 2025.xlsx"
]

estructuras = {}

for archivo in archivos:
    if os.path.exists(archivo):
        headers = analizar_archivo(archivo)
        estructuras[archivo] = headers
    else:
        print(f"❌ Archivo no encontrado: {archivo}")

print(f"\n🎯 RESUMEN DE ESTRUCTURAS:")
print("=" * 60)
for archivo, headers in estructuras.items():
    print(f"\n📁 {archivo}:")
    print(f"   Campos principales: {len(headers)}")
    if headers:
        print(f"   Primeros campos: {headers[:5]}")
