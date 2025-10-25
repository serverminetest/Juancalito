"""
Generador de Excel Automático para Inventarios
==============================================

Este script genera un archivo Excel con:
- Hojas separadas por categoría (ALMACEN GENERAL, QUIMICOS, POSCOSECHA)
- Fórmulas automáticas para calcular saldo real
- Datos actualizados desde la base de datos
- Formato profesional

Uso:
    python generar_excel_inventario.py
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.formula import ArrayFormula
import os
from datetime import datetime

def generar_excel_inventario(periodo='2025-10', output_file='Inventario_Automatico.xlsx'):
    """
    Genera un archivo Excel con inventario automático
    
    Args:
        periodo (str): Período a generar (formato: 2025-10)
        output_file (str): Nombre del archivo de salida
    """
    
    # Simular datos de la base de datos (en producción vendría de SQLAlchemy)
    # En la implementación real, esto vendría de:
    # productos = Producto.query.filter_by(periodo=periodo, activo=True).all()
    
    datos_ejemplo = {
        'ALMACEN GENERAL': [
            {'codigo': 'AG-001', 'nombre': 'CAJAS CARTON', 'unidad': 'UNIDAD', 'saldo_inicial': 100, 'entradas': 50, 'salidas': 30, 'precio_unitario': 5000},
            {'codigo': 'AG-002', 'nombre': 'ALAMBRE', 'unidad': 'KG', 'saldo_inicial': 200, 'entradas': 100, 'salidas': 80, 'precio_unitario': 15000},
            {'codigo': 'AG-003', 'nombre': 'PLASTICO', 'unidad': 'M2', 'saldo_inicial': 50, 'entradas': 25, 'salidas': 15, 'precio_unitario': 8000},
        ],
        'QUIMICOS': [
            {'codigo': 'Q-001', 'nombre': 'SULFATO', 'unidad': 'KG', 'saldo_inicial': 80, 'entradas': 40, 'salidas': 25, 'precio_unitario': 25000},
            {'codigo': 'Q-002', 'nombre': 'FERTILIZANTE', 'unidad': 'KG', 'saldo_inicial': 150, 'entradas': 75, 'salidas': 60, 'precio_unitario': 18000},
        ],
        'POSCOSECHA': [
            {'codigo': 'P-001', 'nombre': 'BANDEJAS', 'unidad': 'UNIDAD', 'saldo_inicial': 300, 'entradas': 150, 'salidas': 100, 'precio_unitario': 3000},
            {'codigo': 'P-002', 'nombre': 'BOLSAS', 'unidad': 'UNIDAD', 'saldo_inicial': 500, 'entradas': 200, 'salidas': 180, 'precio_unitario': 2000},
        ]
    }
    
    # Crear workbook
    wb = Workbook()
    
    # Eliminar hoja por defecto
    wb.remove(wb.active)
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Crear hoja para cada categoría
    for categoria, productos in datos_ejemplo.items():
        ws = wb.create_sheet(title=categoria)
        
        # Título de la hoja
        ws['A1'] = f'INVENTARIO {categoria} - PERÍODO {periodo}'
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = center_alignment
        ws.merge_cells('A1:H1')
        
        # Fecha de generación
        ws['A2'] = f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A2'].font = Font(italic=True)
        ws.merge_cells('A2:H2')
        
        # Encabezados
        headers = ['CÓDIGO', 'NOMBRE', 'UNIDAD', 'SALDO INICIAL', 'ENTRADAS', 'SALIDAS', 'SALDO REAL', 'PRECIO UNIT.', 'VALOR TOTAL']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Datos de productos
        for row, producto in enumerate(productos, 5):
            # Datos básicos
            ws.cell(row=row, column=1, value=producto['codigo']).border = border
            ws.cell(row=row, column=2, value=producto['nombre']).border = border
            ws.cell(row=row, column=3, value=producto['unidad']).border = border
            ws.cell(row=row, column=4, value=producto['saldo_inicial']).border = border
            ws.cell(row=row, column=5, value=producto['entradas']).border = border
            ws.cell(row=row, column=6, value=producto['salidas']).border = border
            
            # FÓRMULA: Saldo Real = Saldo Inicial + Entradas - Salidas
            formula_cell = ws.cell(row=row, column=7)
            formula_cell.value = f'=D{row}+E{row}-F{row}'
            formula_cell.border = border
            formula_cell.alignment = center_alignment
            
            # Precio unitario
            ws.cell(row=row, column=8, value=producto['precio_unitario']).border = border
            
            # FÓRMULA: Valor Total = Saldo Real × Precio Unitario
            valor_cell = ws.cell(row=row, column=9)
            valor_cell.value = f'=G{row}*H{row}'
            valor_cell.border = border
            valor_cell.alignment = center_alignment
        
        # Fila de totales
        total_row = len(productos) + 6
        ws.cell(row=total_row, column=1, value='TOTALES').font = Font(bold=True)
        ws.cell(row=total_row, column=1).border = border
        
        # FÓRMULAS DE TOTALES
        ws.cell(row=total_row, column=4, value=f'=SUM(D5:D{total_row-1})').border = border  # Total Saldo Inicial
        ws.cell(row=total_row, column=5, value=f'=SUM(E5:E{total_row-1})').border = border  # Total Entradas
        ws.cell(row=total_row, column=6, value=f'=SUM(F5:F{total_row-1})').border = border  # Total Salidas
        ws.cell(row=total_row, column=7, value=f'=SUM(G5:G{total_row-1})').border = border  # Total Saldo Real
        ws.cell(row=total_row, column=9, value=f'=SUM(I5:I{total_row-1})').border = border  # Total Valor
        
        # Aplicar formato a totales
        for col in range(1, 10):
            cell = ws.cell(row=total_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            cell.border = border
        
        # Ajustar ancho de columnas
        column_widths = [12, 25, 10, 12, 10, 10, 12, 12, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Crear hoja de resumen
    ws_resumen = wb.create_sheet(title='RESUMEN GENERAL', index=0)
    
    # Título del resumen
    ws_resumen['A1'] = f'RESUMEN GENERAL DE INVENTARIOS - PERÍODO {periodo}'
    ws_resumen['A1'].font = Font(bold=True, size=16)
    ws_resumen['A1'].alignment = center_alignment
    ws_resumen.merge_cells('A1:F1')
    
    # Encabezados del resumen
    resumen_headers = ['CATEGORÍA', 'TOTAL PRODUCTOS', 'TOTAL SALDO INICIAL', 'TOTAL ENTRADAS', 'TOTAL SALIDAS', 'TOTAL SALDO REAL', 'VALOR TOTAL']
    for col, header in enumerate(resumen_headers, 1):
        cell = ws_resumen.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Datos del resumen (en producción vendría de consultas SQL)
    resumen_data = [
        ['ALMACEN GENERAL', 3, 350, 175, 125, 400, 4000000],
        ['QUIMICOS', 2, 230, 115, 85, 260, 5200000],
        ['POSCOSECHA', 2, 800, 350, 280, 870, 2610000],
    ]
    
    for row, data in enumerate(resumen_data, 4):
        for col, value in enumerate(data, 1):
            cell = ws_resumen.cell(row=row, column=col, value=value)
            cell.border = border
            if col == 1:  # Categoría
                cell.font = Font(bold=True)
    
    # Totales del resumen
    total_resumen_row = len(resumen_data) + 5
    ws_resumen.cell(row=total_resumen_row, column=1, value='TOTAL GENERAL').font = Font(bold=True)
    ws_resumen.cell(row=total_resumen_row, column=1).border = border
    
    # Fórmulas de totales en resumen
    for col in range(2, 8):
        formula_cell = ws_resumen.cell(row=total_resumen_row, column=col)
        formula_cell.value = f'=SUM({ws_resumen.cell(row=4, column=col).column_letter}4:{ws_resumen.cell(row=total_resumen_row-1, column=col).column_letter}{total_resumen_row-1})'
        formula_cell.border = border
        formula_cell.font = Font(bold=True)
        formula_cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    
    # Ajustar ancho de columnas en resumen
    resumen_widths = [20, 15, 18, 15, 15, 18, 15]
    for col, width in enumerate(resumen_widths, 1):
        ws_resumen.column_dimensions[ws_resumen.cell(row=1, column=col).column_letter].width = width
    
    # Guardar archivo
    wb.save(output_file)
    print(f"✅ Excel generado exitosamente: {output_file}")
    return output_file

if __name__ == "__main__":
    # Generar Excel de ejemplo
    generar_excel_inventario('2025-10', 'Inventario_Automatico.xlsx')
