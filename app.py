from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file  # pyright: ignore[reportMissingImports]
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, date, timedelta, timezone
from sqlalchemy import text
import os
import qrcode
import io
import hashlib
import secrets
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import shutil

# Importar sistema de notificaciones
from notificaciones import (
    notificacion_manager, 
    notificar_asistencia_entrada, 
    notificar_asistencia_salida,
    notificar_visitante_nuevo,
    notificar_visitante_salida,
    notificar_error,
    notificar_exito,
    obtener_notificaciones_api,
    marcar_notificacion_leida_api,
    limpiar_notificaciones_api
)

# Configurar zona horaria de Colombia (UTC-5)
COLOMBIA_TZ = timezone(timedelta(hours=-5))

# Token global para QR (mismo para asistencia y visitantes)
QR_TOKEN_SECRETO_GLOBAL = os.environ.get('QR_TOKEN_GLOBAL', 'flores_juncalito_qr_global')
def generar_token_qr_constante():
    """Genera un token estable para los QR públicos"""
    return hashlib.sha256(QR_TOKEN_SECRETO_GLOBAL.encode()).hexdigest()[:32]

def colombia_now():
    """Devuelve la fecha y hora actual en zona horaria de Colombia"""
    return datetime.now(COLOMBIA_TZ)

def es_festivo_colombia(fecha):
    """Verifica si una fecha es festivo en Colombia"""
    # Festivos fijos en Colombia
    festivos_fijos = [
        (1, 1),   # Año Nuevo
        (5, 1),   # Día del Trabajo
        (7, 20),  # Día de la Independencia
        (8, 7),   # Batalla de Boyacá
        (12, 8),  # Día de la Inmaculada Concepción
        (12, 25), # Navidad
    ]
    
    # Festivos que dependen del año (Pascua y relacionados)
    # Para simplificar, usaremos una aproximación básica
    # En producción, se recomienda usar una librería como 'holidays' o calcular Pascua
    
    if (fecha.month, fecha.day) in festivos_fijos:
        return True
    
    # Aquí se pueden agregar más festivos calculados (Pascua, etc.)
    # Por ahora, solo verificamos festivos fijos
    
    return False

def calcular_fecha_reintegro(fecha_inicio, cantidad_dias):
    """Calcula la fecha de reintegro excluyendo domingos y festivos de Colombia"""
    if isinstance(fecha_inicio, str):
        fecha_actual = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
    else:
        fecha_actual = fecha_inicio
    dias_agregados = 0
    
    while dias_agregados < cantidad_dias:
        fecha_actual += timedelta(days=1)
        # Si no es domingo (0 = lunes, 6 = domingo) y no es festivo
        if fecha_actual.weekday() != 6 and not es_festivo_colombia(fecha_actual):
            dias_agregados += 1
    
    return fecha_actual

def get_periodo_actual():
    """Devuelve el período actual en formato YYYY-MM"""
    return datetime.now().strftime('%Y-%m')

def get_periodo_desde_params():
    """Obtiene el período desde los parámetros de la URL, por defecto el actual"""
    periodo = request.args.get('periodo', get_periodo_actual())
    # Validar formato YYYY-MM
    try:
        datetime.strptime(periodo, '%Y-%m')
        return periodo
    except ValueError:
        return get_periodo_actual()

def to_colombia_time(dt):
    """Convierte una fecha/hora a zona horaria de Colombia"""
    if dt is None:
        return None
    if dt.tzinfo is None:
        # Si no tiene zona horaria, asumir que es UTC
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(COLOMBIA_TZ)

def generar_contrato_excel(contrato_id):
    """Genera un contrato Excel basado en el template y los datos del empleado"""
    try:
        # Obtener datos del contrato y empleado
        contrato = Contrato.query.get_or_404(contrato_id)
        empleado = contrato.empleado
        
        # Cargar template Excel
        template_path = 'CONTRATO EXCEL FLORE JUNCALITO.xlsx'
        if not os.path.exists(template_path):
            raise FileNotFoundError("Template de contrato no encontrado")
        
        # Crear directorio para contratos generados
        contratos_dir = 'contratos_generados'
        if not os.path.exists(contratos_dir):
            os.makedirs(contratos_dir)
        
        # Generar nombre único para el archivo
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_archivo = f"Contrato_{empleado.nombre_completo.replace(' ', '_')}_{timestamp}.xlsx"
        ruta_archivo = os.path.join(contratos_dir, nombre_archivo)
        
        # Copiar template y cargar workbook
        shutil.copy2(template_path, ruta_archivo)
        workbook = load_workbook(ruta_archivo)
        worksheet = workbook.active
        
        # Datos del empleador (predeterminados según la imagen)
        datos_empleador = {
            'NOMBRE_EMPLEADOR': 'FLORES JUNCALITO S.A.S',
            'DIRECCION_EMPLEADOR': 'CALLE 19* C N. 88-07'
        }
        
        # Datos del empleado
        datos_empleado = {
            'NOMBRE_EMPLEADO': empleado.nombre_completo.upper(),
            'NOMBRE_TRABAJADOR': empleado.nombre_completo.upper(),
            'DIRECCION_EMPLEADO': (empleado.direccion_residencia or 'NO ESPECIFICADA').upper(),
            'DIRECCION_TRABAJADOR': (empleado.direccion_residencia or 'NO ESPECIFICADA').upper(),
            'LUGAR_NACIMIENTO': (empleado.ciudad or 'BOGOTÁ, COLOMBIA').upper(),
            'FECHA_NACIMIENTO': convertir_fecha_espanol(empleado.fecha_nacimiento),
            'CARGO_EMPLEADO': (empleado.cargo_puesto or 'NO ESPECIFICADO').upper(),
            'CARGO': (empleado.cargo_puesto or 'NO ESPECIFICADO').upper(),
            'SALARIO_NUMEROS': f"$ {contrato.salario:,.0f}",
            'SALARIO': f"$ {contrato.salario:,.0f}",
            'SALARIO_LETRAS': convertir_numero_a_letras(contrato.salario),
            'VALOR_LETRAS': convertir_numero_a_letras(contrato.salario),
            'FECHA_INICIO_LABORES': convertir_fecha_espanol(contrato.fecha_inicio),
            'FECHA_INICIO': convertir_fecha_espanol(contrato.fecha_inicio),
            'FECHA_FIN': convertir_fecha_espanol(contrato.fecha_fin) if contrato.fecha_fin else 'INDEFINIDO',
            'VENCE_EL_DIA': convertir_fecha_espanol(contrato.fecha_fin) if contrato.fecha_fin else 'INDEFINIDO',
            'LUGAR_LABORES': 'FLORES JUNCALITO S.A.S',
            'LUGAR_TRABAJO': 'FLORES JUNCALITO S.A.S',
            'CIUDAD_CONTRATACION': 'EL ROSAL CUNDINAMARCA',
            'CIUDAD_CONTRATO': 'EL ROSAL CUNDINAMARCA',
            'TIPO_SALARIO': 'ORDINARIO',
            'PERIODOS_PAGO': 'MENSUAL'
        }
        
        # Combinar todos los datos
        todos_datos = {**datos_empleador, **datos_empleado}
        
        # Reemplazar variables en el Excel
        reemplazar_variables_excel(worksheet, todos_datos)
        
        # Guardar el archivo
        workbook.save(ruta_archivo)
        
        # Leer el archivo como datos binarios
        with open(ruta_archivo, 'rb') as f:
            archivo_data = f.read()
        
        # Verificar si la tabla contrato_generado existe antes de insertar
        try:
            # Intentar crear la tabla si no existe
            try:
                with db.engine.connect() as connection:
                    connection.execute(text("""
                        CREATE TABLE IF NOT EXISTS contrato_generado (
                            id SERIAL PRIMARY KEY,
                            empleado_id INTEGER NOT NULL REFERENCES empleado(id),
                            contrato_id INTEGER NOT NULL REFERENCES contrato(id),
                            nombre_archivo VARCHAR(255) NOT NULL,
                            ruta_archivo VARCHAR(500) NOT NULL,
                            archivo_data BYTEA,
                            fecha_generacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                            activo BOOLEAN DEFAULT TRUE
                        );
                    """))
                    connection.commit()
                print("✅ Tabla contrato_generado creada/verificada en generar_contrato_excel")
            except Exception as create_error:
                print(f"⚠️ No se pudo crear la tabla en generar_contrato_excel: {create_error}")
            
            # Registrar en la base de datos con datos binarios
            contrato_generado = ContratoGenerado(
                empleado_id=empleado.id,
                contrato_id=contrato.id,
                nombre_archivo=nombre_archivo,
                ruta_archivo=ruta_archivo,
                archivo_data=archivo_data
            )
            db.session.add(contrato_generado)
            db.session.commit()
            
            # Limpiar archivo temporal (opcional, ya que tenemos los datos en BD)
            try:
                os.remove(ruta_archivo)
                print(f"✅ Archivo temporal eliminado: {ruta_archivo}")
            except:
                pass  # No es crítico si no se puede eliminar
            
            print(f"✅ Contrato generado y guardado en BD: {nombre_archivo}")
            return contrato_generado
        except Exception as db_error:
            print(f"Error al guardar en base de datos: {str(db_error)}")
            # Si hay error de BD, crear un objeto mock que simule ContratoGenerado
            class MockContratoGenerado:
                def __init__(self, nombre_archivo, ruta_archivo, empleado, contrato, archivo_data):
                    self.nombre_archivo = nombre_archivo
                    self.ruta_archivo = ruta_archivo
                    self.archivo_data = archivo_data
                    self.empleado = empleado
                    self.contrato = contrato
                    self.id = None  # No tiene ID porque no se guardó en BD
            
            return MockContratoGenerado(nombre_archivo, ruta_archivo, empleado, contrato, archivo_data)
        
    except Exception as e:
        print(f"Error al generar contrato: {str(e)}")
        raise

def reemplazar_variables_excel(worksheet, datos):
    """Reemplaza las variables {VARIABLE} en el Excel con los datos reales"""
    try:
        variables_reemplazadas = 0
        variables_no_encontradas = []
        
        # Recorrer todas las celdas del worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Buscar variables en formato {VARIABLE}
                    import re
                    variables_encontradas = re.findall(r'\{([^}]+)\}', cell.value)
                    
                    if variables_encontradas:
                        valor_original = cell.value
                        valor_nuevo = valor_original
                        
                        # Reemplazar cada variable encontrada
                        for variable in variables_encontradas:
                            if variable in datos:
                                valor_nuevo = valor_nuevo.replace(f'{{{variable}}}', str(datos[variable]))
                                variables_reemplazadas += 1
                            else:
                                variables_no_encontradas.append(variable)
                                print(f"⚠️ Variable no encontrada: {variable}")
                        
                        # Actualizar el valor de la celda
                        cell.value = valor_nuevo
                        if valor_original != valor_nuevo:
                            print(f"✅ Reemplazado: {valor_original} -> {valor_nuevo}")
        
        print(f"✅ Variables reemplazadas exitosamente: {variables_reemplazadas} reemplazos")
        if variables_no_encontradas:
            print(f"⚠️ Variables no encontradas: {', '.join(set(variables_no_encontradas))}")
        
        # Mostrar todas las variables disponibles
        print(f"📋 Variables disponibles: {', '.join(datos.keys())}")
        
    except Exception as e:
        print(f"❌ Error al reemplazar variables: {str(e)}")

def convertir_numero_a_letras(numero):
    """Convierte un número a letras (mejorado para salarios)"""
    try:
        if numero == 0:
            return 'CERO PESOS'
        
        # Convertir a entero para evitar decimales
        numero = int(numero)
        
        # Nombres de números
        unidades = ['', 'UNO', 'DOS', 'TRES', 'CUATRO', 'CINCO', 'SEIS', 'SIETE', 'OCHO', 'NUEVE']
        decenas = ['', '', 'VEINTE', 'TREINTA', 'CUARENTA', 'CINCUENTA', 'SESENTA', 'SETENTA', 'OCHENTA', 'NOVENTA']
        centenas = ['', 'CIENTO', 'DOSCIENTOS', 'TRESCIENTOS', 'CUATROCIENTOS', 'QUINIENTOS', 'SEISCIENTOS', 'SETECIENTOS', 'OCHOCIENTOS', 'NOVECIENTOS']
        
        # Casos especiales
        if numero == 100:
            return 'CIEN PESOS'
        if numero == 1000:
            return 'MIL PESOS'
        if numero == 1000000:
            return 'UN MILLON DE PESOS'
        
        resultado = ''
        
        # Millones
        millones = numero // 1000000
        if millones > 0:
            if millones == 1:
                resultado += 'UN MILLON '
            elif millones < 10:
                resultado += f"{unidades[millones]} MILLONES "
            else:
                # Para números mayores a 9 millones, convertir a letras
                resultado += convertir_centenas_miles(millones) + " MILLONES "
            numero = numero % 1000000
        
        # Miles
        miles = numero // 1000
        if miles > 0:
            if miles == 1:
                resultado += 'MIL '
            elif miles < 10:
                resultado += f"{unidades[miles]} MIL "
            else:
                # Para números mayores a 9 mil, convertir a letras
                resultado += convertir_centenas_miles(miles) + " MIL "
            numero = numero % 1000
        
        # Centenas, decenas y unidades restantes
        if numero > 0:
            resultado += convertir_centenas_miles(numero) + " "
        
        resultado += 'PESOS'
        return resultado.strip()
        
    except Exception as e:
        print(f"Error al convertir número a letras: {str(e)}")
        return f"{numero} PESOS"

def convertir_centenas_miles(numero):
    """Convierte números de 1 a 999 a letras"""
    if numero == 0:
        return ''
    
    unidades = ['', 'UNO', 'DOS', 'TRES', 'CUATRO', 'CINCO', 'SEIS', 'SIETE', 'OCHO', 'NUEVE']
    decenas = ['', '', 'VEINTE', 'TREINTA', 'CUARENTA', 'CINCUENTA', 'SESENTA', 'SETENTA', 'OCHENTA', 'NOVENTA']
    centenas = ['', 'CIENTO', 'DOSCIENTOS', 'TRESCIENTOS', 'CUATROCIENTOS', 'QUINIENTOS', 'SEISCIENTOS', 'SETECIENTOS', 'OCHOCIENTOS', 'NOVECIENTOS']
    
    resultado = ''
    
    # Centenas
    centena = numero // 100
    if centena > 0:
        if centena == 1 and numero % 100 == 0:
            resultado += 'CIEN'
        else:
            resultado += centenas[centena]
        numero = numero % 100
    
    # Decenas y unidades
    if numero > 0:
        if numero < 10:
            resultado += unidades[numero]
        elif numero < 20:
            if numero == 10:
                resultado += 'DIEZ'
            elif numero == 11:
                resultado += 'ONCE'
            elif numero == 12:
                resultado += 'DOCE'
            elif numero == 13:
                resultado += 'TRECE'
            elif numero == 14:
                resultado += 'CATORCE'
            elif numero == 15:
                resultado += 'QUINCE'
            elif numero == 16:
                resultado += 'DIECISEIS'
            elif numero == 17:
                resultado += 'DIECISIETE'
            elif numero == 18:
                resultado += 'DIECIOCHO'
            elif numero == 19:
                resultado += 'DIECINUEVE'
        else:
            decena = numero // 10
            unidad = numero % 10
            if unidad == 0:
                resultado += decenas[decena]
            else:
                resultado += decenas[decena] + " Y " + unidades[unidad]
    
    return resultado

def convertir_fecha_espanol(fecha):
    """Convierte una fecha a formato español"""
    if fecha is None:
        return 'INDEFINIDO'
    
    # Mapeo de meses en inglés a español
    meses_espanol = {
        1: 'ENERO', 2: 'FEBRERO', 3: 'MARZO', 4: 'ABRIL',
        5: 'MAYO', 6: 'JUNIO', 7: 'JULIO', 8: 'AGOSTO',
        9: 'SEPTIEMBRE', 10: 'OCTUBRE', 11: 'NOVIEMBRE', 12: 'DICIEMBRE'
    }
    
    try:
        dia = fecha.day
        mes = meses_espanol[fecha.month]
        año = fecha.year
        
        return f"{dia} DE {mes} DE {año}"
    except Exception as e:
        print(f"Error al convertir fecha: {str(e)}")
        return str(fecha)

def convertir_excel_a_html(worksheet, contrato_generado):
    """Convierte una hoja de Excel a HTML para vista previa con formato fiel"""
    try:
        html = '<div class="vista-previa-excel">'
        html += f'<div class="text-center mb-4">'
        html += f'<h4 class="text-primary"><i class="fas fa-file-excel me-2"></i>Vista Previa del Contrato</h4>'
        html += f'<p class="text-muted">Empleado: <strong>{contrato_generado.empleado.nombre_completo}</strong></p>'
        html += f'</div>'
        
        # Crear tabla HTML con estilos más fieles al Excel
        html += '<div class="table-responsive" style="overflow-x: auto;">'
        html += '<table class="table table-bordered" style="font-family: Arial, sans-serif; font-size: 11px; border-collapse: collapse; width: 100%;">'
        
        # Obtener dimensiones de la hoja
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # Procesar cada fila
        for row in range(1, min(max_row + 1, 100)):  # Aumentar límite para ver más contenido
            html += '<tr style="height: 20px;">'
            
            for col in range(1, min(max_col + 1, 30)):  # Aumentar límite de columnas
                cell = worksheet.cell(row=row, column=col)
                cell_value = str(cell.value) if cell.value is not None else ''
                
                # Aplicar estilos más detallados
                cell_style = 'padding: 4px 8px; border: 1px solid #000; vertical-align: top; '
                
                # Estilos de fuente
                if cell.font:
                    if cell.font.bold:
                        cell_style += 'font-weight: bold; '
                    if cell.font.italic:
                        cell_style += 'font-style: italic; '
                    if cell.font.size:
                        cell_style += f'font-size: {cell.font.size}px; '
                    if cell.font.color and cell.font.color.rgb:
                        cell_style += f'color: {cell.font.color.rgb}; '
                
                # Estilos de relleno
                if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                    cell_style += f'background-color: {cell.fill.start_color.rgb}; '
                
                # Estilos de alineación
                if cell.alignment:
                    if cell.alignment.horizontal == 'center':
                        cell_style += 'text-align: center; '
                    elif cell.alignment.horizontal == 'right':
                        cell_style += 'text-align: right; '
                    elif cell.alignment.horizontal == 'left':
                        cell_style += 'text-align: left; '
                    
                    if cell.alignment.vertical == 'center':
                        cell_style += 'vertical-align: middle; '
                    elif cell.alignment.vertical == 'bottom':
                        cell_style += 'vertical-align: bottom; '
                
                # Estilos de borde
                if cell.border:
                    if cell.border.left and cell.border.left.style:
                        cell_style += 'border-left: 2px solid #000; '
                    if cell.border.right and cell.border.right.style:
                        cell_style += 'border-right: 2px solid #000; '
                    if cell.border.top and cell.border.top.style:
                        cell_style += 'border-top: 2px solid #000; '
                    if cell.border.bottom and cell.border.bottom.style:
                        cell_style += 'border-bottom: 2px solid #000; '
                
                # Determinar el tipo de celda
                if row == 1 or (cell.font and cell.font.bold):
                    html += f'<th style="{cell_style}">{cell_value}</th>'
                else:
                    html += f'<td style="{cell_style}">{cell_value}</td>'
            
            html += '</tr>'
        
        html += '</table>'
        html += '</div>'
        
        # Información adicional
        html += '<div class="mt-3">'
        html += '<div class="alert alert-info">'
        html += '<h6><i class="fas fa-info-circle me-2"></i>Información del Contrato:</h6>'
        html += f'<ul class="mb-0">'
        html += f'<li><strong>Empleado:</strong> {contrato_generado.empleado.nombre_completo}</li>'
        html += f'<li><strong>Cédula:</strong> {contrato_generado.empleado.cedula}</li>'
        html += f'<li><strong>Cargo:</strong> {contrato_generado.empleado.cargo_puesto or "No especificado"}</li>'
        html += f'<li><strong>Salario:</strong> $ {contrato_generado.contrato.salario:,.0f}</li>'
        html += f'<li><strong>Tipo de Contrato:</strong> {contrato_generado.contrato.tipo_contrato}</li>'
        html += f'<li><strong>Fecha de Inicio:</strong> {contrato_generado.contrato.fecha_inicio.strftime("%d/%m/%Y")}</li>'
        if contrato_generado.contrato.fecha_fin:
            html += f'<li><strong>Fecha de Fin:</strong> {contrato_generado.contrato.fecha_fin.strftime("%d/%m/%Y")}</li>'
        else:
            html += f'<li><strong>Fecha de Fin:</strong> Indefinido</li>'
        html += f'</ul>'
        html += '</div>'
        html += '</div>'
        
        html += '</div>'
        
        return html
        
    except Exception as e:
        print(f"Error al convertir Excel a HTML: {str(e)}")
        return f'''
        <div class="alert alert-warning">
            <h6><i class="fas fa-exclamation-triangle me-2"></i>Error en Vista Previa</h6>
            <p>No se pudo generar la vista previa del contrato. Error: {str(e)}</p>
            <p>Puedes descargar el archivo Excel para verlo directamente.</p>
        </div>
        '''

app = Flask(__name__)

# Configuración de la aplicación
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'tu_clave_secreta_muy_segura_aqui')

# Registrar filtro de zona horaria para templates
@app.template_filter('colombia_time')
def colombia_time_filter(dt):
    """Filtro para convertir fechas a zona horaria de Colombia en templates"""
    return to_colombia_time(dt)

# Context processor para pasar variables globales a todos los templates
@app.context_processor
def inject_global_vars():
    """Inyecta variables globales en todos los templates"""
    try:
        solicitudes_pendientes = SolicitudEmpleado.query.filter_by(estado='PENDIENTE').count()
    except:
        solicitudes_pendientes = 0
    return dict(solicitudes_pendientes=solicitudes_pendientes)

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Configuración de base de datos
database_url = os.environ.get('DATABASE_URL')
if database_url:
    # Producción: usar PostgreSQL con psycopg
    # Convertir postgresql:// a postgresql+psycopg:// para usar psycopg en lugar de psycopg2
    if database_url.startswith('postgresql://'):
        database_url = database_url.replace('postgresql://', 'postgresql+psycopg://', 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = database_url
    print(f"🔗 Usando PostgreSQL con psycopg en producción")
else:
    # Verificar si estamos en producción (Railway, Heroku, etc.)
    if os.environ.get('RAILWAY_ENVIRONMENT') or os.environ.get('DYNO') or os.environ.get('PORT'):
        # Estamos en producción pero no hay DATABASE_URL - ERROR
        raise RuntimeError(
            "❌ ERROR: DATABASE_URL no configurada en producción.\n"
            "💡 Solución: Agrega una base de datos PostgreSQL en Railway:\n"
            "   1. Ve a tu proyecto en Railway\n"
            "   2. Haz clic en '+ New' → 'Database' → 'PostgreSQL'\n"
            "   3. Railway configurará automáticamente DATABASE_URL"
        )
    else:
        # Desarrollo local: usar SQLite
        app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///empleados.db'
        print("💾 Usando SQLite para desarrollo local")

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Modelos de Base de Datos
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    username = db.Column(db.String(80), nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)
    is_admin = db.Column(db.Boolean, default=True)

class Empleado(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    
    # Información Personal
    nombre_completo = db.Column(db.String(200), nullable=False)
    cedula = db.Column(db.String(20), unique=True, nullable=False)
    fecha_nacimiento = db.Column(db.Date, nullable=False)
    genero = db.Column(db.String(20), nullable=False)  # Masculino, Femenino, Otro
    estado_civil = db.Column(db.String(30), nullable=False)  # Soltero, Casado, etc.
    fecha_expedicion_documento = db.Column(db.Date)  # Fecha de expedición del documento
    lugar_expedicion_documento = db.Column(db.String(200))  # Lugar de expedición del documento
    
    # Contacto
    telefono_principal = db.Column(db.String(20), nullable=False)
    telefono_secundario = db.Column(db.String(20))
    email_personal = db.Column(db.String(120))  # Ahora es opcional
    email_corporativo = db.Column(db.String(120))
    
    # Dirección
    direccion_residencia = db.Column(db.Text, nullable=False)
    ciudad = db.Column(db.String(100), nullable=False)
    departamento = db.Column(db.String(100), nullable=False)
    codigo_postal = db.Column(db.String(10))
    
    # Información Laboral
    cargo_puesto = db.Column(db.String(100), nullable=False)
    departamento_laboral = db.Column(db.String(50), nullable=False)  # Poscosecha, Cultivo, Administrativo, etc.
    fecha_ingreso = db.Column(db.Date, nullable=False)
    tipo_contrato = db.Column(db.String(30), nullable=False)  # Temporal, Indefinido, etc.
    salario_base = db.Column(db.Float, nullable=False)
    tipo_salario = db.Column(db.String(20), nullable=False)  # Mensual, Quincenal, etc.
    jornada_laboral = db.Column(db.String(30), nullable=False)  # Tiempo completo, Medio tiempo, etc.
    ubicacion_trabajo = db.Column(db.String(20), nullable=False)  # Oficina, Remoto, Híbrido
    estado_empleado = db.Column(db.String(20), nullable=False, default='Activo')  # Activo, Inactivo, Suspendido
    supervisor = db.Column(db.String(100))
    horario = db.Column(db.String(100))
    
    # Información de Seguridad Social
    eps = db.Column(db.String(100), nullable=False)
    arl = db.Column(db.String(100), nullable=False)
    afp = db.Column(db.String(100), nullable=False)
    caja_compensacion = db.Column(db.String(100))
    
    # Contacto de Emergencia
    nombre_contacto_emergencia = db.Column(db.String(200), nullable=False)
    telefono_emergencia = db.Column(db.String(20), nullable=False)
    parentesco = db.Column(db.String(50), nullable=False)
    
    # Campos del sistema
    created_at = db.Column(db.DateTime, default=colombia_now)
    updated_at = db.Column(db.DateTime, default=colombia_now, onupdate=colombia_now)

class Contrato(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    empleado_id = db.Column(db.Integer, db.ForeignKey('empleado.id'), nullable=False)
    tipo_contrato = db.Column(db.String(50), nullable=False)  # Temporal, Indefinido, etc.
    fecha_inicio = db.Column(db.Date, nullable=False)
    fecha_fin = db.Column(db.Date)
    salario = db.Column(db.Float, nullable=False)
    descripcion = db.Column(db.Text)
    activo = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=colombia_now)
    
    # Relación con Empleado
    empleado = db.relationship('Empleado', backref=db.backref('contratos', lazy=True))

class ContratoGenerado(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    empleado_id = db.Column(db.Integer, db.ForeignKey('empleado.id'), nullable=False)
    contrato_id = db.Column(db.Integer, db.ForeignKey('contrato.id'), nullable=False)
    nombre_archivo = db.Column(db.String(255), nullable=False)
    ruta_archivo = db.Column(db.String(500), nullable=False)
    archivo_data = db.Column(db.LargeBinary, nullable=True)  # Datos binarios del archivo
    fecha_generacion = db.Column(db.DateTime, default=colombia_now)
    activo = db.Column(db.Boolean, default=True)
    
    # Relaciones
    empleado = db.relationship('Empleado', backref='contratos_generados')
    contrato = db.relationship('Contrato', backref='documentos_generados')

# Modelos para Sistema de Inventarios
class Producto(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(50), nullable=False)
    nombre = db.Column(db.String(200), nullable=False)
    descripcion = db.Column(db.Text)
    categoria = db.Column(db.String(50), nullable=False)  # ALMACEN GENERAL, QUIMICOS, POSCOSECHA
    periodo = db.Column(db.String(7), nullable=False)  # Formato: 2025-09 (año-mes)
    unidad_medida = db.Column(db.String(20), nullable=False)  # kg, litros, unidades, etc.
    precio_unitario = db.Column(db.Numeric(10, 2), default=0)
    stock_minimo = db.Column(db.Integer, default=0)
    saldo_inicial = db.Column(db.Integer, default=0)  # Saldo al inicio del período
    stock_actual = db.Column(db.Integer, default=0)
    ubicacion = db.Column(db.String(100))  # Estante, sección, etc.
    proveedor = db.Column(db.String(200))
    fecha_vencimiento = db.Column(db.Date)
    lote = db.Column(db.String(50))
    activo = db.Column(db.Boolean, default=True)
    mes_cerrado = db.Column(db.Boolean, default=False)  # Si el mes está cerrado, no se puede editar
    created_at = db.Column(db.DateTime, default=colombia_now)
    updated_at = db.Column(db.DateTime, default=colombia_now, onupdate=colombia_now)
    
    # Índice único para código + categoría + período (permite mismo código en diferentes meses/categorías)
    __table_args__ = (db.UniqueConstraint('codigo', 'categoria', 'periodo', name='_producto_codigo_categoria_periodo_uc'),)
    
    # Relación con movimientos
    movimientos = db.relationship('MovimientoInventario', backref='producto', lazy=True)
    
    def calcular_entradas(self):
        """Calcula el total de entradas del período"""
        return sum(m.calcular_cantidad_total() for m in self.movimientos if m.tipo_movimiento == 'ENTRADA')
    
    def calcular_salidas(self):
        """Calcula el total de salidas del período"""
        return sum(m.calcular_cantidad_total() for m in self.movimientos if m.tipo_movimiento == 'SALIDA')
    
    def calcular_saldo_final(self):
        """Calcula el saldo final: Saldo Inicial + Entradas - Salidas"""
        return self.saldo_inicial + self.calcular_entradas() - self.calcular_salidas()
    
    def calcular_stock_desde_movimientos(self):
        """Calcula el stock basado solo en movimientos, sin usar saldo_inicial"""
        entradas = sum(m.calcular_cantidad_total() for m in self.movimientos if m.tipo_movimiento == 'ENTRADA')
        salidas = sum(m.calcular_cantidad_total() for m in self.movimientos if m.tipo_movimiento == 'SALIDA')
        return entradas - salidas
    
    def recalcular_stock(self):
        """Recalcula y actualiza el stock_actual basado en saldo inicial y movimientos"""
        self.stock_actual = self.calcular_saldo_final()
        return self.stock_actual
    
    def verificar_stock_bajo(self):
        """Verifica si el stock está por debajo del mínimo"""
        return self.stock_actual < self.stock_minimo if self.stock_minimo > 0 else False
    
    def debe_tener_precio(self):
        """Determina si el producto debe tener precio unitario según su categoría"""
        return self.categoria in ['QUIMICOS', 'POSCOSECHA']
    
    @staticmethod
    def generar_codigo_automatico(categoria, periodo=None):
        """Genera un código automático basado en la categoría"""
        if periodo is None:
            periodo = get_periodo_actual()
        
        # Prefijos por categoría
        prefijos = {
            'ALMACEN GENERAL': 'ALM',
            'QUIMICOS': 'QUI', 
            'POSCOSECHA': 'POS'
        }
        
        prefijo = prefijos.get(categoria, 'GEN')
        
        # Buscar el último código de esta categoría en el período
        ultimo_producto = Producto.query.filter(
            Producto.categoria == categoria,
            Producto.periodo == periodo
        ).order_by(Producto.id.desc()).first()
        
        if ultimo_producto and ultimo_producto.codigo:
            # Extraer el número del último código
            try:
                # Buscar el patrón PREFIJO-NUMERO
                import re
                match = re.search(rf'^{prefijo}-(\d+)$', ultimo_producto.codigo)
                if match:
                    ultimo_numero = int(match.group(1))
                    nuevo_numero = ultimo_numero + 1
                else:
                    nuevo_numero = 1
            except:
                nuevo_numero = 1
        else:
            nuevo_numero = 1
        
        # Formatear con ceros a la izquierda (3 dígitos)
        return f"{prefijo}-{nuevo_numero:03d}"

class MovimientoInventario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    producto_id = db.Column(db.Integer, db.ForeignKey('producto.id'), nullable=False)
    periodo = db.Column(db.String(7), nullable=False)  # Formato: 2025-09 (año-mes)
    tipo_movimiento = db.Column(db.String(20), nullable=False)  # ENTRADA, SALIDA
    cantidad = db.Column(db.Integer, nullable=False)
    precio_unitario = db.Column(db.Numeric(10, 2))
    total = db.Column(db.Numeric(10, 2))
    motivo = db.Column(db.String(200))
    referencia = db.Column(db.String(100))  # Factura, orden de compra, etc.
    responsable = db.Column(db.String(200))
    observaciones = db.Column(db.Text)
    fecha_movimiento = db.Column(db.DateTime, default=colombia_now)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    proveedor = db.Column(db.String(200))
    
    # Nuevos campos para sistema de empaques
    tipo_ingreso = db.Column(db.String(20), default='INDIVIDUAL')  # EMPAQUE, INDIVIDUAL
    cantidad_empaques = db.Column(db.Integer, nullable=True)
    contenido_por_empaque = db.Column(db.Numeric(10, 2), nullable=True)
    precio_por_empaque = db.Column(db.Numeric(15, 2), nullable=True)
    
    # Relación con usuario
    usuario = db.relationship('User', backref='movimientos_inventario')
    
    def debe_tener_precio(self):
        """Determina si el movimiento debe tener precio unitario y total"""
        return self.tipo_movimiento == 'ENTRADA'
    
    def calcular_cantidad_total(self):
        """Calcula la cantidad total en unidad base según el tipo de ingreso"""
        if self.tipo_ingreso == 'EMPAQUE' and self.cantidad_empaques and self.contenido_por_empaque:
            return int(self.cantidad_empaques * self.contenido_por_empaque)
        return self.cantidad
    
    def calcular_valor_total(self):
        """Calcula el valor total según el tipo de ingreso"""
        if self.tipo_ingreso == 'EMPAQUE' and self.cantidad_empaques and self.precio_por_empaque:
            return float(self.cantidad_empaques * self.precio_por_empaque)
        elif self.precio_unitario:
            return float(self.cantidad * self.precio_unitario)
        return 0.0
    
    def es_ingreso_por_empaques(self):
        """Verifica si es un ingreso por empaques"""
        return self.tipo_ingreso == 'EMPAQUE' and self.tipo_movimiento == 'ENTRADA'
    
    def obtener_descripcion_ingreso(self):
        """Retorna una descripción del tipo de ingreso"""
        if self.tipo_ingreso == 'EMPAQUE' and self.cantidad_empaques and self.contenido_por_empaque:
            return f"{self.cantidad_empaques} empaques de {self.contenido_por_empaque} c/u"
        return f"{self.cantidad} unidades individuales"

class Notificacion(db.Model):
    __tablename__ = 'notificacion'
    
    id = db.Column(db.Integer, primary_key=True)
    titulo = db.Column(db.String(200), nullable=False)
    mensaje = db.Column(db.Text, nullable=False)
    tipo = db.Column(db.String(20), nullable=False, default='info')  # success, info, warning, error
    tipo_sonido = db.Column(db.String(20), nullable=False, default='alerta')  # entrada, salida, visitante, alerta
    icono = db.Column(db.String(50), nullable=False, default='fas fa-bell')
    leida = db.Column(db.Boolean, nullable=False, default=False)
    fecha_creacion = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    usuario_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    
    # Relaciones
    usuario = db.relationship('User', backref='notificaciones')

class Asistencia(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    empleado_id = db.Column(db.Integer, db.ForeignKey('empleado.id'), nullable=False)
    fecha = db.Column(db.Date, nullable=False)
    hora_entrada = db.Column(db.Time)
    hora_salida = db.Column(db.Time)
    horas_trabajadas = db.Column(db.Float)
    observaciones = db.Column(db.Text)
    token_diario = db.Column(db.String(100))  # Token del día para validación
    created_at = db.Column(db.DateTime, default=colombia_now)
    
    empleado = db.relationship('Empleado', backref=db.backref('asistencias', lazy=True))
    
    # Índice único para evitar asistencia duplicada por día
    __table_args__ = (db.UniqueConstraint('empleado_id', 'fecha', name='unique_attendance_per_day'),)

class SolicitudEmpleado(db.Model):
    """Modelo para solicitudes de empleados (vacaciones, licencias, etc.)"""
    id = db.Column(db.Integer, primary_key=True)
    empleado_id = db.Column(db.Integer, db.ForeignKey('empleado.id'), nullable=False)
    
    # Tipo de solicitud
    tipo_solicitud = db.Column(db.String(50), nullable=False)  # VACACIONES, LICENCIA_LUTO, CALAMIDAD, INCAPACIDAD, PERMISO_REMUNERADO, RETIRO_CESANTIAS
    
    # Fechas
    fecha_inicio = db.Column(db.Date, nullable=False)
    fecha_fin = db.Column(db.Date, nullable=True)  # Opcional para algunos tipos
    
    # Información de la solicitud
    motivo = db.Column(db.Text, nullable=False)
    observaciones = db.Column(db.Text)
    
    # Campos adicionales específicos por tipo (almacenados como JSON)
    datos_adicionales = db.Column(db.Text)  # JSON con campos específicos según el tipo
    
    # Estado de la solicitud
    estado = db.Column(db.String(20), default='PENDIENTE')  # PENDIENTE, APROBADA, RECHAZADA
    
    # Aprobación/Rechazo
    aprobado_por_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    fecha_aprobacion = db.Column(db.DateTime, nullable=True)
    comentario_admin = db.Column(db.Text)  # Comentario al aprobar/rechazar
    
    # Archivos adjuntos (almacenados como BYTEA)
    adjuntos_data = db.Column(db.LargeBinary, nullable=True)  # JSON con información de archivos
    adjuntos_nombres = db.Column(db.Text)  # Nombres de archivos separados por |
    
    # Documentos del admin (respuesta)
    documentos_admin_data = db.Column(db.LargeBinary, nullable=True)
    documentos_admin_nombres = db.Column(db.Text)
    
    # Campos del sistema
    created_at = db.Column(db.DateTime, default=colombia_now)
    updated_at = db.Column(db.DateTime, default=colombia_now, onupdate=colombia_now)
    
    # Relaciones
    empleado = db.relationship('Empleado', backref=db.backref('solicitudes', lazy=True))
    aprobado_por = db.relationship('User', backref=db.backref('solicitudes_aprobadas', lazy=True))

class Visitante(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    
    # Información Personal
    nombre = db.Column(db.String(100), nullable=False)
    apellido = db.Column(db.String(100), nullable=False)
    documento = db.Column(db.String(20), nullable=False)  # Cédula, pasaporte, etc.
    eps = db.Column(db.String(100), nullable=False)  # EPS del visitante
    rh = db.Column(db.String(10), nullable=False)  # Tipo de sangre
    
    # Contacto
    telefono = db.Column(db.String(20), nullable=False)
    empresa = db.Column(db.String(100))
    motivo_visita = db.Column(db.Text, nullable=False)
    
    # Control de Entrada/Salida
    fecha_entrada = db.Column(db.DateTime, default=colombia_now)
    fecha_salida = db.Column(db.DateTime)
    estado_visita = db.Column(db.String(20), default='En visita')  # En visita, Finalizada
    
    # Contacto de Emergencia
    nombre_contacto_emergencia = db.Column(db.String(200), nullable=False)
    telefono_emergencia = db.Column(db.String(20), nullable=False)
    parentesco = db.Column(db.String(50), nullable=False)
    
    # Campos del sistema
    activo = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=colombia_now)
    updated_at = db.Column(db.DateTime, default=colombia_now, onupdate=colombia_now)


def obtener_visitantes_recurrentes():
    """Obtiene la lista de visitantes registrados previamente (último registro por documento)."""
    subconsulta = db.session.query(
        db.func.max(Visitante.id).label('max_id')
    ).group_by(Visitante.documento).subquery()

    visitantes_unicos = Visitante.query.filter(
        Visitante.id.in_(subconsulta)
    ).order_by(Visitante.nombre.asc()).all()

    return visitantes_unicos

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Funciones para el sistema de QR y tokens
def generar_token_diario():
    """Genera un token estable para los QR públicos"""
    return generar_token_qr_constante()

def validar_token_diario(token):
    """Valida si el token corresponde al token público configurado"""
    token_actual = generar_token_diario()
    return token == token_actual

def generar_token_diario_visitantes():
    """Genera el mismo token estable usado en asistencia"""
    return generar_token_qr_constante()

def validar_token_diario_visitantes(token):
    """Valida si el token de visitantes corresponde al token público"""
    token_actual = generar_token_diario_visitantes()
    return token == token_actual

def generar_qr_solicitudes():
    """Genera un código QR para solicitudes de empleados"""
    token = generar_token_qr_constante()  # Mismo token estático
    url_solicitudes = f"{request.url_root}solicitudes-publico/{token}"
    
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(url_solicitudes)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Guardar en memoria
    img_buffer = io.BytesIO()
    img.save(img_buffer, format='PNG')
    img_buffer.seek(0)
    
    return img_buffer, token, url_solicitudes

def generar_qr_asistencia():
    """Genera un código QR para la asistencia del día"""
    token = generar_token_diario()
    url_asistencia = f"{request.url_root}asistencia-publica/{token}"
    
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(url_asistencia)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Guardar en memoria
    img_buffer = io.BytesIO()
    img.save(img_buffer, format='PNG')
    img_buffer.seek(0)
    
    return img_buffer, token, url_asistencia

def generar_qr_visitantes():
    """Genera un código QR para el registro de visitantes del día"""
    token = generar_token_diario_visitantes()
    url_visitantes = f"{request.url_root}visitantes-publico/{token}"
    
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(url_visitantes)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Guardar en memoria
    img_buffer = io.BytesIO()
    img.save(img_buffer, format='PNG')
    img_buffer.seek(0)
    
    return img_buffer, token, url_visitantes

# Rutas de Autenticación
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        user = User.query.filter_by(email=email).first()
        
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            return redirect(url_for('dashboard'))
        else:
            flash('Email o contraseña incorrectos', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# Dashboard Principal
@app.route('/')
@app.route('/dashboard')
@login_required
def dashboard():
    # Estadísticas principales
    try:
        total_empleados = Empleado.query.filter_by(estado_empleado='Activo').count()
        total_empleados_inactivos = Empleado.query.filter_by(estado_empleado='Inactivo').count()
    except Exception as e:
        print(f"⚠️ Error obteniendo estadísticas de empleados: {e}")
        total_empleados = 0
        total_empleados_inactivos = 0
    
    # Visitantes
    try:
        total_visitantes_hoy = Visitante.query.filter(
            Visitante.fecha_entrada >= datetime.now().date(),
            Visitante.activo == True
        ).count()
        total_visitantes_mes = Visitante.query.filter(
            Visitante.fecha_entrada >= date.today().replace(day=1),
            Visitante.activo == True
        ).count()
    except Exception as e:
        print(f"⚠️ Error obteniendo estadísticas de visitantes: {e}")
        total_visitantes_hoy = 0
        total_visitantes_mes = 0
    
    # Asistencias - Mejorado para incluir todos los meses
    try:
        asistencias_hoy = Asistencia.query.filter_by(fecha=date.today()).count()
        asistencias_semana = Asistencia.query.filter(
            Asistencia.fecha >= date.today() - timedelta(days=7)
        ).count()
        # Asistencias del mes actual
        primer_dia_mes = date.today().replace(day=1)
        asistencias_mes = Asistencia.query.filter(
            Asistencia.fecha >= primer_dia_mes
        ).count()
        # Asistencias del año actual
        primer_dia_ano = date.today().replace(month=1, day=1)
        asistencias_ano = Asistencia.query.filter(
            Asistencia.fecha >= primer_dia_ano
        ).count()
        # Total de horas trabajadas del mes
        asistencias_mes_completas = Asistencia.query.filter(
            Asistencia.fecha >= primer_dia_mes,
            Asistencia.horas_trabajadas.isnot(None)
        ).all()
        horas_trabajadas_mes = sum(a.horas_trabajadas for a in asistencias_mes_completas if a.horas_trabajadas)
    except Exception as e:
        print(f"⚠️ Error obteniendo estadísticas de asistencias: {e}")
        asistencias_hoy = 0
        asistencias_semana = 0
        asistencias_mes = 0
        asistencias_ano = 0
        horas_trabajadas_mes = 0
    
    # Contratos
    try:
        contratos_vencer = Contrato.query.filter(
            Contrato.fecha_fin <= date.today() + timedelta(days=30),
            Contrato.activo == True
        ).count()
        total_contratos_activos = Contrato.query.filter_by(activo=True).count()
    except Exception as e:
        print(f"⚠️ Error obteniendo estadísticas de contratos: {e}")
        contratos_vencer = 0
        total_contratos_activos = 0
    
    # Inventarios (si existen)
    try:
        total_productos = Producto.query.filter_by(activo=True).count()
        productos_stock_bajo = Producto.query.filter(
            Producto.stock_actual <= Producto.stock_minimo,
            Producto.activo == True
        ).count()
    except:
        total_productos = 0
        productos_stock_bajo = 0
    
    # Contratos generados
    try:
        contratos_generados_hoy = ContratoGenerado.query.filter(
            ContratoGenerado.fecha_generacion >= datetime.now().date()
        ).count()
    except Exception as e:
        print(f"⚠️ Error obteniendo contratos generados: {e}")
        contratos_generados_hoy = 0
    
    # Empleados recientes (últimos 30 días)
    try:
        empleados_recientes = Empleado.query.filter(
            Empleado.fecha_ingreso >= date.today() - timedelta(days=30)
        ).count()
    except Exception as e:
        print(f"⚠️ Error obteniendo empleados recientes: {e}")
        empleados_recientes = 0
    
    # Solicitudes pendientes
    try:
        solicitudes_pendientes = SolicitudEmpleado.query.filter_by(estado='PENDIENTE').count()
    except Exception as e:
        print(f"⚠️ Error obteniendo solicitudes pendientes: {e}")
        solicitudes_pendientes = 0
    
    return render_template('dashboard.html', 
                         total_empleados=total_empleados,
                         total_empleados_inactivos=total_empleados_inactivos,
                         total_visitantes_hoy=total_visitantes_hoy,
                         total_visitantes_mes=total_visitantes_mes,
                         asistencias_hoy=asistencias_hoy,
                         asistencias_semana=asistencias_semana,
                         asistencias_mes=asistencias_mes,
                         asistencias_ano=asistencias_ano,
                         horas_trabajadas_mes=horas_trabajadas_mes,
                         contratos_vencer=contratos_vencer,
                         total_contratos_activos=total_contratos_activos,
                         total_productos=total_productos,
                         productos_stock_bajo=productos_stock_bajo,
                         contratos_generados_hoy=contratos_generados_hoy,
                         empleados_recientes=empleados_recientes,
                         solicitudes_pendientes=solicitudes_pendientes)

# Gestión de Empleados
@app.route('/empleados')
@login_required
def empleados():
    # Mostrar todos los empleados, no solo los activos
    empleados = Empleado.query.order_by(Empleado.created_at.desc()).all()
    return render_template('empleados.html', empleados=empleados)

@app.route('/empleados/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_empleado():
    if request.method == 'POST':
        # Manejar fecha_expedicion_documento opcional
        fecha_expedicion = None
        if request.form.get('fecha_expedicion_documento'):
            fecha_expedicion = datetime.strptime(request.form['fecha_expedicion_documento'], '%Y-%m-%d').date()
        
        empleado = Empleado(
            # Información Personal
            nombre_completo=request.form['nombre_completo'],
            cedula=request.form['cedula'],
            fecha_nacimiento=datetime.strptime(request.form['fecha_nacimiento'], '%Y-%m-%d').date(),
            genero=request.form['genero'],
            estado_civil=request.form['estado_civil'],
            fecha_expedicion_documento=fecha_expedicion,
            lugar_expedicion_documento=request.form.get('lugar_expedicion_documento', ''),
            
            # Contacto
            telefono_principal=request.form['telefono_principal'],
            telefono_secundario=request.form.get('telefono_secundario', ''),
            email_personal=request.form.get('email_personal', ''),
            email_corporativo=request.form.get('email_corporativo', ''),
            
            # Dirección
            direccion_residencia=request.form['direccion_residencia'],
            ciudad=request.form['ciudad'],
            departamento=request.form['departamento'],
            codigo_postal=request.form.get('codigo_postal', ''),
            
            # Información Laboral
            cargo_puesto=request.form['cargo_puesto'],
            departamento_laboral=request.form['departamento_laboral'],
            fecha_ingreso=datetime.strptime(request.form['fecha_ingreso'], '%Y-%m-%d').date(),
            tipo_contrato=request.form['tipo_contrato'],
            salario_base=float(request.form['salario_base']),
            tipo_salario=request.form['tipo_salario'],
            jornada_laboral=request.form['jornada_laboral'],
            ubicacion_trabajo=request.form['ubicacion_trabajo'],
            estado_empleado=request.form['estado_empleado'],
            supervisor=request.form.get('supervisor', ''),
            horario=request.form.get('horario', ''),
            
            # Seguridad Social
            eps=request.form['eps'],
            arl=request.form['arl'],
            afp=request.form['afp'],
            caja_compensacion=request.form.get('caja_compensacion', ''),
            
            # Contacto de Emergencia
            nombre_contacto_emergencia=request.form['nombre_contacto_emergencia'],
            telefono_emergencia=request.form['telefono_emergencia'],
            parentesco=request.form['parentesco']
        )
        db.session.add(empleado)
        db.session.commit()
        flash('Empleado creado exitosamente', 'success')
        return redirect(url_for('empleados'))
    
    return render_template('nuevo_empleado.html')

@app.route('/empleados/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_empleado(id):
    empleado = Empleado.query.get_or_404(id)
    
    if request.method == 'POST':
        # Información Personal
        empleado.nombre_completo = request.form['nombre_completo']
        empleado.cedula = request.form['cedula']
        empleado.fecha_nacimiento = datetime.strptime(request.form['fecha_nacimiento'], '%Y-%m-%d').date()
        empleado.genero = request.form['genero']
        empleado.estado_civil = request.form['estado_civil']
        # Manejar fecha_expedicion_documento opcional
        if request.form.get('fecha_expedicion_documento'):
            empleado.fecha_expedicion_documento = datetime.strptime(request.form['fecha_expedicion_documento'], '%Y-%m-%d').date()
        else:
            empleado.fecha_expedicion_documento = None
        empleado.lugar_expedicion_documento = request.form.get('lugar_expedicion_documento', '')
        
        # Contacto
        empleado.telefono_principal = request.form['telefono_principal']
        empleado.telefono_secundario = request.form.get('telefono_secundario', '')
        empleado.email_personal = request.form.get('email_personal', '')
        empleado.email_corporativo = request.form.get('email_corporativo', '')
        
        # Dirección
        empleado.direccion_residencia = request.form['direccion_residencia']
        empleado.ciudad = request.form['ciudad']
        empleado.departamento = request.form['departamento']
        empleado.codigo_postal = request.form.get('codigo_postal', '')
        
        # Información Laboral
        empleado.cargo_puesto = request.form['cargo_puesto']
        empleado.departamento_laboral = request.form['departamento_laboral']
        empleado.fecha_ingreso = datetime.strptime(request.form['fecha_ingreso'], '%Y-%m-%d').date()
        empleado.tipo_contrato = request.form['tipo_contrato']
        empleado.salario_base = float(request.form['salario_base'])
        empleado.tipo_salario = request.form['tipo_salario']
        empleado.jornada_laboral = request.form['jornada_laboral']
        empleado.ubicacion_trabajo = request.form['ubicacion_trabajo']
        empleado.estado_empleado = request.form['estado_empleado']
        empleado.supervisor = request.form.get('supervisor', '')
        empleado.horario = request.form.get('horario', '')
        
        # Seguridad Social
        empleado.eps = request.form['eps']
        empleado.arl = request.form['arl']
        empleado.afp = request.form['afp']
        empleado.caja_compensacion = request.form.get('caja_compensacion', '')
        
        # Contacto de Emergencia
        empleado.nombre_contacto_emergencia = request.form['nombre_contacto_emergencia']
        empleado.telefono_emergencia = request.form['telefono_emergencia']
        empleado.parentesco = request.form['parentesco']
        
        db.session.commit()
        flash('Empleado actualizado exitosamente', 'success')
        return redirect(url_for('empleados'))
    
    return render_template('editar_empleado.html', empleado=empleado)

@app.route('/empleados/<int:id>')
@login_required
def ver_empleado(id):
    empleado = Empleado.query.get_or_404(id)
    return render_template('ver_empleado.html', empleado=empleado)

@app.route('/empleados/<int:id>/eliminar', methods=['POST'])
@login_required
def eliminar_empleado(id):
    empleado = Empleado.query.get_or_404(id)
    empleado.estado_empleado = 'Inactivo'
    db.session.commit()
    flash('Empleado desactivado exitosamente', 'success')
    return redirect(url_for('empleados'))

# Gestión de Solicitudes de Empleados
@app.route('/solicitudes')
@login_required
def solicitudes():
    """Lista todas las solicitudes"""
    estado_filtro = request.args.get('estado', 'TODAS')
    tipo_filtro = request.args.get('tipo', 'TODAS')
    
    query = SolicitudEmpleado.query
    
    if estado_filtro != 'TODAS':
        query = query.filter_by(estado=estado_filtro)
    
    if tipo_filtro != 'TODAS':
        query = query.filter_by(tipo_solicitud=tipo_filtro)
    
    solicitudes = query.order_by(SolicitudEmpleado.created_at.desc()).all()
    
    # Generar QR para solicitudes
    qr_buffer_solicitudes, token_solicitudes, url_qr_solicitudes = generar_qr_solicitudes()
    
    return render_template('solicitudes.html', 
                         solicitudes=solicitudes,
                         estado_filtro=estado_filtro,
                         tipo_filtro=tipo_filtro,
                         url_qr_solicitudes=url_qr_solicitudes)

@app.route('/solicitudes/<int:id>')
@login_required
def ver_solicitud(id):
    """Ver detalles de una solicitud"""
    solicitud = SolicitudEmpleado.query.get_or_404(id)
    
    # Procesar adjuntos si existen
    adjuntos = []
    if solicitud.adjuntos_data:
        try:
            import json
            adjuntos_json = json.loads(solicitud.adjuntos_data.decode())
            adjuntos = [{'nombre': a['nombre']} for a in adjuntos_json]
        except:
            pass
    
    # Procesar documentos del admin si existen
    documentos_admin = []
    if solicitud.documentos_admin_data:
        try:
            import json
            docs_json = json.loads(solicitud.documentos_admin_data.decode())
            documentos_admin = [{'nombre': d['nombre']} for d in docs_json]
        except:
            pass
    
    return render_template('ver_solicitud.html', 
                         solicitud=solicitud,
                         adjuntos=adjuntos,
                         documentos_admin=documentos_admin)

@app.route('/solicitudes/<int:id>/aprobar', methods=['POST'])
@login_required
def aprobar_solicitud(id):
    """Aprobar una solicitud"""
    solicitud = SolicitudEmpleado.query.get_or_404(id)
    
    if solicitud.estado != 'PENDIENTE':
        flash('Esta solicitud ya fue procesada', 'warning')
        return redirect(url_for('ver_solicitud', id=id))
    
    comentario = request.form.get('comentario', '').strip()
    
    # Procesar documentos del admin si se subieron
    documentos_admin_data = None
    documentos_admin_nombres = []
    if 'documentos_admin' in request.files:
        archivos = request.files.getlist('documentos_admin')
        archivos_data_list = []
        for archivo in archivos:
            if archivo and archivo.filename:
                if not archivo.filename.lower().endswith(('.pdf', '.jpg', '.jpeg', '.png', '.doc', '.docx')):
                    flash('Solo se permiten archivos PDF, imágenes o documentos Word', 'error')
                    return redirect(url_for('ver_solicitud', id=id))
                
                archivo.seek(0, os.SEEK_END)
                tamaño = archivo.tell()
                archivo.seek(0)
                if tamaño > 5 * 1024 * 1024:
                    flash('Cada archivo debe ser menor a 5MB', 'error')
                    return redirect(url_for('ver_solicitud', id=id))
                
                archivos_data_list.append({
                    'nombre': archivo.filename,
                    'data': archivo.read()
                })
                documentos_admin_nombres.append(archivo.filename)
        
        if archivos_data_list:
            import json
            documentos_admin_data = json.dumps([{'nombre': a['nombre'], 'data': a['data'].hex()} for a in archivos_data_list]).encode()
    
    solicitud.estado = 'APROBADA'
    solicitud.aprobado_por_id = current_user.id
    solicitud.fecha_aprobacion = colombia_now()
    solicitud.comentario_admin = comentario or None
    if documentos_admin_data:
        solicitud.documentos_admin_data = documentos_admin_data
        solicitud.documentos_admin_nombres = '|'.join(documentos_admin_nombres)
    
    try:
        db.session.commit()
        flash('Solicitud aprobada exitosamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Error al aprobar la solicitud', 'error')
    
    return redirect(url_for('ver_solicitud', id=id))

@app.route('/solicitudes/<int:id>/rechazar', methods=['POST'])
@login_required
def rechazar_solicitud(id):
    """Rechazar una solicitud"""
    solicitud = SolicitudEmpleado.query.get_or_404(id)
    
    if solicitud.estado != 'PENDIENTE':
        flash('Esta solicitud ya fue procesada', 'warning')
        return redirect(url_for('ver_solicitud', id=id))
    
    comentario = request.form.get('comentario', '').strip()
    
    if not comentario:
        flash('Debe especificar el motivo del rechazo', 'error')
        return redirect(url_for('ver_solicitud', id=id))
    
    solicitud.estado = 'RECHAZADA'
    solicitud.aprobado_por_id = current_user.id
    solicitud.fecha_aprobacion = colombia_now()
    solicitud.comentario_admin = comentario
    
    try:
        db.session.commit()
        flash('Solicitud rechazada', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Error al rechazar la solicitud', 'error')
    
    return redirect(url_for('ver_solicitud', id=id))

@app.route('/solicitudes/<int:id>/adjunto/<int:adjunto_idx>')
@login_required
def descargar_adjunto_solicitud(id, adjunto_idx):
    """Descargar un archivo adjunto de una solicitud"""
    solicitud = SolicitudEmpleado.query.get_or_404(id)
    
    if not solicitud.adjuntos_data:
        flash('No hay archivos adjuntos', 'error')
        return redirect(url_for('ver_solicitud', id=id))
    
    try:
        import json
        adjuntos_json = json.loads(solicitud.adjuntos_data.decode())
        if adjunto_idx >= len(adjuntos_json):
            flash('Archivo no encontrado', 'error')
            return redirect(url_for('ver_solicitud', id=id))
        
        adjunto = adjuntos_json[adjunto_idx]
        archivo_data = bytes.fromhex(adjunto['data'])
        
        return send_file(
            io.BytesIO(archivo_data),
            mimetype='application/octet-stream',
            as_attachment=True,
            download_name=adjunto['nombre']
        )
    except Exception as e:
        flash('Error al descargar el archivo', 'error')
        return redirect(url_for('ver_solicitud', id=id))

@app.route('/solicitudes/<int:id>/documento-admin/<int:doc_idx>')
@login_required
def descargar_documento_admin(id, doc_idx):
    """Descargar un documento del admin"""
    solicitud = SolicitudEmpleado.query.get_or_404(id)
    
    if not solicitud.documentos_admin_data:
        flash('No hay documentos', 'error')
        return redirect(url_for('ver_solicitud', id=id))
    
    try:
        import json
        docs_json = json.loads(solicitud.documentos_admin_data.decode())
        if doc_idx >= len(docs_json):
            flash('Documento no encontrado', 'error')
            return redirect(url_for('ver_solicitud', id=id))
        
        doc = docs_json[doc_idx]
        archivo_data = bytes.fromhex(doc['data'])
        
        return send_file(
            io.BytesIO(archivo_data),
            mimetype='application/octet-stream',
            as_attachment=True,
            download_name=doc['nombre']
        )
    except Exception as e:
        flash('Error al descargar el documento', 'error')
        return redirect(url_for('ver_solicitud', id=id))

# Gestión de Contratos - Rutas movidas a la sección completa más abajo

# Gestión de Asistencia
@app.route('/asistencia')
@login_required
def asistencia():
    fecha = request.args.get('fecha', date.today().strftime('%Y-%m-%d'))
    fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
    
    asistencias = Asistencia.query.filter_by(fecha=fecha_obj).all()
    empleados = Empleado.query.filter_by(estado_empleado='Activo').all()
    
    # Generar QR para el día actual
    qr_buffer, token, url_qr = generar_qr_asistencia()
    
    # Generar QR para solicitudes
    qr_buffer_solicitudes, token_solicitudes, url_qr_solicitudes = generar_qr_solicitudes()
    
    return render_template('asistencia.html', 
                         asistencias=asistencias, 
                         empleados=empleados, 
                         fecha=fecha,
                         token_diario=token,
                         url_qr=url_qr,
                         url_qr_solicitudes=url_qr_solicitudes)

@app.route('/asistencia/qr')
@login_required
def generar_qr_imagen():
    """Genera y devuelve la imagen del código QR"""
    qr_buffer, token, url_qr = generar_qr_asistencia()
    return send_file(qr_buffer, mimetype='image/png')

@app.route('/visitantes/qr')
@login_required
def generar_qr_visitantes_imagen():
    """Genera y devuelve la imagen del código QR para visitantes"""
    qr_buffer, token, url_qr = generar_qr_visitantes()
    return send_file(qr_buffer, mimetype='image/png')

@app.route('/solicitudes/qr')
@login_required
def generar_qr_solicitudes_imagen():
    """Genera y devuelve la imagen del código QR para solicitudes"""
    qr_buffer, token, url_qr = generar_qr_solicitudes()
    return send_file(qr_buffer, mimetype='image/png')

# Ruta pública para asistencia (sin login requerido)
@app.route('/asistencia-publica/<token>', methods=['GET', 'POST'])
def asistencia_publica(token):
    """Página pública para que los empleados marquen asistencia"""
    # Validar que el token sea del día actual
    if not validar_token_diario(token):
        flash('El código QR ha expirado. Solicite un nuevo código al administrador.', 'error')
        return render_template('asistencia_publica.html', token=token, error=True)
    
    if request.method == 'POST':
        documento = request.form.get('documento', '').strip()
        nombre = request.form.get('nombre', '').strip()
        tipo_registro = request.form.get('tipo_registro', '').strip()
        
        if not documento or not nombre or not tipo_registro:
            flash('Por favor complete todos los campos y seleccione el tipo de registro', 'error')
            return redirect(url_for('asistencia_publica', token=token))
        
        # Buscar empleado por documento (exacto)
        empleado = Empleado.query.filter_by(cedula=documento).first()
        
        # Si no se encuentra por documento exacto, buscar por nombre (ignorando acentos)
        if not empleado:
            # Normalizar nombre para búsqueda (quitar acentos)
            import unicodedata
            nombre_normalizado = unicodedata.normalize('NFD', nombre).encode('ascii', 'ignore').decode('ascii').lower()
            
            # Buscar empleados que coincidan con el nombre normalizado
            empleados = Empleado.query.filter_by(estado_empleado='Activo').all()
            for emp in empleados:
                nombre_emp_normalizado = unicodedata.normalize('NFD', emp.nombre_completo).encode('ascii', 'ignore').decode('ascii').lower()
                if nombre_normalizado in nombre_emp_normalizado or nombre_emp_normalizado in nombre_normalizado:
                    empleado = emp
                    break
        
        if not empleado:
            flash('No se encontró un empleado con ese documento o nombre. Verifique los datos ingresados.', 'error')
            return redirect(url_for('asistencia_publica', token=token))
        
        # Verificar que el nombre coincida (validación más flexible)
        nombre_empleado = empleado.nombre_completo.lower().strip()
        nombre_ingresado = nombre.lower().strip()
        
        # Permitir coincidencias parciales y diferentes formatos
        if not (nombre_empleado == nombre_ingresado or 
                nombre_empleado in nombre_ingresado or 
                nombre_ingresado in nombre_empleado):
            flash(f'El nombre ingresado no coincide con el empleado registrado. Empleado: {empleado.nombre_completo}', 'error')
            return redirect(url_for('asistencia_publica', token=token))
        
        # Verificar que el empleado esté activo
        if empleado.estado_empleado != 'Activo':
            flash('El empleado no está activo en el sistema', 'error')
            return redirect(url_for('asistencia_publica', token=token))
        
        fecha_hoy = date.today()
        hora_actual = colombia_now().time()
        
        # Buscar asistencia existente para hoy
        asistencia_existente = Asistencia.query.filter_by(
            empleado_id=empleado.id, 
            fecha=fecha_hoy
        ).first()
        
        if tipo_registro == 'entrada':
            if asistencia_existente:
                flash(f'Ya se registró entrada para {empleado.nombre_completo} hoy a las {asistencia_existente.hora_entrada.strftime("%H:%M")}', 'warning')
                return redirect(url_for('asistencia_publica', token=token))
            
            # Registrar entrada
            asistencia = Asistencia(
                empleado_id=empleado.id,
                fecha=fecha_hoy,
                hora_entrada=hora_actual,
                token_diario=token
            )
            
            try:
                db.session.add(asistencia)
                db.session.commit()
                
                # Enviar notificación
                notificar_asistencia_entrada(
                    empleado.nombre_completo, 
                    colombia_now().strftime("%H:%M")
                )
                
                flash(f'Entrada registrada exitosamente para {empleado.nombre_completo} a las {colombia_now().strftime("%H:%M")}', 'success')
                # Usar redirect para evitar reenvío al recargar (patrón PRG)
                return redirect(url_for('asistencia_publica', token=token))
            except Exception as e:
                db.session.rollback()
                flash('Error al registrar la entrada. Intente nuevamente.', 'error')
        
        elif tipo_registro == 'salida':
            if not asistencia_existente:
                flash(f'No se encontró registro de entrada para {empleado.nombre_completo} hoy. Debe registrar entrada primero.', 'error')
                return redirect(url_for('asistencia_publica', token=token))
            
            if asistencia_existente.hora_salida:
                flash(f'Ya se registró salida para {empleado.nombre_completo} hoy a las {asistencia_existente.hora_salida.strftime("%H:%M")}', 'warning')
                return redirect(url_for('asistencia_publica', token=token))
            
            # Registrar salida
            asistencia_existente.hora_salida = hora_actual
            asistencia_existente.token_diario = token
            
            try:
                db.session.commit()
                
                # Enviar notificación
                notificar_asistencia_salida(
                    empleado.nombre_completo, 
                    hora_actual.strftime("%H:%M")
                )
                
                flash(f'Salida registrada exitosamente para {empleado.nombre_completo} a las {colombia_now().strftime("%H:%M")}', 'success')
                # Usar redirect para evitar reenvío al recargar (patrón PRG)
                return redirect(url_for('asistencia_publica', token=token))
            except Exception as e:
                db.session.rollback()
                flash('Error al registrar la salida. Intente nuevamente.', 'error')
        
        # Si hay errores, también hacer redirect
        return redirect(url_for('asistencia_publica', token=token))
    
    return render_template('asistencia_publica.html', token=token)

# Ruta pública para visitantes (sin login requerido)
@app.route('/visitantes-publico/<token>', methods=['GET', 'POST'])
def visitantes_publico(token):
    """Página pública para que los visitantes se registren"""
    modo_activo = 'nuevo'
    visitantes_recurrentes = obtener_visitantes_recurrentes()

    # Validar que el token sea del día actual
    if not validar_token_diario_visitantes(token):
        flash('El código QR ha expirado. Solicite un nuevo código al administrador.', 'error')
        return render_template(
            'visitantes_publico.html',
            token=token,
            error=True,
            visitantes_recurrentes=visitantes_recurrentes,
            modo_activo=modo_activo
        )
    
    if request.method == 'POST':
        modo_registro = request.form.get('modo_registro', 'nuevo')
        modo_activo = modo_registro
        
        if modo_registro == 'recurrente':
            visitante_recurrente_id = request.form.get('visitante_recurrente_id')
            documento_verificacion = request.form.get('documento_verificacion', '').strip()

            if not visitante_recurrente_id or not documento_verificacion:
                flash('Seleccione su nombre y escriba su documento para continuar.', 'error')
                return redirect(url_for('visitantes_publico', token=token))

            visitante_referencia = Visitante.query.get(int(visitante_recurrente_id))

            if not visitante_referencia:
                flash('No encontramos el visitante seleccionado. Intente nuevamente.', 'error')
                return redirect(url_for('visitantes_publico', token=token))

            if visitante_referencia.documento.strip() != documento_verificacion:
                flash('El documento ingresado no coincide con el registrado anteriormente.', 'error')
                return redirect(url_for('visitantes_publico', token=token))

            fecha_hoy = date.today()
            visitante_existente = Visitante.query.filter(
                Visitante.documento == documento_verificacion,
                db.func.date(Visitante.fecha_entrada) == fecha_hoy,
                Visitante.estado_visita == 'En visita'
            ).first()

            if visitante_existente:
                flash('Ya existe un registro activo para este documento el día de hoy.', 'warning')
                return redirect(url_for('visitantes_publico', token=token))

            visitante = Visitante(
                nombre=visitante_referencia.nombre,
                apellido=visitante_referencia.apellido,
                documento=visitante_referencia.documento,
                eps=visitante_referencia.eps,
                rh=visitante_referencia.rh,
                telefono=visitante_referencia.telefono,
                empresa=visitante_referencia.empresa,
                motivo_visita=visitante_referencia.motivo_visita,
                fecha_entrada=colombia_now(),
                estado_visita='En visita',
                nombre_contacto_emergencia=visitante_referencia.nombre_contacto_emergencia,
                telefono_emergencia=visitante_referencia.telefono_emergencia,
                parentesco=visitante_referencia.parentesco,
                activo=True
            )

            try:
                db.session.add(visitante)
                db.session.commit()

                notificar_visitante_nuevo(
                    f"{visitante.nombre} {visitante.apellido}",
                    visitante.empresa or "Sin empresa"
                )

                flash(f'¡Bienvenido nuevamente {visitante.nombre}! Tu entrada rápida quedó registrada a las {colombia_now().strftime("%H:%M")}', 'success')
                # Usar redirect para evitar reenvío al recargar (patrón PRG)
                return redirect(url_for('visitantes_publico', token=token))
            except Exception:
                db.session.rollback()
                flash('Error al registrar la entrada rápida. Intente nuevamente.', 'error')
                return redirect(url_for('visitantes_publico', token=token))

        nombre = request.form.get('nombre', '').strip()
        apellido = request.form.get('apellido', '').strip()
        documento = request.form.get('documento', '').strip()
        eps = request.form.get('eps', '').strip()
        rh = request.form.get('rh', '').strip()
        telefono = request.form.get('telefono', '').strip()
        empresa = request.form.get('empresa', '').strip()
        motivo_visita = request.form.get('motivo_visita', '').strip()
        nombre_contacto_emergencia = request.form.get('nombre_contacto_emergencia', '').strip()
        telefono_emergencia = request.form.get('telefono_emergencia', '').strip()
        parentesco = request.form.get('parentesco', '').strip()
        
        # Validar campos requeridos
        campos_requeridos = {
            'nombre': nombre,
            'apellido': apellido,
            'documento': documento,
            'eps': eps,
            'rh': rh,
            'telefono': telefono,
            'empresa': empresa,
            'motivo_visita': motivo_visita,
            'nombre_contacto_emergencia': nombre_contacto_emergencia,
            'telefono_emergencia': telefono_emergencia,
            'parentesco': parentesco
        }
        
        campos_faltantes = [campo for campo, valor in campos_requeridos.items() if not valor]
        if campos_faltantes:
            flash('Por favor complete todos los campos requeridos', 'error')
            return redirect(url_for('visitantes_publico', token=token))
        
        # Verificar si ya existe un visitante con el mismo documento hoy
        fecha_hoy = date.today()
        visitante_existente = Visitante.query.filter(
            Visitante.documento == documento,
            db.func.date(Visitante.fecha_entrada) == fecha_hoy
        ).first()
        
        if visitante_existente:
            flash(f'Ya existe un registro de visitante con documento {documento} para hoy', 'warning')
            return redirect(url_for('visitantes_publico', token=token))
        
        # Crear nuevo visitante
        visitante = Visitante(
            nombre=nombre,
            apellido=apellido,
            documento=documento,
            eps=eps,
            rh=rh,
            telefono=telefono,
            empresa=empresa,
            motivo_visita=motivo_visita,
            fecha_entrada=colombia_now(),
            estado_visita='En visita',
            nombre_contacto_emergencia=nombre_contacto_emergencia,
            telefono_emergencia=telefono_emergencia,
            parentesco=parentesco,
            activo=True
        )
        
        try:
            db.session.add(visitante)
            db.session.commit()
            
            # Enviar notificación
            notificar_visitante_nuevo(
                f"{nombre} {apellido}",
                empresa or "Sin empresa"
            )
            
            flash(f'Visitante {nombre} {apellido} registrado exitosamente a las {colombia_now().strftime("%H:%M")}', 'success')
            # Usar redirect para evitar reenvío al recargar (patrón PRG)
            return redirect(url_for('visitantes_publico', token=token))
        except Exception as e:
            db.session.rollback()
            flash('Error al registrar el visitante. Intente nuevamente.', 'error')
            return redirect(url_for('visitantes_publico', token=token))
    
    return render_template(
        'visitantes_publico.html',
        token=token,
        visitantes_recurrentes=visitantes_recurrentes,
        modo_activo=modo_activo
    )

# Ruta pública para solicitudes de empleados (sin login requerido)
@app.route('/solicitudes-publico/<token>', methods=['GET', 'POST'])
def solicitudes_publico(token):
    """Página pública para que los empleados realicen solicitudes"""
    # Validar token (mismo token estático)
    if not validar_token_diario(token):
        flash('El código QR no es válido. Solicite un nuevo código al administrador.', 'error')
        return render_template('solicitudes_publico.html', token=token, error=True)
    
    if request.method == 'POST':
        # Verificar que no se esté reenviando el formulario
        if 'form_submitted' in session and session.get('form_submitted') == True:
            flash('La solicitud ya fue enviada. Por favor, no recargue la página.', 'warning')
            return redirect(url_for('solicitudes_publico', token=token))
        
        documento = request.form.get('documento', '').strip()
        nombre = request.form.get('nombre', '').strip()
        tipo_solicitud = request.form.get('tipo_solicitud', '').strip()
        motivo = request.form.get('motivo', '').strip()
        observaciones = request.form.get('observaciones', '').strip()
        
        # Obtener campos según el tipo de solicitud
        datos_adicionales = {}
        fecha_inicio = None
        fecha_fin = None
        
        if tipo_solicitud == 'LICENCIA_LUTO':
            fecha_inicio = request.form.get('fecha_inicio', '').strip()
            datos_adicionales = {
                'cantidad_dias_semestral': request.form.get('cantidad_dias_semestral', '').strip(),
                'ano': request.form.get('ano', '').strip(),
                'periodo': request.form.get('periodo', '').strip(),
                'cantidad_dias_disponibles': request.form.get('cantidad_dias_disponibles', '').strip()
            }
            if not all([documento, nombre, tipo_solicitud, fecha_inicio] + list(datos_adicionales.values())):
                flash('Por favor complete todos los campos requeridos', 'error')
                return redirect(url_for('solicitudes_publico', token=token))
                
        elif tipo_solicitud == 'INCAPACIDAD':
            fecha_inicio = request.form.get('fecha_inicio_eps', '').strip()
            datos_adicionales = {
                'numero_incapacidad': request.form.get('numero_incapacidad', '').strip(),
                'cantidad_dias': request.form.get('cantidad_dias_incapacidad', '').strip(),
                'fecha_inicio_eps': fecha_inicio,
                'observaciones': request.form.get('observaciones_incapacidad', '').strip()
            }
            if not all([documento, nombre, tipo_solicitud, fecha_inicio, datos_adicionales['numero_incapacidad'], datos_adicionales['cantidad_dias']]):
                flash('Por favor complete todos los campos requeridos', 'error')
                return redirect(url_for('solicitudes_publico', token=token))
                
        elif tipo_solicitud == 'CALAMIDAD':
            fecha_inicio = request.form.get('fecha_inicio_calamidad', '').strip()
            cantidad_dias = request.form.get('cantidad_dias_calamidad', '').strip()
            datos_adicionales = {
                'cantidad_dias': cantidad_dias
            }
            if not all([documento, nombre, tipo_solicitud, fecha_inicio, cantidad_dias]):
                flash('Por favor complete todos los campos requeridos', 'error')
                return redirect(url_for('solicitudes_publico', token=token))
                
        elif tipo_solicitud == 'PERMISO_REMUNERADO':
            fecha_inicio = request.form.get('fecha_inicio_permiso', '').strip()
            motivo_permiso = request.form.get('motivo_permiso', '').strip()
            datos_adicionales = {
                'numero_horas': request.form.get('numero_horas', '').strip(),
                'motivo_permiso': motivo_permiso,
                'observaciones': request.form.get('observaciones_permiso', '').strip()
            }
            motivo = motivo_permiso  # Usar motivo_permiso como motivo principal
            if not all([documento, nombre, tipo_solicitud, fecha_inicio, motivo_permiso, datos_adicionales['numero_horas']]):
                flash('Por favor complete todos los campos requeridos', 'error')
                return redirect(url_for('solicitudes_publico', token=token))
                
        elif tipo_solicitud == 'VACACIONES':
            fecha_inicio = request.form.get('fecha_inicio_vacaciones', '').strip()
            cantidad_dias = request.form.get('cantidad_dias_vacaciones', '').strip()
            fecha_reintegro = request.form.get('fecha_reintegro', '').strip()
            datos_adicionales = {
                'cantidad_dias': cantidad_dias,
                'fecha_reintegro': fecha_reintegro
            }
            fecha_fin = fecha_reintegro  # Usar fecha de reintegro como fecha_fin
            if not all([documento, nombre, tipo_solicitud, fecha_inicio, cantidad_dias, fecha_reintegro]):
                flash('Por favor complete todos los campos requeridos', 'error')
                return redirect(url_for('solicitudes_publico', token=token))
                
        elif tipo_solicitud == 'RETIRO_CESANTIAS':
            motivo_cesantias = request.form.get('motivo_cesantias', '').strip()
            datos_adicionales = {
                'motivo_cesantias': motivo_cesantias,
                'observaciones': request.form.get('observaciones_cesantias', '').strip()
            }
            motivo = motivo_cesantias  # Usar motivo_cesantias como motivo principal
            fecha_inicio = date.today().isoformat()  # Fecha actual
            if not all([documento, nombre, tipo_solicitud, motivo_cesantias]):
                flash('Por favor complete todos los campos requeridos', 'error')
                return redirect(url_for('solicitudes_publico', token=token))
        else:
            # Tipo genérico (por si acaso)
            fecha_inicio = request.form.get('fecha_inicio', '').strip()
            fecha_fin = request.form.get('fecha_fin', '').strip()
            if not all([documento, nombre, tipo_solicitud, fecha_inicio, fecha_fin, motivo]):
                flash('Por favor complete todos los campos requeridos', 'error')
                return redirect(url_for('solicitudes_publico', token=token))
        
        # Buscar empleado
        empleado = Empleado.query.filter_by(cedula=documento).first()
        if not empleado:
            flash('No se encontró un empleado con ese documento. Verifique los datos.', 'error')
            return redirect(url_for('solicitudes_publico', token=token))
        
        # Validar nombre
        if empleado.nombre_completo.lower().strip() != nombre.lower().strip():
            flash(f'El nombre no coincide. Empleado registrado: {empleado.nombre_completo}', 'error')
            return redirect(url_for('solicitudes_publico', token=token))
        
        # Validar que el empleado esté activo
        if empleado.estado_empleado != 'Activo':
            flash('El empleado no está activo en el sistema', 'error')
            return redirect(url_for('solicitudes_publico', token=token))
        
        # Validar fechas
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            if fecha_fin:
                fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                if fecha_fin_obj < fecha_inicio_obj:
                    flash('La fecha de fin debe ser posterior a la fecha de inicio', 'error')
                    return redirect(url_for('solicitudes_publico', token=token))
            else:
                fecha_fin_obj = None
        except ValueError:
            flash('Formato de fecha inválido', 'error')
            return redirect(url_for('solicitudes_publico', token=token))
        
        # Procesar archivos adjuntos
        adjuntos_data = None
        adjuntos_nombres = []
        if 'adjuntos' in request.files:
            archivos = request.files.getlist('adjuntos')
            archivos_data_list = []
            for archivo in archivos:
                if archivo and archivo.filename:
                    # Validar tipo de archivo
                    if not archivo.filename.lower().endswith(('.pdf', '.jpg', '.jpeg', '.png', '.doc', '.docx')):
                        flash('Solo se permiten archivos PDF, imágenes o documentos Word', 'error')
                        return redirect(url_for('solicitudes_publico', token=token))
                    
                    # Validar tamaño (máximo 5MB por archivo)
                    archivo.seek(0, os.SEEK_END)
                    tamaño = archivo.tell()
                    archivo.seek(0)
                    if tamaño > 5 * 1024 * 1024:  # 5MB
                        flash('Cada archivo debe ser menor a 5MB', 'error')
                        return redirect(url_for('solicitudes_publico', token=token))
                    
                    archivos_data_list.append({
                        'nombre': archivo.filename,
                        'data': archivo.read()
                    })
                    adjuntos_nombres.append(archivo.filename)
            
            if archivos_data_list:
                import json
                adjuntos_data = json.dumps([{'nombre': a['nombre'], 'data': a['data'].hex()} for a in archivos_data_list]).encode()
        
        # Serializar datos adicionales como JSON
        import json
        datos_adicionales_json = json.dumps(datos_adicionales) if datos_adicionales else None
        
        # Crear solicitud
        solicitud = SolicitudEmpleado(
            empleado_id=empleado.id,
            tipo_solicitud=tipo_solicitud,
            fecha_inicio=fecha_inicio_obj,
            fecha_fin=fecha_fin_obj,
            motivo=motivo,
            observaciones=observaciones or None,
            datos_adicionales=datos_adicionales_json,
            estado='PENDIENTE',
            adjuntos_data=adjuntos_data,
            adjuntos_nombres='|'.join(adjuntos_nombres) if adjuntos_nombres else None
        )
        
        try:
            db.session.add(solicitud)
            db.session.commit()
            
            # Notificar al admin
            from notificaciones import notificacion_manager
            notificacion_manager.agregar_notificacion(
                titulo="Nueva Solicitud de Empleado",
                mensaje=f"{empleado.nombre_completo} ha enviado una solicitud de {tipo_solicitud.replace('_', ' ').title()}",
                tipo='warning',
                tipo_sonido='alerta',
                icono='fas fa-file-alt'
            )
            
            # Marcar formulario como enviado
            session['form_submitted'] = True
            
            flash('Solicitud enviada exitosamente. Será revisada por el administrador.', 'success')
            return redirect(url_for('solicitudes_publico', token=token))
        except Exception as e:
            db.session.rollback()
            print(f"Error al crear solicitud: {e}")
            flash('Error al enviar la solicitud. Intente nuevamente.', 'error')
            return redirect(url_for('solicitudes_publico', token=token))
    
    # Limpiar flag de formulario enviado al cargar la página
    session.pop('form_submitted', None)
    
    return render_template('solicitudes_publico.html', token=token)

@app.route('/asistencia/registrar', methods=['POST'])
@login_required
def registrar_asistencia():
    empleado_id = int(request.form['empleado_id'])  # Convertir a entero
    fecha = datetime.strptime(request.form['fecha'], '%Y-%m-%d').date()
    tipo_registro = request.form.get('tipo_registro', '').strip()
    observaciones = request.form.get('observaciones', '')
    
    if not tipo_registro:
        flash('Por favor seleccione el tipo de registro (entrada o salida)', 'error')
        return redirect(url_for('asistencia'))
    
    # Buscar empleado
    empleado = Empleado.query.get(empleado_id)
    if not empleado:
        flash('Empleado no encontrado', 'error')
        return redirect(url_for('asistencia'))
    
    # Buscar asistencia existente para hoy
    asistencia_existente = Asistencia.query.filter_by(
        empleado_id=empleado_id, 
        fecha=fecha
    ).first()
    
    if tipo_registro == 'entrada':
        if asistencia_existente:
            flash(f'Ya se registró entrada para {empleado.nombre_completo} hoy a las {asistencia_existente.hora_entrada.strftime("%H:%M")}', 'warning')
            return redirect(url_for('asistencia'))
        
        # Registrar entrada
        asistencia = Asistencia(
            empleado_id=empleado_id,
            fecha=fecha,
            hora_entrada=colombia_now().time(),
            observaciones=observaciones,
            token_diario='Manual'  # Marcar como registro manual
        )
        
        try:
            db.session.add(asistencia)
            db.session.commit()
            
            # Enviar notificación
            notificar_asistencia_entrada(
                empleado.nombre_completo, 
                colombia_now().strftime("%H:%M")
            )
            
            flash(f'Entrada registrada exitosamente para {empleado.nombre_completo} a las {colombia_now().strftime("%H:%M")}', 'success')
        except Exception as e:
            db.session.rollback()
            flash('Error al registrar la entrada. Intente nuevamente.', 'error')
    
    elif tipo_registro == 'salida':
        if not asistencia_existente:
            flash(f'No se encontró registro de entrada para {empleado.nombre_completo} hoy. Debe registrar entrada primero.', 'error')
            return redirect(url_for('asistencia'))
        
        if asistencia_existente.hora_salida:
            flash(f'Ya se registró salida para {empleado.nombre_completo} hoy a las {asistencia_existente.hora_salida.strftime("%H:%M")}', 'warning')
            return redirect(url_for('asistencia'))
        
        # Registrar salida
        hora_salida = colombia_now().time()
        asistencia_existente.hora_salida = hora_salida
        
        # Calcular horas trabajadas
        entrada = datetime.combine(fecha, asistencia_existente.hora_entrada)
        salida = datetime.combine(fecha, hora_salida)
        asistencia_existente.horas_trabajadas = (salida - entrada).total_seconds() / 3600
        
        # Actualizar observaciones si se proporcionaron
        if observaciones:
            asistencia_existente.observaciones = observaciones
        
        try:
            db.session.commit()
            
            # Enviar notificación
            notificar_asistencia_salida(
                empleado.nombre_completo, 
                hora_salida.strftime("%H:%M")
            )
            
            flash(f'Salida registrada exitosamente para {empleado.nombre_completo} a las {colombia_now().strftime("%H:%M")}', 'success')
        except Exception as e:
            db.session.rollback()
            flash('Error al registrar la salida. Intente nuevamente.', 'error')
    
    return redirect(url_for('asistencia'))

@app.route('/asistencia/eliminar/<int:id>', methods=['DELETE'])
@login_required
def eliminar_asistencia(id):
    """Eliminar una asistencia (completa o incompleta)"""
    try:
        asistencia = Asistencia.query.get_or_404(id)
        
        # Permitir eliminar cualquier asistencia (completa o incompleta)
        db.session.delete(asistencia)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Asistencia eliminada exitosamente'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error al eliminar la asistencia: {str(e)}'
        }), 500

@app.route('/asistencia/editar/<int:id>', methods=['PUT'])
@login_required
def editar_asistencia(id):
    """Editar una asistencia (solo hora de entrada y observaciones)"""
    try:
        asistencia = Asistencia.query.get_or_404(id)
        data = request.get_json()
        
        # Solo permitir editar si no tiene salida registrada
        if asistencia.hora_salida:
            return jsonify({
                'success': False,
                'message': 'No se puede editar una asistencia que ya tiene salida registrada'
            }), 400
        
        # Actualizar hora de entrada si se proporciona
        if 'hora_entrada' in data and data['hora_entrada']:
            try:
                hora_entrada = datetime.strptime(data['hora_entrada'], '%H:%M').time()
                asistencia.hora_entrada = hora_entrada
            except ValueError:
                return jsonify({
                    'success': False,
                    'message': 'Formato de hora inválido. Use HH:MM'
                }), 400
        
        # Actualizar observaciones
        if 'observaciones' in data:
            asistencia.observaciones = data['observaciones']
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Asistencia actualizada exitosamente'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error al editar la asistencia: {str(e)}'
        }), 500

# Gestión de Visitantes
@app.route('/visitantes')
@login_required
def visitantes():
    visitantes = Visitante.query.order_by(Visitante.created_at.desc()).all()
    
    # Generar QR para visitantes
    qr_buffer, token, url_qr = generar_qr_visitantes()
    
    return render_template('visitantes.html', 
                         visitantes=visitantes,
                         token_diario=token,
                         url_qr=url_qr)

@app.route('/visitantes/detalles/<int:id>')
@login_required
def detalles_visitante(id):
    visitante = Visitante.query.get_or_404(id)
    return render_template('detalles_visitante.html', visitante=visitante)

@app.route('/asistencia/detalles/<int:id>')
@login_required
def detalles_asistencia(id):
    asistencia = Asistencia.query.get_or_404(id)
    return render_template('detalles_asistencia.html', asistencia=asistencia)

# ===== RUTAS PARA CONTRATOS =====

@app.route('/contratos')
@login_required
def contratos():
    """Lista todos los contratos con opciones CRUD"""
    contratos = Contrato.query.join(Empleado).order_by(Contrato.created_at.desc()).all()
    return render_template('contratos.html', contratos=contratos)

@app.route('/contratos/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_contrato():
    """Crear nuevo contrato"""
    if request.method == 'POST':
        contrato = Contrato(
            empleado_id=request.form['empleado_id'],
            tipo_contrato=request.form['tipo_contrato'],
            fecha_inicio=datetime.strptime(request.form['fecha_inicio'], '%Y-%m-%d').date(),
            fecha_fin=datetime.strptime(request.form['fecha_fin'], '%Y-%m-%d').date() if request.form.get('fecha_fin') else None,
            salario=float(request.form['salario']),
            descripcion=request.form.get('descripcion', '')
        )
        db.session.add(contrato)
        db.session.commit()
        flash('Contrato creado exitosamente', 'success')
        return redirect(url_for('contratos'))
    
    empleados = Empleado.query.filter_by(estado_empleado='Activo').all()
    return render_template('nuevo_contrato.html', empleados=empleados)

@app.route('/contratos/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_contrato(id):
    """Editar contrato existente"""
    contrato = Contrato.query.get_or_404(id)
    
    if request.method == 'POST':
        contrato.empleado_id = request.form['empleado_id']
        contrato.tipo_contrato = request.form['tipo_contrato']
        contrato.fecha_inicio = datetime.strptime(request.form['fecha_inicio'], '%Y-%m-%d').date()
        contrato.fecha_fin = datetime.strptime(request.form['fecha_fin'], '%Y-%m-%d').date() if request.form.get('fecha_fin') else None
        contrato.salario = float(request.form['salario'])
        contrato.descripcion = request.form.get('descripcion', '')
        
        db.session.commit()
        flash('Contrato actualizado exitosamente', 'success')
        return redirect(url_for('contratos'))
    
    empleados = Empleado.query.filter_by(estado_empleado='Activo').all()
    return render_template('editar_contrato.html', contrato=contrato, empleados=empleados)

@app.route('/contratos/desactivar/<int:id>')
@login_required
def desactivar_contrato(id):
    """Desactivar contrato"""
    contrato = Contrato.query.get_or_404(id)
    contrato.activo = False
    db.session.commit()
    flash('Contrato desactivado exitosamente', 'success')
    return redirect(url_for('contratos'))

@app.route('/contratos/activar/<int:id>')
@login_required
def activar_contrato(id):
    """Activar contrato"""
    contrato = Contrato.query.get_or_404(id)
    contrato.activo = True
    db.session.commit()
    flash('Contrato activado exitosamente', 'success')
    return redirect(url_for('contratos'))

@app.route('/contratos/eliminar/<int:id>', methods=['DELETE'])
@login_required
def eliminar_contrato(id):
    """Eliminar contrato"""
    try:
        contrato = Contrato.query.get_or_404(id)
        db.session.delete(contrato)
        db.session.commit()
        return jsonify({
            'success': True,
            'message': 'Contrato eliminado exitosamente'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error al eliminar el contrato: {str(e)}'
        }), 500

@app.route('/contratos/generar/<int:id>')
@login_required
def generar_contrato(id):
    """Generar contrato Excel con control de duplicados"""
    try:
        # Verificar si ya existe un contrato generado para este empleado
        contrato = Contrato.query.get_or_404(id)
        empleado_id = contrato.empleado_id
        
        # Buscar contratos generados existentes para este empleado
        contrato_existente = ContratoGenerado.query.filter_by(empleado_id=empleado_id).first()
        
        if contrato_existente:
            # Si ya existe, preguntar si quiere regenerar
            flash(f'Ya existe un contrato generado para {contrato.empleado.nombre_completo}. Usa "Regenerar" para crear uno nuevo.', 'warning')
            return redirect(url_for('contratos_generados'))
        
        # Si no existe, generar nuevo contrato
        contrato_generado = generar_contrato_excel(id)
        flash(f'Contrato generado exitosamente: {contrato_generado.nombre_archivo}', 'success')
        return redirect(url_for('contratos'))
        
    except Exception as e:
        flash(f'Error al generar el contrato: {str(e)}', 'error')
        return redirect(url_for('contratos'))

@app.route('/contratos/generados')
@login_required
def contratos_generados():
    """Lista de contratos generados"""
    try:
        # Intentar crear la tabla si no existe
        try:
            with db.engine.connect() as connection:
                connection.execute(text("""
                    CREATE TABLE IF NOT EXISTS contrato_generado (
                        id SERIAL PRIMARY KEY,
                        empleado_id INTEGER NOT NULL REFERENCES empleado(id),
                        contrato_id INTEGER NOT NULL REFERENCES contrato(id),
                        nombre_archivo VARCHAR(255) NOT NULL,
                        ruta_archivo VARCHAR(500) NOT NULL,
                        fecha_generacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        activo BOOLEAN DEFAULT TRUE
                    );
                """))
                connection.commit()
            print("✅ Tabla contrato_generado creada/verificada")
        except Exception as create_error:
            print(f"⚠️ No se pudo crear la tabla: {create_error}")
        
        contratos_generados = ContratoGenerado.query.join(Empleado).join(Contrato).order_by(ContratoGenerado.fecha_generacion.desc()).all()
        return render_template('contratos_generados.html', contratos_generados=contratos_generados)
    except Exception as e:
        print(f"Error al cargar contratos generados: {str(e)}")
        flash('Error al cargar contratos generados. La tabla puede no existir aún.', 'error')
        return render_template('contratos_generados.html', contratos_generados=[])

@app.route('/contratos/descargar/<int:id>')
@login_required
def descargar_contrato(id):
    """Descargar contrato generado"""
    contrato_generado = ContratoGenerado.query.get_or_404(id)
    
    # Verificar si tenemos datos binarios en la base de datos
    if contrato_generado.archivo_data:
        # Crear respuesta desde datos binarios
        from io import BytesIO
        return send_file(
            BytesIO(contrato_generado.archivo_data),
            as_attachment=True,
            download_name=contrato_generado.nombre_archivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    else:
        # Fallback: intentar desde archivo (para contratos antiguos)
        if os.path.exists(contrato_generado.ruta_archivo):
            return send_file(
                contrato_generado.ruta_archivo,
                as_attachment=True,
                download_name=contrato_generado.nombre_archivo,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            flash('El archivo del contrato no existe', 'error')
            return redirect(url_for('contratos_generados'))

@app.route('/contratos/regenerar/<int:id>')
@login_required
def regenerar_contrato(id):
    """Regenerar contrato (eliminar el anterior y crear uno nuevo)"""
    try:
        # Intentar crear la tabla si no existe
        try:
            with db.engine.connect() as connection:
                connection.execute(text("""
                    CREATE TABLE IF NOT EXISTS contrato_generado (
                        id SERIAL PRIMARY KEY,
                        empleado_id INTEGER NOT NULL REFERENCES empleado(id),
                        contrato_id INTEGER NOT NULL REFERENCES contrato(id),
                        nombre_archivo VARCHAR(255) NOT NULL,
                        ruta_archivo VARCHAR(500) NOT NULL,
                        fecha_generacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        activo BOOLEAN DEFAULT TRUE
                    );
                """))
                connection.commit()
            print("✅ Tabla contrato_generado creada/verificada en regenerar")
        except Exception as create_error:
            print(f"⚠️ No se pudo crear la tabla en regenerar: {create_error}")
        
        # Obtener el contrato generado actual
        contrato_generado = ContratoGenerado.query.get_or_404(id)
        contrato_id = contrato_generado.contrato_id
        empleado_nombre = contrato_generado.empleado.nombre_completo
        
        # Eliminar el archivo anterior si existe
        if os.path.exists(contrato_generado.ruta_archivo):
            os.remove(contrato_generado.ruta_archivo)
            print(f"✅ Archivo anterior eliminado: {contrato_generado.ruta_archivo}")
        
        # Eliminar el registro de la base de datos
        db.session.delete(contrato_generado)
        db.session.commit()
        print(f"✅ Registro anterior eliminado de la base de datos")
        
        # Generar nuevo contrato
        nuevo_contrato = generar_contrato_excel(contrato_id)
        
        flash(f'Contrato regenerado exitosamente para {empleado_nombre}', 'success')
        return redirect(url_for('contratos_generados'))
        
    except Exception as e:
        print(f"Error al regenerar contrato: {str(e)}")
        flash(f'Error al regenerar el contrato: {str(e)}', 'error')
        return redirect(url_for('contratos_generados'))

@app.route('/contratos/vista_previa_simple/<int:id>')
@login_required
def vista_previa_simple(id):
    """Vista previa simple del contrato generado"""
    try:
        contrato_generado = ContratoGenerado.query.get_or_404(id)
        
        # Verificar si tenemos datos binarios en la base de datos
        if contrato_generado.archivo_data:
            # Leer desde datos binarios
            from io import BytesIO
            from openpyxl import load_workbook
            workbook = load_workbook(BytesIO(contrato_generado.archivo_data))
        else:
            # Fallback: intentar desde archivo (para contratos antiguos)
            if not os.path.exists(contrato_generado.ruta_archivo):
                return f"""
                <!DOCTYPE html>
                <html>
                <head>
                    <title>Error - Vista Previa</title>
                    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
                </head>
                <body>
                    <div class="container mt-5">
                        <div class="alert alert-danger">
                            <h4>Error</h4>
                            <p>El archivo del contrato no existe</p>
                        </div>
                    </div>
                </body>
                </html>
                """, 404
            
            # Leer el archivo Excel
            from openpyxl import load_workbook
            workbook = load_workbook(contrato_generado.ruta_archivo)
        
        worksheet = workbook.active
        
        # Crear página HTML simple y limpia
        pagina_completa = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Vista Previa - {contrato_generado.empleado.nombre_completo}</title>
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
            <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/css/all.min.css" rel="stylesheet">
            <style>
                body {{ 
                    font-family: Arial, sans-serif; 
                    margin: 0; 
                    padding: 20px;
                    background-color: #f8f9fa;
                }}
                .excel-table {{
                    border-collapse: collapse;
                    font-size: 10px;
                    font-family: Arial, sans-serif;
                    margin: 0 auto;
                    background-color: white;
                    box-shadow: 0 0 10px rgba(0,0,0,0.1);
                }}
                .excel-table td, .excel-table th {{
                    border: 1px solid #d0d7de;
                    padding: 2px 4px;
                    text-align: left;
                    vertical-align: top;
                    white-space: nowrap;
                    min-width: 60px;
                    max-width: 200px;
                    overflow: hidden;
                    text-overflow: ellipsis;
                }}
                .excel-table th {{
                    background-color: #f6f8fa;
                    font-weight: bold;
                    font-size: 10px;
                }}
                .excel-container {{
                    overflow: auto;
                    border: 1px solid #d0d7de;
                    margin: 20px auto;
                    max-width: 95%;
                    background-color: white;
                    border-radius: 6px;
                }}
                .excel-table tr:nth-child(even) {{
                    background-color: #f8f9fa;
                }}
                .excel-table tr:hover {{
                    background-color: #e3f2fd;
                }}
                @media print {{
                    .no-print {{ display: none !important; }}
                    body {{ padding: 0; }}
                    .excel-container {{ 
                        max-width: 100%; 
                        box-shadow: none; 
                        border: none;
                    }}
                    .excel-table {{ font-size: 8px; }}
                    .excel-table td, .excel-table th {{ 
                        padding: 1px 2px; 
                        min-width: 40px;
                    }}
                }}
                @media (max-width: 768px) {{
                    .excel-table {{ font-size: 8px; }}
                    .excel-table td, .excel-table th {{ 
                        padding: 1px 2px; 
                        min-width: 40px;
                    }}
                }}
            </style>
        </head>
        <body>
            <div class="container-fluid">
                <div class="d-flex justify-content-between align-items-center mb-4 no-print">
                    <h2><i class="fas fa-file-excel"></i> Vista Previa del Contrato</h2>
                    <div class="d-flex align-items-center">
                        <div class="me-3">
                            <label class="form-label me-2">Zoom:</label>
                            <select id="zoomSelect" class="form-select form-select-sm" style="width: auto;">
                                <option value="0.5">50%</option>
                                <option value="0.75">75%</option>
                                <option value="1" selected>100%</option>
                                <option value="1.25">125%</option>
                                <option value="1.5">150%</option>
                            </select>
                        </div>
                        <button onclick="window.print()" class="btn btn-info me-2">
                            <i class="fas fa-print"></i> Imprimir
                        </button>
                        <a href="{{ url_for('descargar_contrato', id=contrato_generado.id) }}" class="btn btn-success me-2">
                            <i class="fas fa-download"></i> Descargar Excel
                        </a>
                        <button onclick="window.close()" class="btn btn-secondary">
                            <i class="fas fa-times"></i> Cerrar
                        </button>
                    </div>
                </div>
                
                <div class="excel-container" id="excelContainer">
                    <table class="excel-table">
        """
        
        # Obtener dimensiones de la hoja
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # Procesar cada fila
        for row in range(1, min(max_row + 1, 100)):  # Aumentar límite para ver más contenido
            pagina_completa += '<tr>'
            
            for col in range(1, min(max_col + 1, 30)):  # Aumentar límite de columnas
                cell = worksheet.cell(row=row, column=col)
                cell_value = str(cell.value) if cell.value is not None else ''
                
                # Limpiar y formatear el valor de la celda
                if cell_value and len(cell_value) > 50:
                    cell_value = cell_value[:47] + '...'
                
                # Aplicar estilos más detallados
                cell_style = 'font-size: 10px; '
                
                # Estilos de fuente
                if cell.font:
                    if cell.font.bold:
                        cell_style += 'font-weight: bold; '
                    if cell.font.italic:
                        cell_style += 'font-style: italic; '
                    if cell.font.size:
                        cell_style += f'font-size: {min(cell.font.size, 12)}px; '
                
                # Estilos de relleno
                if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                    cell_style += f'background-color: {cell.fill.start_color.rgb}; '
                
                # Estilos de alineación
                if cell.alignment:
                    if cell.alignment.horizontal == 'center':
                        cell_style += 'text-align: center; '
                    elif cell.alignment.horizontal == 'right':
                        cell_style += 'text-align: right; '
                    elif cell.alignment.horizontal == 'left':
                        cell_style += 'text-align: left; '
                    
                    if cell.alignment.vertical == 'center':
                        cell_style += 'vertical-align: middle; '
                    elif cell.alignment.vertical == 'bottom':
                        cell_style += 'vertical-align: bottom; '
                
                # Estilos de borde
                if cell.border:
                    if cell.border.left and cell.border.left.style:
                        cell_style += 'border-left: 2px solid #000; '
                    if cell.border.right and cell.border.right.style:
                        cell_style += 'border-right: 2px solid #000; '
                    if cell.border.top and cell.border.top.style:
                        cell_style += 'border-top: 2px solid #000; '
                    if cell.border.bottom and cell.border.bottom.style:
                        cell_style += 'border-bottom: 2px solid #000; '
                
                # Determinar el tipo de celda
                if row == 1 or (cell.font and cell.font.bold):
                    pagina_completa += f'<th style="{cell_style}">{cell_value}</th>'
                else:
                    pagina_completa += f'<td style="{cell_style}">{cell_value}</td>'
            
            pagina_completa += '</tr>'
        
        pagina_completa += """
                    </table>
                </div>
            </div>
            
            <script>
                // Control de zoom
                document.getElementById('zoomSelect').addEventListener('change', function() {
                    const zoom = parseFloat(this.value);
                    const container = document.getElementById('excelContainer');
                    const table = container.querySelector('.excel-table');
                    
                    // Aplicar zoom
                    container.style.transform = `scale(${zoom})`;
                    container.style.transformOrigin = 'top left';
                    
                    // Ajustar el contenedor para el zoom
                    if (zoom !== 1) {
                        container.style.width = `${100 / zoom}%`;
                        container.style.height = `${100 / zoom}%`;
                    } else {
                        container.style.width = '100%';
                        container.style.height = 'auto';
                    }
                });
                
                // Zoom con rueda del mouse
                document.getElementById('excelContainer').addEventListener('wheel', function(e) {
                    if (e.ctrlKey) {
                        e.preventDefault();
                        const zoomSelect = document.getElementById('zoomSelect');
                        const currentZoom = parseFloat(zoomSelect.value);
                        const options = zoomSelect.options;
                        let newIndex = zoomSelect.selectedIndex;
                        
                        if (e.deltaY < 0) { // Zoom in
                            newIndex = Math.min(newIndex + 1, options.length - 1);
                        } else { // Zoom out
                            newIndex = Math.max(newIndex - 1, 0);
                        }
                        
                        zoomSelect.selectedIndex = newIndex;
                        zoomSelect.dispatchEvent(new Event('change'));
                    }
                });
            </script>
        </body>
        </html>
        """
        
        return pagina_completa
        
    except Exception as e:
        print(f"Error al generar vista previa simple: {str(e)}")
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Error - Vista Previa</title>
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
        </head>
        <body>
            <div class="container mt-5">
                <div class="alert alert-danger">
                    <h4>Error</h4>
                    <p>Error al generar vista previa: {str(e)}</p>
                </div>
            </div>
        </body>
        </html>
        """, 500


@app.route('/contratos/vista_previa/<int:id>')
@login_required
def vista_previa_contrato(id):
    """Vista previa del contrato generado"""
    try:
        contrato_generado = ContratoGenerado.query.get_or_404(id)
        
        # Verificar si tenemos datos binarios en la base de datos
        if contrato_generado.archivo_data:
            # Leer desde datos binarios
            from io import BytesIO
            from openpyxl import load_workbook
            workbook = load_workbook(BytesIO(contrato_generado.archivo_data))
        else:
            # Fallback: intentar desde archivo (para contratos antiguos)
            if not os.path.exists(contrato_generado.ruta_archivo):
                return f"""
                <!DOCTYPE html>
                <html>
                <head>
                    <title>Error - Vista Previa</title>
                    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
                </head>
                <body>
                    <div class="container mt-5">
                        <div class="alert alert-danger">
                            <h4>Error</h4>
                            <p>El archivo del contrato no existe</p>
                        </div>
                    </div>
                </body>
                </html>
                """, 404
            
            # Leer el archivo Excel
            from openpyxl import load_workbook
            workbook = load_workbook(contrato_generado.ruta_archivo)
        
        worksheet = workbook.active
        
        # Convertir a HTML
        html_content = convertir_excel_a_html(worksheet, contrato_generado)
        
        # Crear página HTML completa
        pagina_completa = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Vista Previa - {contrato_generado.empleado.nombre_completo}</title>
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
            <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/css/all.min.css" rel="stylesheet">
            <style>
                .vista-previa-excel {{ max-width: 100%; }}
                .table {{ 
                    font-size: 11px; 
                    font-family: Arial, sans-serif;
                    border-collapse: collapse;
                }}
                .table td, .table th {{
                    border: 1px solid #000;
                    padding: 4px 8px;
                    vertical-align: top;
                }}
                .table-responsive {{
                    overflow-x: auto;
                    border: 1px solid #ddd;
                }}
                @media print {{
                    .no-print {{ display: none !important; }}
                    body {{ font-size: 10px; }}
                    .table {{ font-size: 9px; }}
                    .table td, .table th {{
                        padding: 2px 4px;
                    }}
                }}
            </style>
        </head>
        <body>
            <div class="container-fluid">
                <div class="d-flex justify-content-between align-items-center mb-4 no-print">
                    <h2><i class="fas fa-file-excel"></i> Vista Previa del Contrato</h2>
                    <div>
                        <button onclick="window.print()" class="btn btn-info me-2">
                            <i class="fas fa-print"></i> Imprimir
                        </button>
                        <a href="{{ url_for('descargar_contrato', id=contrato_generado.id) }}" class="btn btn-success me-2">
                            <i class="fas fa-download"></i> Descargar Excel
                        </a>
                        <button onclick="window.close()" class="btn btn-secondary">
                            <i class="fas fa-times"></i> Cerrar
                        </button>
                    </div>
                </div>
                {html_content}
            </div>
        </body>
        </html>
        """
        
        return pagina_completa
        
    except Exception as e:
        print(f"Error al generar vista previa: {str(e)}")
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Error - Vista Previa</title>
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
        </head>
        <body>
            <div class="container mt-5">
                <div class="alert alert-danger">
                    <h4>Error</h4>
                    <p>Error al generar vista previa: {str(e)}</p>
                </div>
            </div>
        </body>
        </html>
        """, 500

@app.route('/contratos/eliminar_generado/<int:id>', methods=['DELETE', 'POST'])
@login_required
def eliminar_contrato_generado(id):
    """Eliminar contrato generado"""
    try:
        # Intentar crear la tabla si no existe
        try:
            with db.engine.connect() as connection:
                connection.execute(text("""
                    CREATE TABLE IF NOT EXISTS contrato_generado (
                        id SERIAL PRIMARY KEY,
                        empleado_id INTEGER NOT NULL REFERENCES empleado(id),
                        contrato_id INTEGER NOT NULL REFERENCES contrato(id),
                        nombre_archivo VARCHAR(255) NOT NULL,
                        ruta_archivo VARCHAR(500) NOT NULL,
                        fecha_generacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        activo BOOLEAN DEFAULT TRUE
                    );
                """))
                connection.commit()
            print("✅ Tabla contrato_generado creada/verificada en eliminar")
        except Exception as create_error:
            print(f"⚠️ No se pudo crear la tabla en eliminar: {create_error}")
        
        contrato_generado = ContratoGenerado.query.get_or_404(id)
        empleado_nombre = contrato_generado.empleado.nombre_completo
        
        # Eliminar el archivo si existe (para contratos antiguos)
        if os.path.exists(contrato_generado.ruta_archivo):
            os.remove(contrato_generado.ruta_archivo)
            print(f"✅ Archivo eliminado: {contrato_generado.ruta_archivo}")
        
        # Eliminar el registro de la base de datos (incluye datos binarios)
        db.session.delete(contrato_generado)
        db.session.commit()
        
        # Si es una petición AJAX, devolver JSON
        if request.headers.get('Content-Type') == 'application/json' or request.method == 'DELETE':
            return jsonify({
                'success': True,
                'message': f'Contrato de {empleado_nombre} eliminado exitosamente'
            })
        else:
            # Si es una petición normal, redirigir
            flash(f'Contrato de {empleado_nombre} eliminado exitosamente', 'success')
            return redirect(url_for('contratos_generados'))
        
    except Exception as e:
        db.session.rollback()
        print(f"Error al eliminar contrato generado: {str(e)}")
        
        # Si es una petición AJAX, devolver JSON
        if request.headers.get('Content-Type') == 'application/json' or request.method == 'DELETE':
            return jsonify({
                'success': False,
                'message': f'Error al eliminar el contrato: {str(e)}'
            }), 500
        else:
            # Si es una petición normal, redirigir con error
            flash(f'Error al eliminar contrato: {str(e)}', 'error')
            return redirect(url_for('contratos_generados'))

@app.route('/admin/arreglar_contratos')
@login_required
def arreglar_contratos():
    """Arreglar contratos generados existentes"""
    try:
        # Verificar si es administrador
        if not current_user.is_admin:
            flash('Acceso denegado. Solo administradores pueden acceder a esta función.', 'error')
            return redirect(url_for('dashboard'))
        
        print("🚀 Arreglando contratos generados...")
        
        # Verificar si la columna archivo_data existe
        try:
            with db.engine.connect() as conn:
                result = conn.execute(text("""
                    SELECT column_name 
                    FROM information_schema.columns 
                    WHERE table_name = 'contrato_generado' 
                    AND column_name = 'archivo_data';
                """))
                
                if result.fetchone():
                    print("✅ La columna archivo_data ya existe")
                    columna_existe = True
                else:
                    print("📝 Agregando columna archivo_data...")
                    conn.execute(text("""
                        ALTER TABLE contrato_generado 
                        ADD COLUMN archivo_data BYTEA;
                    """))
                    print("✅ Columna archivo_data agregada")
                    columna_existe = False
                
                # Contar contratos existentes
                result = conn.execute(text("SELECT COUNT(*) FROM contrato_generado;"))
                count_antes = result.fetchone()[0]
                
                if count_antes > 0:
                    print(f"🗑️ Eliminando {count_antes} contratos generados existentes (sin datos binarios)...")
                    conn.execute(text("DELETE FROM contrato_generado;"))
                    print("✅ Contratos existentes eliminados")
                
                # Confirmar cambios
                conn.commit()
                
                mensaje = f"✅ Arreglo completado exitosamente! Columna archivo_data: {'ya existía' if columna_existe else 'agregada'}. Contratos antiguos eliminados: {count_antes}"
                flash(mensaje, 'success')
                print(mensaje)
                
                return redirect(url_for('contratos_generados'))
                
        except Exception as e:
            print(f"❌ Error durante el arreglo: {str(e)}")
            flash(f'Error al arreglar contratos: {str(e)}', 'error')
            return redirect(url_for('contratos_generados'))
            
    except Exception as e:
        print(f"❌ Error general: {str(e)}")
        flash(f'Error general: {str(e)}', 'error')
        return redirect(url_for('dashboard'))

@app.route('/visitantes/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_visitante():
    if request.method == 'POST':
        visitante = Visitante(
            nombre=request.form['nombre'],
            apellido=request.form['apellido'],
            documento=request.form['documento'],
            eps=request.form['eps'],
            rh=request.form['rh'],
            telefono=request.form['telefono'],
            empresa=request.form.get('empresa', ''),
            motivo_visita=request.form['motivo_visita'],
            nombre_contacto_emergencia=request.form['nombre_contacto_emergencia'],
            telefono_emergencia=request.form['telefono_emergencia'],
            parentesco=request.form['parentesco'],
            estado_visita='Pendiente',  # Estado inicial - esperando entrada
            activo=False  # No activo hasta que se registre la entrada
        )
        db.session.add(visitante)
        db.session.commit()
        flash('Visitante registrado exitosamente. Use el botón de Entrada/Salida para registrar su llegada.', 'success')
        return redirect(url_for('visitantes'))
    
    return render_template('nuevo_visitante.html')

@app.route('/visitantes/<int:id>/entrada-salida', methods=['POST'])
@login_required
def registrar_entrada_salida_visitante(id):
    visitante = Visitante.query.get_or_404(id)
    
    if visitante.estado_visita == 'En visita':
        # Registrar salida
        visitante.fecha_salida = colombia_now()
        visitante.estado_visita = 'Finalizada'
        visitante.activo = False
        db.session.commit()
        
        # Enviar notificación
        notificar_visitante_salida(f"{visitante.nombre} {visitante.apellido}")
        
        flash(f'Salida registrada para {visitante.nombre} {visitante.apellido} a las {visitante.fecha_salida.strftime("%H:%M")}', 'success')
    else:
        # Registrar entrada (nuevo visitante)
        visitante.fecha_entrada = colombia_now()
        visitante.estado_visita = 'En visita'
        visitante.activo = True
        visitante.fecha_salida = None
        db.session.commit()
        
        # Enviar notificación
        notificar_visitante_nuevo(
            f"{visitante.nombre} {visitante.apellido}",
            visitante.empresa or "Sin empresa"
        )
        
        flash(f'Entrada registrada para {visitante.nombre} {visitante.apellido} a las {visitante.fecha_entrada.strftime("%H:%M")}', 'success')
    
    return redirect(url_for('visitantes'))

# Reportes
@app.route('/reportes')
@login_required
def reportes():
    return render_template('reportes.html')

@app.route('/reportes/asistencia')
@login_required
def reporte_asistencia():
    fecha_inicio = request.args.get('fecha_inicio', (date.today() - timedelta(days=30)).strftime('%Y-%m-%d'))
    fecha_fin = request.args.get('fecha_fin', date.today().strftime('%Y-%m-%d'))
    
    asistencias = Asistencia.query.filter(
        Asistencia.fecha >= datetime.strptime(fecha_inicio, '%Y-%m-%d').date(),
        Asistencia.fecha <= datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    ).all()
    
    return render_template('reporte_asistencia.html', asistencias=asistencias, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)

@app.route('/reportes/empleados')
@login_required
def reporte_empleados():
    empleados = Empleado.query.filter_by(estado_empleado='Activo').all()
    return render_template('reporte_empleados.html', empleados=empleados)

@app.route('/reportes/visitantes')
@login_required
def reporte_visitantes():
    fecha_inicio = request.args.get('fecha_inicio', (date.today() - timedelta(days=30)).strftime('%Y-%m-%d'))
    fecha_fin = request.args.get('fecha_fin', date.today().strftime('%Y-%m-%d'))
    
    visitantes = Visitante.query.filter(
        Visitante.fecha_entrada >= datetime.strptime(fecha_inicio, '%Y-%m-%d').date(),
        Visitante.fecha_entrada <= datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    ).all()
    
    return render_template('reporte_visitantes.html', visitantes=visitantes, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)

# Sistema de Backups
@app.route('/backups')
@login_required
def backups():
    """Página de gestión de backups"""
    if not current_user.is_admin:
        flash('Solo los administradores pueden acceder a esta sección', 'error')
        return redirect(url_for('dashboard'))
    
    # Listar backups existentes
    backups_dir = 'backups'
    backups_list = []
    if os.path.exists(backups_dir):
        for archivo in os.listdir(backups_dir):
            if archivo.endswith('.db') or archivo.endswith('.sql'):
                ruta_completa = os.path.join(backups_dir, archivo)
                tamaño = os.path.getsize(ruta_completa)
                fecha_mod = datetime.fromtimestamp(os.path.getmtime(ruta_completa))
                backups_list.append({
                    'nombre': archivo,
                    'tamaño': tamaño,
                    'fecha': fecha_mod,
                    'ruta': ruta_completa
                })
        backups_list.sort(key=lambda x: x['fecha'], reverse=True)
    
    return render_template('backups.html', backups=backups_list)

@app.route('/backups/crear', methods=['POST'])
@login_required
def crear_backup():
    """Crear un backup de la base de datos"""
    if not current_user.is_admin:
        flash('Solo los administradores pueden crear backups', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        backups_dir = 'backups'
        if not os.path.exists(backups_dir):
            os.makedirs(backups_dir)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Detectar tipo de base de datos
        database_url = app.config.get('SQLALCHEMY_DATABASE_URI', '')
        
        if 'sqlite' in database_url.lower():
            # Backup SQLite
            db_path = database_url.replace('sqlite:///', '')
            if os.path.exists(db_path):
                backup_path = os.path.join(backups_dir, f'backup_{timestamp}.db')
                shutil.copy2(db_path, backup_path)
                flash(f'Backup creado exitosamente: {os.path.basename(backup_path)}', 'success')
            else:
                flash('No se encontró la base de datos SQLite', 'error')
        else:
            # Backup PostgreSQL usando pg_dump
            import subprocess
            backup_path = os.path.join(backups_dir, f'backup_{timestamp}.sql')
            
            # Extraer información de conexión
            if database_url.startswith('postgresql+psycopg://'):
                database_url = database_url.replace('postgresql+psycopg://', 'postgresql://', 1)
            
            # Intentar crear backup
            try:
                # Usar pg_dump si está disponible
                result = subprocess.run(
                    ['pg_dump', database_url],
                    capture_output=True,
                    text=True,
                    timeout=300
                )
                if result.returncode == 0:
                    with open(backup_path, 'w', encoding='utf-8') as f:
                        f.write(result.stdout)
                    flash(f'Backup creado exitosamente: {os.path.basename(backup_path)}', 'success')
                else:
                    flash('Error al crear backup de PostgreSQL. Asegúrese de tener pg_dump instalado.', 'error')
            except FileNotFoundError:
                flash('pg_dump no está instalado. No se puede crear backup de PostgreSQL automáticamente.', 'error')
            except Exception as e:
                flash(f'Error al crear backup: {str(e)}', 'error')
    
    except Exception as e:
        flash(f'Error al crear backup: {str(e)}', 'error')
    
    return redirect(url_for('backups'))

@app.route('/backups/descargar/<path:nombre>')
@login_required
def descargar_backup(nombre):
    """Descargar un backup"""
    if not current_user.is_admin:
        flash('Solo los administradores pueden descargar backups', 'error')
        return redirect(url_for('dashboard'))
    
    backups_dir = 'backups'
    ruta_backup = os.path.join(backups_dir, nombre)
    
    if os.path.exists(ruta_backup) and os.path.commonpath([backups_dir, ruta_backup]) == backups_dir:
        return send_file(ruta_backup, as_attachment=True, download_name=nombre)
    else:
        flash('Backup no encontrado', 'error')
        return redirect(url_for('backups'))

@app.route('/backups/eliminar/<path:nombre>', methods=['POST'])
@login_required
def eliminar_backup(nombre):
    """Eliminar un backup"""
    if not current_user.is_admin:
        flash('Solo los administradores pueden eliminar backups', 'error')
        return redirect(url_for('dashboard'))
    
    backups_dir = 'backups'
    ruta_backup = os.path.join(backups_dir, nombre)
    
    if os.path.exists(ruta_backup) and os.path.commonpath([backups_dir, ruta_backup]) == backups_dir:
        try:
            os.remove(ruta_backup)
            flash('Backup eliminado exitosamente', 'success')
        except Exception as e:
            flash(f'Error al eliminar backup: {str(e)}', 'error')
    else:
        flash('Backup no encontrado', 'error')
    
    return redirect(url_for('backups'))

# Sistema de Control de Cesantías
@app.route('/cesantias')
@login_required
def cesantias():
    """Página principal de control de cesantías"""
    if not current_user.is_admin:
        flash('Solo los administradores pueden acceder a esta sección', 'error')
        return redirect(url_for('dashboard'))
    
    # Obtener todas las solicitudes de retiro de cesantías aprobadas
    solicitudes_cesantias = SolicitudEmpleado.query.filter_by(
        tipo_solicitud='RETIRO_CESANTIAS',
        estado='APROBADA'
    ).order_by(SolicitudEmpleado.fecha_aprobacion.desc()).all()
    
    # Obtener también las pendientes para mostrar
    solicitudes_pendientes = SolicitudEmpleado.query.filter_by(
        tipo_solicitud='RETIRO_CESANTIAS',
        estado='PENDIENTE'
    ).order_by(SolicitudEmpleado.created_at.desc()).all()
    
    # Estadísticas
    total_retiros = len(solicitudes_cesantias)
    retiros_mes_actual = len([s for s in solicitudes_cesantias 
                              if s.fecha_aprobacion and 
                              s.fecha_aprobacion.month == date.today().month and
                              s.fecha_aprobacion.year == date.today().year])
    pendientes = len(solicitudes_pendientes)
    
    return render_template('cesantias.html',
                         solicitudes_cesantias=solicitudes_cesantias,
                         solicitudes_pendientes=solicitudes_pendientes,
                         total_retiros=total_retiros,
                         retiros_mes_actual=retiros_mes_actual,
                         pendientes=pendientes)

# Inicialización de la base de datos
def init_db():
    try:
        print("🚀 INICIANDO INICIALIZACIÓN DE BASE DE DATOS")
        print("=" * 60)
        
        with app.app_context():
            print("📊 Creando tablas de la base de datos...")
            db.create_all()
            print("✅ Tablas principales creadas")
            
            # Ejecutar migración de tablas de inventario
            print("🔄 Iniciando migración de tablas de inventario...")
            from sqlalchemy import text
            
            # Crear tablas de inventario si no existen
            tablas_inventario = [
                {
                    'nombre': 'categoria_inventario',
                    'sql': """
                        CREATE TABLE IF NOT EXISTS categoria_inventario (
                            id SERIAL PRIMARY KEY,
                            nombre VARCHAR(100) NOT NULL UNIQUE,
                            descripcion TEXT,
                            activa BOOLEAN DEFAULT TRUE,
                            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                        );
                    """
                },
                {
                    'nombre': 'producto',
                    'sql': """
                        CREATE TABLE IF NOT EXISTS producto (
                            id SERIAL PRIMARY KEY,
                            codigo VARCHAR(50) NOT NULL UNIQUE,
                            nombre VARCHAR(200) NOT NULL,
                            descripcion TEXT,
                            categoria_id INTEGER NOT NULL REFERENCES categoria_inventario(id),
                            unidad_medida VARCHAR(20) NOT NULL,
                            precio_unitario NUMERIC(10, 2) DEFAULT 0,
                            stock_minimo INTEGER DEFAULT 0,
                            stock_actual INTEGER DEFAULT 0,
                            ubicacion VARCHAR(100),
                            proveedor VARCHAR(200),
                            fecha_vencimiento DATE,
                            lote VARCHAR(50),
                            activo BOOLEAN DEFAULT TRUE,
                            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                        );
                    """
                },
                {
                    'nombre': 'movimiento_inventario',
                    'sql': """
                        CREATE TABLE IF NOT EXISTS movimiento_inventario (
                            id SERIAL PRIMARY KEY,
                            producto_id INTEGER NOT NULL REFERENCES producto(id),
                            tipo_movimiento VARCHAR(20) NOT NULL,
                            cantidad INTEGER NOT NULL,
                            precio_unitario NUMERIC(10, 2),
                            total NUMERIC(10, 2),
                            motivo VARCHAR(200),
                            referencia VARCHAR(100),
                            responsable VARCHAR(200),
                            observaciones TEXT,
                            fecha_movimiento TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                            created_by INTEGER REFERENCES "user"(id)
                        );
                    """
                }
            ]
            
            for tabla in tablas_inventario:
                print(f"📝 Procesando tabla {tabla['nombre']}...")
                try:
                    with db.engine.connect() as conn:
                        conn.execute(text(tabla['sql']))
                        conn.commit()
                    print(f"✅ Tabla {tabla['nombre']} creada/verificada exitosamente")
                except Exception as e:
                    print(f"❌ ERROR con tabla {tabla['nombre']}: {str(e)}")
                    # Continuar con las demás tablas
                    continue
            
            # Verificar columna archivo_data en contrato_generado
            print("📝 Verificando columna archivo_data en contrato_generado...")
            try:
                with db.engine.connect() as conn:
                    conn.execute(text("""
                        ALTER TABLE contrato_generado 
                        ADD COLUMN IF NOT EXISTS archivo_data BYTEA;
                    """))
                    conn.commit()
                print("✅ Columna archivo_data verificada exitosamente")
            except Exception as e:
                print(f"⚠️ Columna archivo_data: {str(e)}")
            
            print("✅ Migración de inventarios completada")
            
            print("✅ Sistema de inventarios simplificado - categorías fijas: ALMACEN GENERAL, QUIMICOS, POSCOSECHA")
            
            # Crear tabla de notificaciones
            print("🔔 Creando tabla de notificaciones...")
            try:
                with db.engine.connect() as conn:
                    conn.execute(text("""
                        CREATE TABLE IF NOT EXISTS notificacion (
                            id SERIAL PRIMARY KEY,
                            titulo VARCHAR(200) NOT NULL,
                            mensaje TEXT NOT NULL,
                            tipo VARCHAR(20) NOT NULL DEFAULT 'info',
                            tipo_sonido VARCHAR(20) NOT NULL DEFAULT 'alerta',
                            icono VARCHAR(50) NOT NULL DEFAULT 'fas fa-bell',
                            leida BOOLEAN NOT NULL DEFAULT FALSE,
                            fecha_creacion TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                            usuario_id INTEGER REFERENCES "user"(id)
                        )
                    """))
                    conn.commit()
                print("✅ Tabla de notificaciones creada exitosamente")
            except Exception as e:
                print(f"⚠️ Tabla de notificaciones: {str(e)}")
            
            # Crear usuario administrador por defecto
            print("👤 Verificando usuario administrador...")
            admin_user = User.query.filter_by(email='admin@floresjuncalito.com').first()
            if not admin_user:
                admin_user = User(
                    email='admin@floresjuncalito.com',
                    username='Administrador',
                    password_hash=generate_password_hash('nueva_contraseña_2024'),
                    is_admin=True
                )
                db.session.add(admin_user)
                db.session.commit()
                print("✅ Usuario administrador creado: admin@floresjuncalito.com / nueva_contraseña_2024")
            else:
                print("✅ Usuario administrador ya existe")
            
            print("=" * 60)
            print("🎉 BASE DE DATOS INICIALIZADA CORRECTAMENTE")
            print("=" * 60)
            
    except Exception as e:
        print("=" * 60)
        print(f"❌ ERROR CRÍTICO al inicializar la base de datos: {str(e)}")
        print("=" * 60)
        import traceback
        traceback.print_exc()
        raise

# ===== RUTAS PARA SISTEMA DE INVENTARIOS =====

@app.route('/inventarios')
@login_required
def inventarios():
    """Página principal del sistema de inventarios"""
    # Obtener período actual o desde parámetros
    periodo_actual = get_periodo_desde_params()
    
    categorias_fijas = ['ALMACEN GENERAL', 'QUIMICOS', 'POSCOSECHA']
    
    # Obtener productos del período actual
    try:
        productos = Producto.query.filter_by(activo=True, periodo=periodo_actual).all()
    except Exception as e:
        print(f"⚠️ Error obteniendo productos por período: {e}")
        # Fallback: obtener todos los productos activos
        productos = Producto.query.filter_by(activo=True).all()
    
    # Estadísticas rápidas por categoría
    stats_por_categoria = {}
    for categoria in categorias_fijas:
        productos_cat = [p for p in productos if p.categoria == categoria]
        total_productos_cat = len(productos_cat)
        productos_bajo_stock_cat = len([p for p in productos_cat if p.stock_actual <= p.stock_minimo])
        
        # Calcular valor excluyendo precios anormales (> $10M, probablemente errores)
        valor_total_cat = sum(
            p.stock_actual * p.precio_unitario 
            for p in productos_cat 
            if p.precio_unitario and p.precio_unitario < 10000000
        )
        
        # Contar productos con precios anormales
        productos_precio_anormal = len([
            p for p in productos_cat 
            if p.precio_unitario and p.precio_unitario >= 10000000
        ])
        
        stats_por_categoria[categoria] = {
            'total_productos': total_productos_cat,
            'productos_bajo_stock': productos_bajo_stock_cat,
            'valor_total': valor_total_cat,
            'productos_precio_anormal': productos_precio_anormal
        }
    
    # Estadísticas generales
    total_productos = len(productos)
    productos_bajo_stock = len([p for p in productos if p.stock_actual <= p.stock_minimo])
    
    # Calcular valor total excluyendo precios anormales (> $10M)
    valor_total_inventario = sum(
        p.stock_actual * p.precio_unitario 
        for p in productos 
        if p.precio_unitario and p.precio_unitario < 10000000
    )
    
    # Identificar productos con precios anormales
    productos_precio_anormal = [
        p for p in productos 
        if p.precio_unitario and p.precio_unitario >= 10000000
    ]
    
    # Valor que se está excluyendo (probablemente errores)
    valor_excluido = sum(
        p.stock_actual * p.precio_unitario 
        for p in productos_precio_anormal
    ) if productos_precio_anormal else 0
    
    # Obtener períodos disponibles para el selector
    try:
        periodos_disponibles = db.session.query(Producto.periodo).distinct().order_by(Producto.periodo.desc()).all()
        periodos_disponibles = [p[0] for p in periodos_disponibles if p[0] is not None]
    except Exception as e:
        print(f"⚠️ Error obteniendo períodos: {e}")
        periodos_disponibles = [periodo_actual]
    
    return render_template('inventarios.html', 
                         categorias=categorias_fijas,
                         productos=productos,
                         periodo_actual=periodo_actual,
                         periodos_disponibles=periodos_disponibles,
                         stats_por_categoria=stats_por_categoria,
                         total_productos=total_productos,
                         productos_bajo_stock=productos_bajo_stock,
                         valor_total_inventario=valor_total_inventario,
                         productos_precio_anormal=productos_precio_anormal,
                         valor_excluido=valor_excluido)

@app.route('/inventarios/productos')
@login_required
def productos_inventario():
    """Lista de productos del inventario con búsqueda avanzada"""
    categoria = request.args.get('categoria', '')
    busqueda = request.args.get('busqueda', '').strip()
    periodo = request.args.get('periodo', '')
    orden = request.args.get('orden', 'nombre')
    stock_bajo = request.args.get('stock_bajo', '')
    precio_min = request.args.get('precio_min', '').strip()
    precio_max = request.args.get('precio_max', '').strip()
    
    query = Producto.query.filter_by(activo=True)
    
    # Filtro por categoría
    if categoria:
        query = query.filter_by(categoria=categoria)
    
    # Filtro por período
    if periodo:
        query = query.filter_by(periodo=periodo)
    
    # Búsqueda mejorada (código, nombre, descripción, proveedor, ubicación)
    if busqueda:
        search_filter = db.or_(
            Producto.nombre.contains(busqueda),
            Producto.codigo.contains(busqueda),
            Producto.descripcion.contains(busqueda),
            Producto.proveedor.contains(busqueda),
            Producto.ubicacion.contains(busqueda)
        )
        query = query.filter(search_filter)
    
    # Filtro de stock bajo
    if stock_bajo:
        query = query.filter(Producto.stock_actual <= Producto.stock_minimo)
    
    # Filtro por rango de precio
    if precio_min:
        try:
            query = query.filter(Producto.precio_unitario >= float(precio_min))
        except ValueError:
            pass
    
    if precio_max:
        try:
            query = query.filter(Producto.precio_unitario <= float(precio_max))
        except ValueError:
            pass
    
    # Ordenamiento
    if orden == 'nombre':
        query = query.order_by(Producto.nombre)
    elif orden == 'codigo':
        query = query.order_by(Producto.codigo)
    elif orden == 'stock_asc':
        query = query.order_by(Producto.stock_actual)
    elif orden == 'stock_desc':
        query = query.order_by(Producto.stock_actual.desc())
    elif orden == 'precio_asc':
        query = query.order_by(Producto.precio_unitario)
    elif orden == 'precio_desc':
        query = query.order_by(Producto.precio_unitario.desc())
    elif orden == 'categoria':
        query = query.order_by(Producto.categoria, Producto.nombre)
    
    productos = query.all()
    categorias_fijas = ['ALMACEN GENERAL', 'QUIMICOS', 'POSCOSECHA']
    
    # Obtener períodos disponibles
    try:
        periodos_disponibles = db.session.query(Producto.periodo).distinct().order_by(Producto.periodo.desc()).all()
        periodos_disponibles = [p[0] for p in periodos_disponibles if p[0] is not None]
    except:
        periodos_disponibles = []
    
    # Calcular estadísticas de los productos filtrados
    total_productos = len(productos)
    total_stock = sum(p.stock_actual for p in productos)
    
    # Calcular valor total excluyendo precios anormales (> $10M)
    valor_total = sum(
        p.stock_actual * p.precio_unitario 
        for p in productos 
        if p.precio_unitario and p.precio_unitario < 10000000
    )
    
    # Productos con stock bajo
    productos_stock_bajo = sum(1 for p in productos if p.stock_actual <= p.stock_minimo and p.stock_minimo > 0)
    
    # Identificar productos con precios anormales en los resultados
    productos_precio_anormal = [
        p for p in productos 
        if p.precio_unitario and p.precio_unitario >= 10000000
    ]
    
    return render_template('productos_inventario.html', 
                         productos=productos, 
                         categorias=categorias_fijas,
                         categoria_actual=categoria,
                         busqueda_actual=busqueda,
                         periodo_actual=periodo,
                         periodos_disponibles=periodos_disponibles,
                         orden_actual=orden,
                         stock_bajo_activo=stock_bajo,
                         precio_min_actual=precio_min,
                         precio_max_actual=precio_max,
                         total_productos=total_productos,
                         total_stock=total_stock,
                         valor_total=valor_total,
                         productos_stock_bajo=productos_stock_bajo,
                         productos_precio_anormal=productos_precio_anormal)

@app.route('/api/buscar-productos')
@login_required
def api_buscar_productos():
    """API para búsqueda rápida de productos"""
    q = request.args.get('q', '').strip()
    limit = request.args.get('limit', 10, type=int)
    
    if not q or len(q) < 2:
        return jsonify([])
    
    # Buscar en código, nombre y proveedor
    productos = Producto.query.filter(
        Producto.activo == True,
        db.or_(
            Producto.codigo.ilike(f'%{q}%'),
            Producto.nombre.ilike(f'%{q}%'),
            Producto.proveedor.ilike(f'%{q}%')
        )
    ).limit(limit).all()
    
    resultados = []
    for p in productos:
        resultados.append({
            'id': p.id,
            'codigo': p.codigo,
            'nombre': p.nombre,
            'categoria': p.categoria,
            'stock_actual': p.stock_actual,
            'unidad_medida': p.unidad_medida,
            'precio_unitario': float(p.precio_unitario),
            'proveedor': p.proveedor or '',
            'periodo': p.periodo
        })
    
    return jsonify(resultados)

@app.route('/inventarios/productos/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_producto_inventario():
    """Crear nuevo producto"""
    # Obtener período actual o desde parámetros
    periodo_actual = get_periodo_desde_params()
    
    if request.method == 'POST':
        try:
            # Generar código automático basado en la categoría
            categoria = request.form['categoria']
            codigo = Producto.generar_codigo_automatico(categoria, periodo_actual)
            
            nombre = request.form['nombre'].strip()
            descripcion = request.form.get('descripcion', '').strip()
            unidad_medida = request.form['unidad_medida'].strip()
            precio_unitario = float(request.form.get('precio_unitario', 0))
            stock_minimo = int(request.form.get('stock_minimo', 0))
            stock_actual = int(request.form.get('stock_actual', 0))
            ubicacion = request.form.get('ubicacion', '').strip()
            fecha_vencimiento = request.form.get('fecha_vencimiento')
            lote = request.form.get('lote', '').strip()
            
            # Validaciones
            if not nombre or not categoria or not unidad_medida:
                flash('Los campos nombre, categoría y unidad de medida son obligatorios', 'error')
                return redirect(url_for('nuevo_producto_inventario', periodo=periodo_actual))
            
            # Verificar si el código ya existe en la misma categoría y período
            try:
                producto_existente = Producto.query.filter_by(
                    codigo=codigo, 
                    categoria=categoria, 
                    periodo=periodo_actual
                ).first()
            except Exception as e:
                print(f"⚠️ Error verificando código único: {e}")
                # Fallback: verificar solo por código y categoría
                producto_existente = Producto.query.filter_by(
                    codigo=codigo, 
                    categoria=categoria
                ).first()
            
            if producto_existente:
                flash(f'Ya existe un producto con el código "{codigo}" en {categoria} para el período {periodo_actual}', 'error')
                return redirect(url_for('nuevo_producto_inventario', periodo=periodo_actual))
            
            # Convertir fecha de vencimiento
            fecha_venc = None
            if fecha_vencimiento:
                fecha_venc = datetime.strptime(fecha_vencimiento, '%Y-%m-%d').date()
            
            nuevo_producto = Producto(
                codigo=codigo,
                nombre=nombre,
                descripcion=descripcion,
                categoria=categoria,
                periodo=periodo_actual,
                unidad_medida=unidad_medida,
                precio_unitario=precio_unitario,
                stock_minimo=stock_minimo,
                stock_actual=stock_actual,
                ubicacion=ubicacion,
                fecha_vencimiento=fecha_venc,
                lote=lote
            )
            
            db.session.add(nuevo_producto)
            db.session.commit()
            
            flash(f'Producto "{nombre}" creado exitosamente para {periodo_actual}', 'success')
            return redirect(url_for('productos_inventario', periodo=periodo_actual))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear el producto: {str(e)}', 'error')
            return redirect(url_for('nuevo_producto_inventario', periodo=periodo_actual))
    
    categorias_fijas = ['ALMACEN GENERAL', 'QUIMICOS', 'POSCOSECHA']
    categoria_predefinida = request.args.get('categoria', '')
    return render_template('nuevo_producto_inventario.html', 
                         categorias=categorias_fijas,
                         categoria_predefinida=categoria_predefinida,
                         periodo_actual=periodo_actual)

@app.route('/inventarios/movimientos')
@login_required
def movimientos_inventario():
    """Historial de movimientos de inventario con búsqueda avanzada"""
    # Parámetros de búsqueda y filtros
    producto_id = request.args.get('producto_id', type=int)
    busqueda = request.args.get('busqueda', '').strip()
    tipo_movimiento = request.args.get('tipo_movimiento', '')
    categoria = request.args.get('categoria', '')
    periodo = request.args.get('periodo', '')
    responsable = request.args.get('responsable', '').strip()
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')
    orden = request.args.get('orden', 'fecha_desc')
    
    # Usar eager loading para cargar productos con todos sus datos
    from sqlalchemy.orm import joinedload
    query = MovimientoInventario.query.options(joinedload(MovimientoInventario.producto)).join(Producto)
    
    # Filtro por producto específico
    if producto_id:
        query = query.filter(MovimientoInventario.producto_id == producto_id)
    
    # Búsqueda por código o nombre de producto, responsable, motivo, referencia
    if busqueda:
        search_filter = db.or_(
            Producto.codigo.ilike(f'%{busqueda}%'),
            Producto.nombre.ilike(f'%{busqueda}%'),
            MovimientoInventario.responsable.ilike(f'%{busqueda}%'),
            MovimientoInventario.motivo.ilike(f'%{busqueda}%'),
            MovimientoInventario.referencia.ilike(f'%{busqueda}%')
        )
        query = query.filter(search_filter)
    
    # Filtro por tipo de movimiento
    if tipo_movimiento:
        query = query.filter(MovimientoInventario.tipo_movimiento == tipo_movimiento)
    
    # Filtro por categoría del producto
    if categoria:
        query = query.filter(Producto.categoria == categoria)
    
    # Filtro por período
    if periodo:
        query = query.filter(MovimientoInventario.periodo == periodo)
    
    # Filtro por responsable
    if responsable:
        query = query.filter(MovimientoInventario.responsable.ilike(f'%{responsable}%'))
    
    # Filtro por rango de fechas
    if fecha_desde:
        try:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d')
            query = query.filter(MovimientoInventario.fecha_movimiento >= fecha_desde_obj)
        except ValueError:
            pass
    
    if fecha_hasta:
        try:
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d')
            # Agregar 1 día para incluir todo el día final
            fecha_hasta_obj = fecha_hasta_obj + timedelta(days=1)
            query = query.filter(MovimientoInventario.fecha_movimiento < fecha_hasta_obj)
        except ValueError:
            pass
    
    # Ordenamiento
    if orden == 'fecha_desc':
        query = query.order_by(MovimientoInventario.fecha_movimiento.desc())
    elif orden == 'fecha_asc':
        query = query.order_by(MovimientoInventario.fecha_movimiento.asc())
    elif orden == 'cantidad_desc':
        query = query.order_by(MovimientoInventario.cantidad.desc())
    elif orden == 'cantidad_asc':
        query = query.order_by(MovimientoInventario.cantidad.asc())
    elif orden == 'producto':
        query = query.order_by(Producto.nombre, MovimientoInventario.fecha_movimiento.desc())
    
    movimientos = query.all()
    productos = Producto.query.filter_by(activo=True).order_by(Producto.nombre).all()
    
    # Obtener categorías y períodos disponibles
    categorias_fijas = ['ALMACEN GENERAL', 'QUIMICOS', 'POSCOSECHA']
    try:
        periodos_disponibles = db.session.query(MovimientoInventario.periodo).distinct().order_by(MovimientoInventario.periodo.desc()).all()
        periodos_disponibles = [p[0] for p in periodos_disponibles if p[0] is not None]
    except:
        periodos_disponibles = []
    
    # Obtener responsables únicos
    try:
        responsables = db.session.query(MovimientoInventario.responsable).distinct().filter(MovimientoInventario.responsable.isnot(None)).all()
        responsables = sorted([r[0] for r in responsables if r[0] and r[0].strip()])
    except:
        responsables = []
    
    # Calcular estadísticas
    total_entradas = sum(m.cantidad for m in movimientos if m.tipo_movimiento == 'ENTRADA')
    total_salidas = sum(m.cantidad for m in movimientos if m.tipo_movimiento == 'SALIDA')
    valor_total_movimientos = sum(m.total or 0 for m in movimientos)
    
    return render_template('movimientos_inventario.html',
                         movimientos=movimientos,
                         productos=productos,
                         categorias=categorias_fijas,
                         periodos_disponibles=periodos_disponibles,
                         responsables=responsables,
                         producto_actual=producto_id,
                         busqueda_actual=busqueda,
                         tipo_actual=tipo_movimiento,
                         categoria_actual=categoria,
                         periodo_actual=periodo,
                         responsable_actual=responsable,
                         fecha_desde_actual=fecha_desde,
                         fecha_hasta_actual=fecha_hasta,
                         orden_actual=orden,
                         total_entradas=total_entradas,
                         total_salidas=total_salidas,
                         valor_total_movimientos=valor_total_movimientos)

# Rutas de categorías eliminadas - se usan categorías fijas: ALMACEN GENERAL, QUIMICOS, POSCOSECHA

@app.route('/fix-database', methods=['GET', 'POST'])
def fix_database():
    """Arreglar la base de datos después de la simplificación de inventarios"""
    if request.method == 'POST':
        try:
            from sqlalchemy import text
            
            # Usar una nueva conexión para evitar problemas de transacción
            engine = db.engine
            with engine.connect() as conn:
                # Ejecutar el script de arreglo
                print("🔧 Iniciando arreglo de base de datos...")
                
                try:
                    conn.execute(text("ALTER TABLE producto DROP CONSTRAINT IF EXISTS producto_categoria_id_fkey"))
                    print("✅ Constraint eliminado")
                except Exception as e:
                    print(f"⚠️ Error eliminando constraint: {e}")
                
                try:
                    conn.execute(text("ALTER TABLE producto DROP COLUMN IF EXISTS categoria_id"))
                    print("✅ Columna categoria_id eliminada")
                except Exception as e:
                    print(f"⚠️ Error eliminando columna: {e}")
                
                try:
                    conn.execute(text("ALTER TABLE producto ADD COLUMN IF NOT EXISTS categoria VARCHAR(50)"))
                    print("✅ Columna categoria agregada")
                except Exception as e:
                    print(f"⚠️ Error agregando columna: {e}")
                
                try:
                    conn.execute(text("DROP TABLE IF EXISTS categoria_inventario"))
                    print("✅ Tabla categoria_inventario eliminada")
                except Exception as e:
                    print(f"⚠️ Error eliminando tabla: {e}")
                
                try:
                    conn.execute(text("UPDATE producto SET categoria = 'ALMACEN GENERAL' WHERE categoria IS NULL OR categoria = ''"))
                    print("✅ Productos actualizados")
                except Exception as e:
                    print(f"⚠️ Error actualizando productos: {e}")
                
                conn.commit()
                print("🎉 Base de datos arreglada correctamente")
            
            return jsonify({'success': True, 'message': 'Base de datos arreglada correctamente'})
            
        except Exception as e:
            print(f"❌ Error al arreglar la base de datos: {str(e)}")
            return jsonify({'success': False, 'message': f'Error: {str(e)}'})
    
    # GET: Mostrar página de arreglo
    return '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Arreglar Base de Datos</title>
        <style>
            body { font-family: Arial, sans-serif; max-width: 600px; margin: 50px auto; padding: 20px; }
            .btn { background: #007bff; color: white; padding: 10px 20px; border: none; border-radius: 5px; cursor: pointer; }
            .btn:hover { background: #0056b3; }
            .result { margin-top: 20px; padding: 10px; border-radius: 5px; }
            .success { background: #d4edda; color: #155724; border: 1px solid #c3e6cb; }
            .error { background: #f8d7da; color: #721c24; border: 1px solid #f5c6cb; }
        </style>
    </head>
    <body>
        <h1>🔧 Arreglar Base de Datos</h1>
        <p>Esta herramienta arreglará la base de datos después de la simplificación del sistema de inventarios.</p>
        <button class="btn" onclick="fixDatabase()">Arreglar Base de Datos</button>
        <div id="result"></div>
        
        <script>
        async function fixDatabase() {
            const resultDiv = document.getElementById('result');
            resultDiv.innerHTML = '<p>Arreglando base de datos...</p>';
            
            try {
                const response = await fetch('/fix-database', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' }
                });
                
                const data = await response.json();
                
                if (data.success) {
                    resultDiv.innerHTML = `<div class="result success">✅ ${data.message}</div>`;
                } else {
                    resultDiv.innerHTML = `<div class="result error">❌ ${data.message}</div>`;
                }
            } catch (error) {
                resultDiv.innerHTML = `<div class="result error">❌ Error: ${error.message}</div>`;
            }
        }
        </script>
    </body>
    </html>
    '''

@app.route('/inventarios/productos/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_producto_inventario(id):
    """Editar producto de inventario"""
    producto = Producto.query.get_or_404(id)
    
    # Verificar si el mes está cerrado (solo admins pueden editar meses cerrados)
    if hasattr(producto, 'mes_cerrado') and producto.mes_cerrado and not current_user.is_admin:
        flash('❌ No se puede editar este producto. El período está cerrado. Contacte al administrador.', 'error')
        return redirect(url_for('productos_inventario', categoria=producto.categoria))
    
    if request.method == 'POST':
        try:
            codigo = request.form['codigo'].strip().upper()
            nombre = request.form['nombre'].strip().upper()
            descripcion = request.form.get('descripcion', '').strip()
            categoria = request.form['categoria']
            unidad_medida = request.form['unidad_medida']
            precio_unitario = float(request.form.get('precio_unitario', 0))
            stock_minimo = int(request.form.get('stock_minimo', 0))
            stock_actual = int(request.form.get('stock_actual', 0))
            saldo_inicial = int(request.form.get('saldo_inicial', 0))
            ubicacion = request.form.get('ubicacion', '').strip()
            fecha_vencimiento = request.form.get('fecha_vencimiento')
            lote = request.form.get('lote', '').strip()
            activo = 'activo' in request.form
            
            # Validaciones
            if not codigo or not nombre or not categoria:
                flash('Código, nombre y categoría son obligatorios', 'error')
                return redirect(url_for('editar_producto_inventario', id=id))
            
            # Verificar si el código ya existe (excluyendo el actual)
            producto_existente = Producto.query.filter(
                Producto.codigo == codigo,
                Producto.id != id
            ).first()
            
            if producto_existente:
                flash(f'Ya existe un producto con el código "{codigo}"', 'error')
                return redirect(url_for('editar_producto_inventario', id=id))
            
            # Actualizar producto
            producto.codigo = codigo
            producto.nombre = nombre
            producto.descripcion = descripcion
            producto.categoria = categoria
            producto.unidad_medida = unidad_medida
            producto.precio_unitario = precio_unitario
            producto.stock_minimo = stock_minimo
            producto.stock_actual = stock_actual
            producto.saldo_inicial = saldo_inicial
            producto.ubicacion = ubicacion
            producto.lote = lote
            producto.activo = activo
            
            if fecha_vencimiento:
                producto.fecha_vencimiento = datetime.strptime(fecha_vencimiento, '%Y-%m-%d').date()
            else:
                producto.fecha_vencimiento = None
            
            # Solo recalcular si no se está editando el stock manualmente
            # El stock_actual ya fue actualizado con el valor del formulario
            # producto.recalcular_stock()  # Comentado para permitir edición manual
            
            db.session.commit()
            
            flash(f'Producto "{nombre}" actualizado exitosamente. Stock: {producto.stock_actual}', 'success')
            return redirect(url_for('productos_inventario'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar el producto: {str(e)}', 'error')
            return redirect(url_for('editar_producto_inventario', id=id))
    
    # GET: Mostrar formulario de edición
    categorias_fijas = ['ALMACEN GENERAL', 'QUIMICOS', 'POSCOSECHA']
    return render_template('editar_producto_inventario.html', producto=producto, categorias=categorias_fijas)

@app.route('/inventarios/productos/eliminar/<int:id>', methods=['POST'])
@login_required
def eliminar_producto_inventario(id):
    """Eliminar producto de inventario"""
    if not current_user.is_admin:
        flash('Solo los administradores pueden eliminar productos', 'error')
        return redirect(url_for('productos_inventario'))
    
    try:
        producto = Producto.query.get_or_404(id)
        
        # Verificar si tiene movimientos asociados
        movimientos_count = MovimientoInventario.query.filter_by(producto_id=id).count()
        if movimientos_count > 0:
            flash(f'No se puede eliminar el producto "{producto.nombre}" porque tiene {movimientos_count} movimientos asociados', 'error')
            return redirect(url_for('productos_inventario'))
        
        # Eliminar producto
        db.session.delete(producto)
        db.session.commit()
        
        flash(f'Producto "{producto.nombre}" eliminado exitosamente', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar el producto: {str(e)}', 'error')
    
    return redirect(url_for('productos_inventario'))

@app.route('/inventarios/productos/activar/<int:id>', methods=['POST'])
@login_required
def activar_producto_inventario(id):
    """Activar producto de inventario"""
    try:
        producto = Producto.query.get_or_404(id)
        producto.activo = True
        db.session.commit()
        
        flash(f'Producto "{producto.nombre}" activado exitosamente', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al activar el producto: {str(e)}', 'error')
    
    return redirect(url_for('productos_inventario'))

@app.route('/inventarios/productos/desactivar/<int:id>', methods=['POST'])
@login_required
def desactivar_producto_inventario(id):
    """Desactivar producto de inventario"""
    try:
        producto = Producto.query.get_or_404(id)
        producto.activo = False
        db.session.commit()
        
        flash(f'Producto "{producto.nombre}" desactivado exitosamente', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al desactivar el producto: {str(e)}', 'error')
    
    return redirect(url_for('productos_inventario'))

@app.route('/inventarios/copiar-mes-anterior', methods=['GET', 'POST'])
@login_required
def copiar_inventario_mes_anterior():
    """Copiar inventario del mes anterior al mes actual"""
    if not current_user.is_admin:
        flash('Solo los administradores pueden copiar inventarios', 'error')
        return redirect(url_for('inventarios'))
    
    if request.method == 'POST':
        try:
            # Obtener mes anterior
            periodo_actual = get_periodo_actual()
            año_actual = int(periodo_actual.split('-')[0])
            mes_actual = int(periodo_actual.split('-')[1])
            
            # Calcular mes anterior
            if mes_actual == 1:
                mes_anterior = 12
                año_anterior = año_actual - 1
            else:
                mes_anterior = mes_actual - 1
                año_anterior = año_actual
            
            periodo_anterior = f"{año_anterior:04d}-{mes_anterior:02d}"
            
            # Obtener productos del mes anterior
            productos_anteriores = Producto.query.filter_by(periodo=periodo_anterior, activo=True).all()
            
            if not productos_anteriores:
                flash(f'No hay productos en el período {periodo_anterior} para copiar', 'warning')
                return redirect(url_for('inventarios'))
            
            # Copiar productos al mes actual
            productos_copiados = 0
            productos_actualizados = 0
            for producto in productos_anteriores:
                # El saldo final del mes anterior se convierte en saldo inicial del nuevo mes
                saldo_final_anterior = producto.calcular_saldo_final()
                
                # Verificar si ya existe un producto con el mismo código en el período actual
                producto_existente = Producto.query.filter_by(
                    codigo=producto.codigo,
                    periodo=periodo_actual
                ).first()
                
                if producto_existente:
                    # Si existe en el período actual, actualizar
                    producto_existente.nombre = producto.nombre
                    producto_existente.descripcion = producto.descripcion
                    producto_existente.categoria = producto.categoria
                    producto_existente.unidad_medida = producto.unidad_medida
                    producto_existente.precio_unitario = producto.precio_unitario
                    producto_existente.stock_minimo = producto.stock_minimo
                    producto_existente.saldo_inicial = saldo_final_anterior
                    producto_existente.stock_actual = saldo_final_anterior
                    producto_existente.ubicacion = producto.ubicacion
                    producto_existente.proveedor = producto.proveedor
                    producto_existente.activo = True
                    producto_existente.mes_cerrado = False
                    productos_actualizados += 1
                else:
                    # Verificar si existe con ese código en otro período
                    producto_otro_periodo = Producto.query.filter_by(codigo=producto.codigo).first()
                    
                    if producto_otro_periodo:
                        # Si existe en otro período, actualizar su período al actual
                        producto_otro_periodo.nombre = producto.nombre
                        producto_otro_periodo.descripcion = producto.descripcion
                        producto_otro_periodo.categoria = producto.categoria
                        producto_otro_periodo.periodo = periodo_actual
                        producto_otro_periodo.unidad_medida = producto.unidad_medida
                        producto_otro_periodo.precio_unitario = producto.precio_unitario
                        producto_otro_periodo.stock_minimo = producto.stock_minimo
                        producto_otro_periodo.saldo_inicial = saldo_final_anterior
                        producto_otro_periodo.stock_actual = saldo_final_anterior
                        producto_otro_periodo.ubicacion = producto.ubicacion
                        producto_otro_periodo.proveedor = producto.proveedor
                        producto_otro_periodo.activo = True
                        producto_otro_periodo.mes_cerrado = False
                        productos_actualizados += 1
                    else:
                        # Si no existe, crear nuevo producto
                        nuevo_producto = Producto(
                            codigo=producto.codigo,
                            nombre=producto.nombre,
                            descripcion=producto.descripcion,
                            categoria=producto.categoria,
                            periodo=periodo_actual,
                            unidad_medida=producto.unidad_medida,
                            precio_unitario=producto.precio_unitario,
                            stock_minimo=producto.stock_minimo,
                            saldo_inicial=saldo_final_anterior,  # ✅ Saldo final del mes anterior
                            stock_actual=saldo_final_anterior,   # ✅ Stock actual = saldo inicial (sin movimientos aún)
                            ubicacion=producto.ubicacion,
                            proveedor=producto.proveedor,
                            activo=True,
                            mes_cerrado=False  # El nuevo mes está abierto
                        )
                        db.session.add(nuevo_producto)
                        productos_copiados += 1
            
            # Cerrar el mes anterior para evitar modificaciones
            for producto in productos_anteriores:
                producto.mes_cerrado = True
            
            db.session.commit()
            
            mensaje = f'✅ Proceso completado: '
            if productos_copiados > 0:
                mensaje += f'{productos_copiados} productos copiados. '
            if productos_actualizados > 0:
                mensaje += f'{productos_actualizados} productos actualizados. '
            mensaje += f'Saldos iniciales configurados automáticamente. Mes {periodo_anterior} cerrado.'
            flash(mensaje, 'success')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error copiando inventario: {str(e)}', 'error')
        
        return redirect(url_for('inventarios'))
    
    # GET: Mostrar página de confirmación
    periodo_actual = get_periodo_actual()
    año_actual = int(periodo_actual.split('-')[0])
    mes_actual = int(periodo_actual.split('-')[1])
    
    # Calcular mes anterior
    if mes_actual == 1:
        mes_anterior = 12
        año_anterior = año_actual - 1
    else:
        mes_anterior = mes_actual - 1
        año_anterior = año_actual
    
    periodo_anterior = f"{año_anterior:04d}-{mes_anterior:02d}"
    
    # Verificar si ya existen productos en el mes actual
    productos_actuales = Producto.query.filter_by(periodo=periodo_actual).count()
    productos_anteriores = Producto.query.filter_by(periodo=periodo_anterior, activo=True).count()
    
    return render_template('copiar_inventario.html',
                         periodo_actual=periodo_actual,
                         periodo_anterior=periodo_anterior,
                         productos_actuales=productos_actuales,
                         productos_anteriores=productos_anteriores)

@app.route('/inventarios/reportes')
@login_required
def reportes_inventarios():
    """Generar reportes mensuales de inventarios"""
    if not current_user.is_admin:
        flash('Solo los administradores pueden generar reportes', 'error')
        return redirect(url_for('inventarios'))
    
    # Obtener período desde parámetros o usar el actual
    periodo = request.args.get('periodo', get_periodo_actual())
    
    # Validar formato del período
    try:
        datetime.strptime(periodo, '%Y-%m')
    except ValueError:
        periodo = get_periodo_actual()
    
    # Obtener productos del período
    productos = Producto.query.filter_by(periodo=periodo, activo=True).all()
    
    # Estadísticas por categoría
    categorias_fijas = ['ALMACEN GENERAL', 'QUIMICOS', 'POSCOSECHA']
    stats_por_categoria = {}
    total_valor = 0
    total_productos = 0
    productos_bajo_stock = 0
    
    for categoria in categorias_fijas:
        productos_cat = [p for p in productos if p.categoria == categoria]
        cantidad_productos = len(productos_cat)
        cantidad_bajo_stock = len([p for p in productos_cat if p.stock_actual <= p.stock_minimo])
        valor_categoria = sum(p.stock_actual * p.precio_unitario for p in productos_cat)
        
        stats_por_categoria[categoria] = {
            'productos': productos_cat,
            'total_productos': cantidad_productos,
            'productos_bajo_stock': cantidad_bajo_stock,
            'valor_total': valor_categoria,
            'stock_total': sum(p.stock_actual for p in productos_cat)
        }
        
        total_valor += valor_categoria
        total_productos += cantidad_productos
        productos_bajo_stock += cantidad_bajo_stock
    
    # Obtener períodos disponibles para el selector
    try:
        periodos_disponibles = db.session.query(Producto.periodo).distinct().order_by(Producto.periodo.desc()).all()
        periodos_disponibles = [p[0] for p in periodos_disponibles if p[0] is not None]
    except Exception as e:
        print(f"⚠️ Error obteniendo períodos: {e}")
        periodos_disponibles = [periodo]
    
    return render_template('reportes_inventarios.html',
                         periodo=periodo,
                         periodos_disponibles=periodos_disponibles,
                         stats_por_categoria=stats_por_categoria,
                         total_productos=total_productos,
                         total_valor=total_valor,
                         productos_bajo_stock=productos_bajo_stock)

@app.route('/inventarios/procedimientos/cerrar-mes/<periodo>', methods=['POST'])
@login_required
def cerrar_mes_procedimiento(periodo):
    """Cerrar mes usando procedimiento almacenado"""
    if not current_user.is_admin:
        flash('Solo los administradores pueden cerrar meses', 'error')
        return redirect(url_for('inventarios'))
    
    try:
        from sqlalchemy import text
        result = db.session.execute(text("SELECT * FROM cerrar_mes_inventario(:periodo)"), 
                                   {'periodo': periodo})
        resultado = result.fetchone()
        
        if resultado:
            flash(resultado[2], 'success')  # mensaje
        else:
            flash(f'Mes {periodo} cerrado exitosamente', 'success')
        
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'Error al cerrar mes: {str(e)}', 'error')
    
    return redirect(url_for('inventarios'))

@app.route('/inventarios/procedimientos/abrir-mes/<periodo>', methods=['POST'])
@login_required
def abrir_mes_procedimiento(periodo):
    """Abrir nuevo mes usando procedimiento almacenado"""
    if not current_user.is_admin:
        flash('Solo los administradores pueden abrir nuevos meses', 'error')
        return redirect(url_for('inventarios'))
    
    try:
        from sqlalchemy import text
        result = db.session.execute(text("SELECT * FROM abrir_nuevo_mes_inventario(:periodo)"), 
                                   {'periodo': periodo})
        resultado = result.fetchone()
        
        if resultado:
            flash(resultado[3], 'success' if '✅' in resultado[3] else 'error')  # mensaje
        else:
            flash(f'Mes {periodo} abierto exitosamente', 'success')
        
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'Error al abrir mes: {str(e)}', 'error')
    
    return redirect(url_for('inventarios'))

@app.route('/inventarios/procedimientos/recalcular-stocks/<periodo>')
@login_required
def recalcular_stocks_procedimiento(periodo):
    """Recalcular stocks usando procedimiento almacenado"""
    if not current_user.is_admin:
        flash('Solo los administradores pueden recalcular stocks', 'error')
        return redirect(url_for('inventarios'))
    
    try:
        from sqlalchemy import text
        result = db.session.execute(text("SELECT * FROM recalcular_stocks(:periodo)"), 
                                   {'periodo': periodo})
        resultado = result.fetchone()
        
        if resultado:
            flash(resultado[2], 'success')  # mensaje
        else:
            flash(f'Stocks del período {periodo} recalculados', 'success')
        
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'Error al recalcular stocks: {str(e)}', 'error')
    
    return redirect(url_for('productos_inventario'))

@app.route('/inventarios/procedimientos/reporte-stock-bajo/<periodo>')
@login_required
def reporte_stock_bajo_procedimiento(periodo):
    """Ver reporte de stock bajo usando procedimiento almacenado"""
    try:
        from sqlalchemy import text
        result = db.session.execute(text("SELECT * FROM reporte_stock_bajo(:periodo)"), 
                                   {'periodo': periodo})
        productos_bajo = result.fetchall()
        
        # Convertir a lista de diccionarios
        productos = []
        for p in productos_bajo:
            productos.append({
                'codigo': p[0],
                'nombre': p[1],
                'categoria': p[2],
                'stock_actual': p[3],
                'stock_minimo': p[4],
                'diferencia': p[5],
                'proveedor': p[6]
            })
        
        return render_template('reporte_stock_bajo.html', 
                             productos=productos,
                             periodo=periodo)
    except Exception as e:
        flash(f'Error al generar reporte: {str(e)}', 'error')
        return redirect(url_for('inventarios'))

@app.route('/inventarios/procedimientos/estadisticas/<periodo>')
@login_required
def estadisticas_mes_procedimiento(periodo):
    """Ver estadísticas del mes usando procedimiento almacenado"""
    try:
        from sqlalchemy import text
        result = db.session.execute(text("SELECT * FROM estadisticas_mes(:periodo)"), 
                                   {'periodo': periodo})
        stats = result.fetchone()
        
        if stats:
            estadisticas = {
                'total_productos': stats[0],
                'productos_activos': stats[1],
                'productos_stock_bajo': stats[2],
                'total_entradas': stats[3],
                'total_salidas': stats[4],
                'valor_total_inventario': float(stats[5]),
                'mes_cerrado': stats[6]
            }
        else:
            estadisticas = None
        
        return render_template('estadisticas_mes.html',
                             estadisticas=estadisticas,
                             periodo=periodo)
    except Exception as e:
        flash(f'Error al obtener estadísticas: {str(e)}', 'error')
        return redirect(url_for('inventarios'))

@app.route('/inventarios/procedimientos/auditoria/<periodo>')
@login_required
def auditoria_movimientos_procedimiento(periodo):
    """Ver auditoría de movimientos usando procedimiento almacenado"""
    if not current_user.is_admin:
        flash('Solo los administradores pueden ver auditorías', 'error')
        return redirect(url_for('inventarios'))
    
    try:
        from sqlalchemy import text
        result = db.session.execute(text("SELECT * FROM auditoria_movimientos(:periodo)"), 
                                   {'periodo': periodo})
        movimientos = result.fetchall()
        
        # Convertir a lista de diccionarios
        auditoria = []
        for m in movimientos:
            auditoria.append({
                'movimiento_id': m[0],
                'producto_codigo': m[1],
                'producto_nombre': m[2],
                'tipo_movimiento': m[3],
                'cantidad': m[4],
                'fecha_movimiento': m[5],
                'problema': m[6]
            })
        
        return render_template('auditoria_movimientos.html',
                             auditoria=auditoria,
                             periodo=periodo)
    except Exception as e:
        flash(f'Error al generar auditoría: {str(e)}', 'error')
        return redirect(url_for('inventarios'))

@app.route('/inventarios/diagnostico-precios/<periodo>')
@login_required
def diagnostico_precios(periodo):
    """Diagnóstico de productos con precios anormales"""
    if not current_user.is_admin:
        flash('Solo los administradores pueden ver diagnósticos', 'error')
        return redirect(url_for('inventarios'))
    
    # Productos con precios muy altos (probablemente errores)
    productos_precio_alto = Producto.query.filter(
        Producto.periodo == periodo,
        Producto.precio_unitario >= 10000000  # > $10M
    ).order_by(Producto.precio_unitario.desc()).all()
    
    # Productos con precios en 0
    productos_sin_precio = Producto.query.filter(
        Producto.periodo == periodo,
        Producto.activo == True,
        db.or_(Producto.precio_unitario == 0, Producto.precio_unitario == None)
    ).order_by(Producto.stock_actual.desc()).all()
    
    # Productos con mayor valor en inventario
    productos_mayor_valor = Producto.query.filter(
        Producto.periodo == periodo,
        Producto.activo == True,
        Producto.precio_unitario < 10000000  # Excluir anormales
    ).order_by((Producto.stock_actual * Producto.precio_unitario).desc()).limit(20).all()
    
    # Calcular valor total real vs valor con anormales
    from sqlalchemy import text
    result = db.session.execute(
        text("""
            SELECT 
                SUM(CASE WHEN precio_unitario < 10000000 THEN stock_actual * precio_unitario ELSE 0 END) as valor_real,
                SUM(CASE WHEN precio_unitario >= 10000000 THEN stock_actual * precio_unitario ELSE 0 END) as valor_anormal,
                SUM(stock_actual * precio_unitario) as valor_total
            FROM producto 
            WHERE periodo = :periodo AND activo = TRUE
        """),
        {'periodo': periodo}
    )
    valores = result.fetchone()
    
    return render_template('diagnostico_precios.html',
                         periodo=periodo,
                         productos_precio_alto=productos_precio_alto,
                         productos_sin_precio=productos_sin_precio,
                         productos_mayor_valor=productos_mayor_valor,
                         valor_real=float(valores[0]) if valores[0] else 0,
                         valor_anormal=float(valores[1]) if valores[1] else 0,
                         valor_total=float(valores[2]) if valores[2] else 0)

@app.route('/inventarios/corregir-precio/<int:producto_id>', methods=['POST'])
@login_required
def corregir_precio_producto(producto_id):
    """Corregir precio de un producto"""
    if not current_user.is_admin:
        flash('Solo los administradores pueden corregir precios', 'error')
        return redirect(url_for('inventarios'))
    
    try:
        producto = Producto.query.get_or_404(producto_id)
        nuevo_precio = float(request.form.get('nuevo_precio', 0))
        
        precio_anterior = producto.precio_unitario
        producto.precio_unitario = nuevo_precio
        
        db.session.commit()
        
        flash(f'✅ Precio de "{producto.nombre}" actualizado: ${precio_anterior:,.0f} → ${nuevo_precio:,.0f}', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al corregir precio: {str(e)}', 'error')
    
    return redirect(request.referrer or url_for('inventarios'))

@app.route('/inventarios/exportar-excel/<periodo>')
@login_required
def exportar_excel_inventario(periodo):
    """Exportar inventario a Excel con fórmulas automáticas"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from flask import send_file
        import io
        import tempfile
        import os
        
        # Obtener productos del período con eager loading de relaciones
        from sqlalchemy.orm import joinedload
        productos = Producto.query.options(joinedload(Producto.movimientos)).filter_by(periodo=periodo, activo=True).all()
        
        if not productos:
            flash(f'No hay productos en el período {periodo}', 'warning')
            return redirect(url_for('inventarios'))
        
        # Crear workbook
        wb = Workbook()
        wb.remove(wb.active)  # Eliminar hoja por defecto
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Agrupar productos por categoría
        categorias = {}
        for producto in productos:
            if producto.categoria not in categorias:
                categorias[producto.categoria] = []
            categorias[producto.categoria].append(producto)
        
        # Crear hoja para cada categoría
        for categoria, productos_cat in categorias.items():
            ws = wb.create_sheet(title=categoria)
            
            # Calcular número máximo de entradas y salidas por separado
            max_entradas = 0
            max_salidas = 0
            for producto in productos_cat:
                entradas_count = len([m for m in producto.movimientos if m.tipo_movimiento == 'ENTRADA'])
                salidas_count = len([m for m in producto.movimientos if m.tipo_movimiento == 'SALIDA'])
                if entradas_count > max_entradas:
                    max_entradas = entradas_count
                if salidas_count > max_salidas:
                    max_salidas = salidas_count
            
            # Calcular número total de columnas
            headers_count = 3  # NOMBRE, SALDO REAL, VALOR TOTAL
            entrada_cols = 6  # Fecha, Factura, Cantidad, Valor Unit, Total, Proveedor
            salida_cols = 3   # Fecha, Cantidad, Unidad
            separator_col_count = 2  # Separador antes de entradas y antes de salidas
            total_cols = headers_count + separator_col_count + (max_entradas * entrada_cols) + (max_salidas * salida_cols)
            
            # Obtener letra de última columna
            from openpyxl.utils import get_column_letter
            last_col_letter = get_column_letter(total_cols)
            
            # Lista para almacenar todos los merges y hacerlos al final
            merges_to_do = []
            
            # Título de la hoja (escribir ANTES del merge)
            title_cell = ws['A1']
            title_cell.value = f'INVENTARIO {categoria} - PERÍODO {periodo}'
            title_cell.font = Font(bold=True, size=16)
            title_cell.alignment = center_alignment
            merges_to_do.append(f'A1:{last_col_letter}1')
            
            # Fecha de generación (escribir ANTES del merge)
            fecha_cell = ws['A2']
            fecha_cell.value = f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
            fecha_cell.font = Font(italic=True)
            merges_to_do.append(f'A2:{last_col_letter}2')
            
            # SECCIÓN 1: RESUMEN DE PRODUCTOS (escribir ANTES del merge)
            resumen_cell = ws['A4']
            resumen_cell.value = 'RESUMEN DE PRODUCTOS'
            resumen_cell.font = Font(bold=True, size=14)
            merges_to_do.append(f'A4:{last_col_letter}4')
            
            # Encabezados principales - solo datos básicos del producto
            headers_resumen = ['NOMBRE', 'SALDO REAL', 'VALOR TOTAL']
            header_cols = ['A', 'B', 'C']
            
            for col_idx, header in enumerate(headers_resumen):
                col_letter = header_cols[col_idx]
                cell = ws[f'{col_letter}5']
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
                # Altura de fila para encabezados
                ws.row_dimensions[5].height = 25
            
            # Encabezados de ENTRADAS y SALIDAS separadas
            start_entradas_col = len(headers_resumen) + 1
            separator1_col = start_entradas_col - 1
            separator1_col_letter = get_column_letter(separator1_col)
            
            # Calcular dónde empiezan las salidas
            start_salidas_col = start_entradas_col + (max_entradas * entrada_cols) + 1
            separator2_col = start_salidas_col - 1
            separator2_col_letter = get_column_letter(separator2_col)
            
            merge_ranges = []
            
            # Encabezados de ENTRADAS
            for ent_idx in range(max_entradas):
                start_col = start_entradas_col + (ent_idx * entrada_cols)
                end_col = start_col + entrada_cols - 1
                
                start_col_letter = get_column_letter(start_col)
                end_col_letter = get_column_letter(end_col)
                
                # Encabezado principal "ENTRADA X"
                header_cell = ws[f'{start_col_letter}4']
                header_cell.value = f'ENTRADA {ent_idx + 1}'
                header_cell.font = Font(bold=True, size=10)
                header_cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
                header_cell.alignment = center_alignment
                header_cell.border = border
                merge_ranges.append(f'{start_col_letter}4:{end_col_letter}4')
                
                # Sub-encabezados: Fecha, Factura, Cantidad, Proveedor, Valor Unit, Total
                sub_headers = ['FECHA', 'FACTURA', 'CANTIDAD', 'PROVEEDOR', 'VALOR UNIT.', 'TOTAL']
                for sub_idx, sub_header in enumerate(sub_headers):
                    sub_col_letter = get_column_letter(start_col + sub_idx)
                    sub_cell = ws[f'{sub_col_letter}5']
                    sub_cell.value = sub_header
                    sub_cell.font = header_font
                    sub_cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
                    sub_cell.alignment = center_alignment
                    sub_cell.border = border
            
            # Encabezados de SALIDAS
            for sal_idx in range(max_salidas):
                start_col = start_salidas_col + (sal_idx * salida_cols)
                end_col = start_col + salida_cols - 1
                
                start_col_letter = get_column_letter(start_col)
                end_col_letter = get_column_letter(end_col)
                
                # Encabezado principal "SALIDA X"
                header_cell = ws[f'{start_col_letter}4']
                header_cell.value = f'SALIDA {sal_idx + 1}'
                header_cell.font = Font(bold=True, size=10)
                header_cell.fill = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")
                header_cell.alignment = center_alignment
                header_cell.border = border
                merge_ranges.append(f'{start_col_letter}4:{end_col_letter}4')
                
                # Sub-encabezados: Fecha, Cantidad, Unidad
                sub_headers = ['FECHA', 'CANTIDAD', 'UNIDAD']
                for sub_idx, sub_header in enumerate(sub_headers):
                    sub_col_letter = get_column_letter(start_col + sub_idx)
                    sub_cell = ws[f'{sub_col_letter}5']
                    sub_cell.value = sub_header
                    sub_cell.font = header_font
                    sub_cell.fill = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")
                    sub_cell.alignment = center_alignment
                    sub_cell.border = border
            
            # Agregar columnas separadoras en encabezados (fila 4 y 5)
            separator1_header_4 = ws[f'{separator1_col_letter}4']
            separator1_header_4.value = ''
            separator1_header_4.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            separator1_header_4.border = border
            
            separator1_header_5 = ws[f'{separator1_col_letter}5']
            separator1_header_5.value = ''
            separator1_header_5.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            separator1_header_5.border = border
            
            separator2_header_4 = ws[f'{separator2_col_letter}4']
            separator2_header_4.value = ''
            separator2_header_4.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            separator2_header_4.border = border
            
            separator2_header_5 = ws[f'{separator2_col_letter}5']
            separator2_header_5.value = ''
            separator2_header_5.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            separator2_header_5.border = border
            
            # Agregar merges a la lista
            merges_to_do.extend(merge_ranges)
            
            # Datos de productos con mejor formato y organización
            from openpyxl.utils import get_column_letter
            fill_even = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")  # Gris claro para filas pares
            fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Blanco para filas impares
            
            for row_idx, producto in enumerate(productos_cat, 6):
                row = row_idx
                # Alternar colores de fila para mejor legibilidad
                row_fill = fill_even if row_idx % 2 == 0 else fill_odd
                
                # NOMBRE (solo nombre, sin código ni unidad)
                nombre_cell = ws[f'A{row}']
                nombre_cell.value = producto.nombre
                nombre_cell.border = border
                nombre_cell.fill = row_fill
                
                # Calcular saldo real desde movimientos
                saldo_inicial = getattr(producto, 'saldo_inicial', 0) or 0
                entradas = sum(m.calcular_cantidad_total() for m in producto.movimientos if m.tipo_movimiento == 'ENTRADA')
                salidas = sum(m.calcular_cantidad_total() for m in producto.movimientos if m.tipo_movimiento == 'SALIDA')
                saldo_real = saldo_inicial + entradas - salidas
                
                # SALDO REAL (valor calculado)
                saldo_real_cell = ws[f'B{row}']
                saldo_real_cell.value = saldo_real
                saldo_real_cell.border = border
                saldo_real_cell.fill = row_fill
                saldo_real_cell.alignment = center_alignment
                saldo_real_cell.font = Font(bold=True)
                
                # VALOR TOTAL (solo para productos que deben tener precio)
                if producto.debe_tener_precio():
                    precio = float(producto.precio_unitario) if producto.precio_unitario else 0
                    valor_total = saldo_real * precio
                    valor_cell = ws[f'C{row}']
                    valor_cell.value = valor_total
                    valor_cell.number_format = '#,##0'
                    valor_cell.border = border
                    valor_cell.fill = row_fill
                    valor_cell.alignment = center_alignment
                    valor_cell.font = Font(bold=True)
                else:
                    # Para ALMACEN GENERAL, mostrar "-"
                    valor_cell = ws[f'C{row}']
                    valor_cell.value = '-'
                    valor_cell.border = border
                    valor_cell.fill = row_fill
                    valor_cell.alignment = center_alignment
                
                # Separadores
                separator1_cell = ws[f'{separator1_col_letter}{row}']
                separator1_cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                separator1_cell.border = border
                
                separator2_cell = ws[f'{separator2_col_letter}{row}']
                separator2_cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                separator2_cell.border = border
                
                # Obtener entradas y salidas separadas y ordenadas por fecha
                entradas_producto = sorted([m for m in producto.movimientos if m.tipo_movimiento == 'ENTRADA'], 
                                         key=lambda x: x.fecha_movimiento)
                salidas_producto = sorted([m for m in producto.movimientos if m.tipo_movimiento == 'SALIDA'], 
                                         key=lambda x: x.fecha_movimiento)
                
                # Escribir ENTRADAS
                for ent_idx, entrada in enumerate(entradas_producto):
                    start_col = start_entradas_col + (ent_idx * entrada_cols)
                    
                    # Fecha
                    fecha_cell = ws[f'{get_column_letter(start_col)}{row}']
                    fecha_cell.value = entrada.fecha_movimiento.strftime('%d/%m/%Y')
                    fecha_cell.border = border
                    fecha_cell.alignment = center_alignment
                    
                    # Factura (referencia)
                    factura_cell = ws[f'{get_column_letter(start_col + 1)}{row}']
                    factura_cell.value = entrada.referencia or ''
                    factura_cell.border = border
                    
                    # Cantidad con unidad
                    cantidad_total = entrada.calcular_cantidad_total()
                    unidad_entrada = producto.unidad_medida.upper()
                    if unidad_entrada in ['L', 'LITRO', 'LITROS']:
                        unidad_display = 'L'
                    elif unidad_entrada in ['KG', 'KILO', 'KILOS', 'KILOGRAMO', 'KILOGRAMOS']:
                        unidad_display = 'KG'
                    elif unidad_entrada in ['G', 'GRAMO', 'GRAMOS']:
                        unidad_display = 'G'
                    elif unidad_entrada in ['ML', 'MILILITRO', 'MILILITROS']:
                        unidad_display = 'ML'
                    elif unidad_entrada in ['CC', 'CENTIMETRO CUBICO', 'CENTIMETROS CUBICOS']:
                        unidad_display = 'CC'
                    else:
                        unidad_display = producto.unidad_medida
                    
                    cantidad_cell = ws[f'{get_column_letter(start_col + 2)}{row}']
                    cantidad_cell.value = f'{cantidad_total} {unidad_display}'
                    cantidad_cell.border = border
                    cantidad_cell.alignment = center_alignment
                    
                    # Proveedor (antes de Valor Unit) - leer directamente del producto del loop principal
                    proveedor_cell = ws[f'{get_column_letter(start_col + 3)}{row}']
                    # Usar el producto del loop principal que ya tenemos cargado
                    proveedor_valor = ''
                    if producto.proveedor:
                        proveedor_valor = str(producto.proveedor).strip()
                    proveedor_cell.value = proveedor_valor
                    proveedor_cell.border = border
                    
                    # Valor Unitario
                    valor_unit_cell = ws[f'{get_column_letter(start_col + 4)}{row}']
                    if entrada.precio_unitario:
                        valor_unit_cell.value = float(entrada.precio_unitario)
                        valor_unit_cell.number_format = '#,##0'
                    else:
                        valor_unit_cell.value = 0
                    valor_unit_cell.border = border
                    valor_unit_cell.alignment = center_alignment
                    
                    # Total
                    total_cell = ws[f'{get_column_letter(start_col + 5)}{row}']
                    if entrada.total:
                        total_cell.value = float(entrada.total)
                        total_cell.number_format = '#,##0'
                    else:
                        total_cell.value = 0
                    total_cell.border = border
                    total_cell.alignment = center_alignment
                    total_cell.font = Font(bold=True)
                
                # Escribir SALIDAS
                for sal_idx, salida in enumerate(salidas_producto):
                    start_col = start_salidas_col + (sal_idx * salida_cols)
                    
                    # Unidad estandarizada del producto
                    unidad = producto.unidad_medida.upper()
                    if unidad in ['L', 'LITRO', 'LITROS']:
                        unidad_display = 'L'
                    elif unidad in ['KG', 'KILO', 'KILOS', 'KILOGRAMO', 'KILOGRAMOS']:
                        unidad_display = 'KG'
                    elif unidad in ['G', 'GRAMO', 'GRAMOS']:
                        unidad_display = 'G'
                    elif unidad in ['ML', 'MILILITRO', 'MILILITROS']:
                        unidad_display = 'ML'
                    elif unidad in ['CC', 'CENTIMETRO CUBICO', 'CENTIMETROS CUBICOS']:
                        unidad_display = 'CC'
                    else:
                        unidad_display = producto.unidad_medida
                    
                    # Fecha
                    fecha_cell = ws[f'{get_column_letter(start_col)}{row}']
                    fecha_cell.value = salida.fecha_movimiento.strftime('%d/%m/%Y')
                    fecha_cell.border = border
                    fecha_cell.alignment = center_alignment
                    
                    # Cantidad (solo número)
                    cantidad_total = salida.calcular_cantidad_total()
                    cantidad_cell = ws[f'{get_column_letter(start_col + 1)}{row}']
                    cantidad_cell.value = cantidad_total
                    cantidad_cell.border = border
                    cantidad_cell.alignment = center_alignment
                    
                    # Unidad (en columna separada)
                    unidad_cell = ws[f'{get_column_letter(start_col + 2)}{row}']
                    unidad_cell.value = unidad_display
                    unidad_cell.border = border
                    unidad_cell.alignment = center_alignment
            
            # FINALMENTE: hacer todos los merges después de escribir todos los datos
            for merge_range in merges_to_do:
                try:
                    ws.merge_cells(merge_range)
                except Exception as e:
                    # Si hay un error en un merge, continuar con los demás
                    print(f"Advertencia: No se pudo hacer merge de {merge_range}: {e}")
            
            # Fila de totales
            total_row = len(productos_cat) + 6
            total_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            total_font = Font(bold=True, color="FFFFFF")
            
            ws[f'A{total_row}'].value = 'TOTALES'
            ws[f'A{total_row}'].font = total_font
            ws[f'A{total_row}'].fill = total_fill
            ws[f'A{total_row}'].border = border
            
            # Fórmulas de totales (solo columnas principales)
            ws[f'B{total_row}'].value = f'=SUM(B6:B{total_row-1})'  # Saldo real
            ws[f'C{total_row}'].value = f'=SUM(C6:C{total_row-1})'  # Valor total
            
            # Formato de totales
            for col in ['B', 'C']:
                cell = ws[f'{col}{total_row}']
                cell.font = total_font
                cell.fill = total_fill
                cell.border = border
                cell.alignment = center_alignment
                if col == 'C':
                    cell.number_format = '#,##0'
            
            # Ajustar ancho de columnas principales
            ws.column_dimensions['A'].width = 40  # NOMBRE - más ancha
            ws.column_dimensions['B'].width = 12  # SALDO REAL
            ws.column_dimensions['C'].width = 15  # VALOR TOTAL
            
            # Separadores
            ws.column_dimensions[separator1_col_letter].width = 3
            ws.column_dimensions[separator2_col_letter].width = 3
            
            # Columnas de ENTRADAS (6 columnas por entrada)
            for ent_idx in range(max_entradas):
                start_col = start_entradas_col + (ent_idx * entrada_cols)
                ws.column_dimensions[get_column_letter(start_col)].width = 11     # FECHA
                ws.column_dimensions[get_column_letter(start_col + 1)].width = 15 # FACTURA
                ws.column_dimensions[get_column_letter(start_col + 2)].width = 10 # CANTIDAD
                ws.column_dimensions[get_column_letter(start_col + 3)].width = 15 # PROVEEDOR
                ws.column_dimensions[get_column_letter(start_col + 4)].width = 12 # VALOR UNIT.
                ws.column_dimensions[get_column_letter(start_col + 5)].width = 12 # TOTAL
            
            # Columnas de SALIDAS (3 columnas por salida)
            for sal_idx in range(max_salidas):
                start_col = start_salidas_col + (sal_idx * salida_cols)
                ws.column_dimensions[get_column_letter(start_col)].width = 11     # FECHA
                ws.column_dimensions[get_column_letter(start_col + 1)].width = 10 # CANTIDAD
                ws.column_dimensions[get_column_letter(start_col + 2)].width = 8  # UNIDAD
            
            # Congelar primeras filas y columnas para mejor navegación
            ws.freeze_panes = 'D6'  # Congelar hasta columna C (datos principales) y fila 5 (encabezados)
        
        # Guardar en archivo temporal
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            wb.save(tmp_file.name)
            tmp_file_path = tmp_file.name
        
        # Enviar archivo
        return send_file(
            tmp_file_path,
            as_attachment=True,
            download_name=f'Inventario_Automatico_{periodo}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Error al generar Excel: {str(e)}', 'error')
        return redirect(url_for('inventarios'))

@app.route('/migrate-inventory-monthly')
@login_required
def migrate_inventory_monthly():
    """Migrar inventarios existentes al sistema mensual"""
    if not current_user.is_admin:
        flash('Solo los administradores pueden migrar inventarios', 'error')
        return redirect(url_for('inventarios'))
    
    try:
        # Agregar columna periodo si no existe
        from sqlalchemy import text
        with db.engine.connect() as conn:
            try:
                conn.execute(text("ALTER TABLE producto ADD COLUMN IF NOT EXISTS periodo VARCHAR(7)"))
                conn.execute(text("ALTER TABLE movimiento_inventario ADD COLUMN IF NOT EXISTS periodo VARCHAR(7)"))
                conn.commit()
                print("✅ Columnas periodo agregadas")
            except Exception as e:
                print(f"⚠️ Error agregando columnas: {e}")
        
        # Obtener productos sin período (después de agregar la columna)
        try:
            productos_sin_periodo = Producto.query.filter_by(periodo=None).all()
            
            if not productos_sin_periodo:
                flash('No hay productos para migrar', 'info')
                return redirect(url_for('inventarios'))
            
            # Asignar período actual a productos existentes
            periodo_actual = get_periodo_actual()
            productos_migrados = 0
            
            for producto in productos_sin_periodo:
                producto.periodo = periodo_actual
                productos_migrados += 1
            
            db.session.commit()
            
            flash(f'{productos_migrados} productos migrados al período {periodo_actual}', 'success')
            
        except Exception as e:
            print(f"⚠️ Error migrando productos: {e}")
            flash(f'Columnas agregadas, pero error migrando productos: {str(e)}', 'warning')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error durante la migración: {str(e)}', 'error')
    
    return redirect(url_for('inventarios'))

@app.route('/inventarios/importar', methods=['GET', 'POST'])
@login_required
def importar_inventarios():
    """Importar inventarios desde archivos Excel"""
    if not current_user.is_admin:
        flash('Solo los administradores pueden importar inventarios', 'error')
        return redirect(url_for('inventarios'))
    
    if request.method == 'POST':
        try:
            # Verificar que se subió un archivo
            if 'archivo_excel' not in request.files:
                flash('No se seleccionó ningún archivo', 'error')
                return redirect(url_for('importar_inventarios'))
            
            archivo = request.files['archivo_excel']
            if archivo.filename == '':
                flash('No se seleccionó ningún archivo', 'error')
                return redirect(url_for('importar_inventarios'))
            
            # Verificar extensión
            if not archivo.filename.lower().endswith(('.xlsx', '.xls')):
                flash('Solo se permiten archivos Excel (.xlsx, .xls)', 'error')
                return redirect(url_for('importar_inventarios'))
            
            # Obtener tipo de inventario y período
            tipo_inventario = request.form.get('tipo_inventario')
            periodo_importacion = request.form.get('periodo', get_periodo_actual())
            
            if not tipo_inventario:
                flash('Debe seleccionar el tipo de inventario', 'error')
                return redirect(url_for('importar_inventarios'))
            
            # Validar formato del período
            try:
                datetime.strptime(periodo_importacion, '%Y-%m')
            except ValueError:
                flash('Formato de período inválido. Use YYYY-MM', 'error')
                return redirect(url_for('importar_inventarios'))
            
            # Guardar archivo temporalmente
            import tempfile
            import os
            from openpyxl import load_workbook
            from sqlalchemy import text
            
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                archivo.save(tmp_file.name)
                
                try:
                    # Cargar archivo Excel
                    wb = load_workbook(tmp_file.name)
                    ws = wb.active
                    
                    # Validar categoría
                    categorias_validas = ['ALMACEN GENERAL', 'QUIMICOS', 'POSCOSECHA']
                    if tipo_inventario not in categorias_validas:
                        flash(f'Error: La categoría "{tipo_inventario}" no es válida. Use: {", ".join(categorias_validas)}', 'error')
                        return redirect(url_for('importar_inventarios'))
                    
                    # Conectar a la base de datos
                    with db.engine.connect() as conn:
                        productos_importados = 0
                        productos_duplicados = 0
                        errores = []
                        
                        # Procesar filas según el tipo de inventario
                        for row in range(2, ws.max_row + 1):  # Saltar encabezado
                            try:
                                # Función auxiliar para limpiar valores
                                def limpiar_valor(celda):
                                    valor = ws[celda].value
                                    if valor is None:
                                        return ''
                                    return str(valor).strip()
                                
                                def limpiar_numero(celda, default=0):
                                    valor = ws[celda].value
                                    if valor is None:
                                        return default
                                    try:
                                        # Convertir a string y limpiar
                                        str_valor = str(valor).strip()
                                        if not str_valor:
                                            return default
                                        # Reemplazar comas por puntos y convertir a float
                                        return float(str_valor.replace(',', '.'))
                                    except (ValueError, TypeError):
                                        return default
                                
                                if tipo_inventario == 'ALMACEN GENERAL':
                                    # Estructura: PRODUCTO, SALDO, FECHA, N. FACTURA, PROVE, CANT, VALOR UND, VALOR TOTAL
                                    producto = limpiar_valor(f'B{row}').upper()
                                    saldo = limpiar_numero(f'C{row}')
                                    proveedor = limpiar_valor(f'F{row}').upper()
                                    valor_und = limpiar_numero(f'H{row}')
                                    
                                elif tipo_inventario == 'QUIMICOS':
                                    # Estructura: CLASE, PRODUCTO, SALDO REAL, FECHA, FACTURA, PROVE, CANT, VALOR C/U, TOTAL
                                    clase = limpiar_valor(f'B{row}').upper()
                                    producto = limpiar_valor(f'C{row}').upper()
                                    saldo = limpiar_numero(f'D{row}')
                                    proveedor = limpiar_valor(f'G{row}').upper()
                                    valor_und = limpiar_numero(f'I{row}')
                                    
                                elif tipo_inventario == 'POSCOSECHA':
                                    # Estructura: PRODUCTO, SALDO, FECHA, N. FACTURA, PROVE, CANT, VALOR UND, VALOR TOTAL
                                    producto = limpiar_valor(f'A{row}').upper()
                                    saldo = limpiar_numero(f'B{row}')
                                    proveedor = limpiar_valor(f'E{row}').upper()
                                    valor_und = limpiar_numero(f'G{row}')
                                
                                # Validar que el producto no esté vacío
                                if not producto or producto == "" or producto == "NONE":
                                    continue
                                
                                # Generar código único basado en el nombre del producto
                                prefijo = {'ALMACEN GENERAL': 'ALM', 'QUIMICOS': 'QUI', 'POSCOSECHA': 'POS'}
                                # Crear código más legible usando las primeras letras del producto
                                codigo_base = ''.join([c for c in producto if c.isalnum()])[:8]
                                codigo = f"{prefijo[tipo_inventario]}-{codigo_base}-{row-1:03d}"
                                
                                # Verificar si ya existe en el mismo período y categoría
                                result = conn.execute(text("""
                                    SELECT id FROM producto WHERE codigo = :codigo AND categoria = :categoria AND periodo = :periodo
                                """), {
                                    'codigo': codigo,
                                    'categoria': tipo_inventario,
                                    'periodo': periodo_importacion
                                })
                                
                                if result.fetchone():
                                    productos_duplicados += 1
                                    continue
                                
                                # Insertar producto con manejo de errores mejorado
                                descripcion = f'Importado desde Excel - {tipo_inventario} - {periodo_importacion}'
                                if tipo_inventario == 'QUIMICOS' and clase:
                                    descripcion += f' - Clase: {clase}'
                                
                                # Asegurar que los valores no sean None
                                precio_final = valor_und if valor_und > 0 else 0
                                saldo_final = int(saldo) if saldo >= 0 else 0
                                proveedor_final = proveedor if proveedor else 'SIN PROVEEDOR'
                                
                                conn.execute(text("""
                                    INSERT INTO producto (
                                        codigo, nombre, descripcion, categoria, periodo, unidad_medida,
                                        precio_unitario, stock_actual, saldo_inicial, proveedor, activo, created_at
                                    ) VALUES (
                                        :codigo, :nombre, :descripcion, :categoria, :periodo, :unidad_medida,
                                        :precio_unitario, :stock_actual, :saldo_inicial, :proveedor, true, CURRENT_TIMESTAMP
                                    )
                                """), {
                                    'codigo': codigo,
                                    'nombre': producto,
                                    'descripcion': descripcion,
                                    'categoria': tipo_inventario,
                                    'periodo': periodo_importacion,
                                    'unidad_medida': 'UNIDAD',
                                    'precio_unitario': precio_final,
                                    'stock_actual': saldo_final,
                                    'saldo_inicial': saldo_final,
                                    'proveedor': proveedor_final
                                })
                                
                                productos_importados += 1
                                
                            except Exception as e:
                                errores.append(f"Fila {row}: {str(e)}")
                                continue
                        
                        conn.commit()
                        
                        # Mensaje de resultado
                        mensaje = f"Importación completada: {productos_importados} productos importados"
                        if productos_duplicados > 0:
                            mensaje += f", {productos_duplicados} duplicados omitidos"
                        if errores:
                            mensaje += f", {len(errores)} errores"
                        
                        flash(mensaje, 'success' if productos_importados > 0 else 'warning')
                        
                        if errores and len(errores) <= 10:  # Mostrar solo los primeros 10 errores
                            for error in errores[:10]:
                                flash(f"Error: {error}", 'error')
                
                finally:
                    # Limpiar archivo temporal
                    try:
                        os.unlink(tmp_file.name)
                    except:
                        pass  # Ignorar errores al eliminar archivo temporal
            
            return redirect(url_for('productos_inventario'))
            
        except Exception as e:
            flash(f'Error durante la importación: {str(e)}', 'error')
            return redirect(url_for('importar_inventarios'))
    
    else:
        # GET: Mostrar formulario de importación
        categorias_fijas = ['ALMACEN GENERAL', 'QUIMICOS', 'POSCOSECHA']
        periodo_actual = get_periodo_actual()
        return render_template('importar_inventarios.html', 
                             categorias=categorias_fijas,
                             periodo_actual=periodo_actual)

@app.route('/inventarios/movimientos/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_movimiento_inventario():
    """Registrar nuevo movimiento de inventario"""
    if request.method == 'POST':
        try:
            producto_id = request.form['producto_id']
            tipo_movimiento = request.form['tipo_movimiento']
            cantidad = int(request.form['cantidad'])
            # Manejar precio unitario con opción NA
            precio_na = request.form.get('precio_na', '0')
            if precio_na == '1':  # NA seleccionado
                precio_unitario = 0
            else:
                precio_unitario = float(request.form.get('precio_unitario', 0))
            
            motivo = request.form.get('motivo', '').strip()
            referencia = request.form.get('referencia', '').strip()
            responsable = request.form.get('responsable', '').strip()
            observaciones = request.form.get('observaciones', '').strip()
            proveedor = request.form.get('proveedor', '').strip()
            
            # Campos para sistema de empaques
            tipo_ingreso = request.form.get('tipo_ingreso', 'INDIVIDUAL')
            
            # Solo asignar valores de empaques si el tipo es EMPAQUE
            if tipo_ingreso == 'EMPAQUE':
                cantidad_empaques = int(request.form.get('cantidad_empaques', 0)) if request.form.get('cantidad_empaques') else None
                contenido_por_empaque = float(request.form.get('contenido_por_empaque', 0)) if request.form.get('contenido_por_empaque') else None
                
                # Manejar precio por empaque con opción NA
                precio_empaque_na = request.form.get('precio_empaque_na', '0')
                if precio_empaque_na == '1':  # NA seleccionado
                    precio_por_empaque = 0
                else:
                    precio_por_empaque = float(request.form.get('precio_por_empaque', 0))
            else:
                cantidad_empaques = None
                contenido_por_empaque = None
                precio_por_empaque = None
            
            # Validaciones
            if not producto_id or not tipo_movimiento or not cantidad:
                flash('Los campos producto, tipo de movimiento y cantidad son obligatorios', 'error')
                return redirect(url_for('nuevo_movimiento_inventario'))
            
            if cantidad <= 0:
                flash('La cantidad debe ser mayor a 0', 'error')
                return redirect(url_for('nuevo_movimiento_inventario'))
            
            producto = Producto.query.get_or_404(producto_id)
            
            # Verificar stock para salidas
            if tipo_movimiento == 'SALIDA' and producto.stock_actual < cantidad:
                flash(f'Stock insuficiente. Stock actual: {producto.stock_actual}', 'error')
                return redirect(url_for('nuevo_movimiento_inventario'))
            
            if tipo_movimiento == 'ENTRADA' and not proveedor:
                flash('El proveedor es obligatorio para las entradas.', 'error')
                return redirect(url_for('nuevo_movimiento_inventario'))
            if tipo_movimiento == 'SALIDA':
                proveedor = ''
            
            # Calcular precio y total según tipo de ingreso
            if tipo_movimiento == 'ENTRADA':
                if tipo_ingreso == 'EMPAQUE' and cantidad_empaques and contenido_por_empaque and precio_por_empaque:
                    # Ingreso por empaques
                    precio_final = precio_por_empaque / contenido_por_empaque  # Precio por unidad base
                    total = cantidad_empaques * precio_por_empaque
                    # Actualizar cantidad a la cantidad total en unidad base
                    cantidad = int(cantidad_empaques * contenido_por_empaque)
                else:
                    # Ingreso individual
                    precio_final = precio_unitario
                    total = cantidad * precio_unitario
            else:
                precio_final = 0
                total = 0
            
            # Obtener período del producto
            periodo_movimiento = producto.periodo if hasattr(producto, 'periodo') and producto.periodo else get_periodo_actual()
            
            # Crear movimiento
            nuevo_movimiento = MovimientoInventario(
                producto_id=producto_id,
                periodo=periodo_movimiento,
                tipo_movimiento=tipo_movimiento,
                cantidad=cantidad,
                precio_unitario=precio_final,
                total=total,
                motivo=motivo,
                referencia=referencia,
                tipo_ingreso=tipo_ingreso,
                cantidad_empaques=cantidad_empaques,
                contenido_por_empaque=contenido_por_empaque,
                precio_por_empaque=precio_por_empaque,
                responsable=responsable,
                observaciones=observaciones,
                created_by=current_user.id,
                proveedor=proveedor or None
            )
            
            # Actualizar stock del producto automáticamente
            db.session.add(nuevo_movimiento)
            db.session.flush()  # Asegura que el movimiento esté en la sesión
            
            # Actualizar stock basado en el movimiento
            stock_anterior = producto.stock_actual
            if tipo_movimiento == 'ENTRADA':
                producto.stock_actual += cantidad
                # Actualizar proveedor del producto si se proporciona
                if proveedor:
                    producto.proveedor = proveedor
            elif tipo_movimiento == 'SALIDA':
                producto.stock_actual -= cantidad
            
            db.session.commit()
            
            # Verificar si stock está bajo
            if producto.verificar_stock_bajo():
                flash(f'⚠️ Movimiento registrado. Stock actual: {producto.stock_actual} (⚠️ STOCK BAJO - Mínimo: {producto.stock_minimo})', 'warning')
            else:
                flash(f'✓ Movimiento registrado exitosamente. Stock actualizado: {stock_anterior} → {producto.stock_actual}', 'success')
            
            return redirect(url_for('movimientos_inventario'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al registrar el movimiento: {str(e)}', 'error')
            return redirect(url_for('nuevo_movimiento_inventario'))
    
    productos = Producto.query.filter_by(activo=True).all()
    return render_template('nuevo_movimiento_inventario.html', productos=productos)

@app.route('/inventarios/movimientos/eliminar/<int:id>', methods=['DELETE'])
@login_required
def eliminar_movimiento_inventario(id):
    """Eliminar un movimiento de inventario y revertir el stock"""
    try:
        movimiento = MovimientoInventario.query.get_or_404(id)
        producto = movimiento.producto
        
        # Guardar información para el mensaje
        stock_anterior = producto.stock_actual
        cantidad = movimiento.cantidad
        tipo_movimiento = movimiento.tipo_movimiento
        
        # Revertir el stock
        if tipo_movimiento == 'ENTRADA':
            # Si era entrada, restar del stock
            producto.stock_actual -= cantidad
        elif tipo_movimiento == 'SALIDA':
            # Si era salida, sumar al stock
            producto.stock_actual += cantidad
        
        # Eliminar el movimiento
        db.session.delete(movimiento)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Movimiento eliminado exitosamente',
            'stock_anterior': stock_anterior,
            'stock_actual': producto.stock_actual,
            'cantidad_revertida': cantidad,
            'tipo_movimiento': tipo_movimiento
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error al eliminar movimiento: {str(e)}'
        }), 500

@app.route('/inventarios/productos/<int:id>/kardex')
@login_required
def kardex_producto(id):
    """Ver kardex detallado de un producto (historial con saldo running)"""
    producto = Producto.query.get_or_404(id)
    
    # Obtener todos los movimientos ordenados por fecha
    movimientos = MovimientoInventario.query.filter_by(producto_id=id).order_by(MovimientoInventario.fecha_movimiento.asc()).all()
    
    # Calcular saldo running para cada movimiento
    kardex = []
    saldo_running = producto.saldo_inicial
    
    for mov in movimientos:
        if mov.tipo_movimiento == 'ENTRADA':
            saldo_running += mov.cantidad
        else:  # SALIDA
            saldo_running -= mov.cantidad
        
        kardex.append({
            'fecha': mov.fecha_movimiento,
            'tipo': mov.tipo_movimiento,
            'cantidad': mov.cantidad,
            'precio_unitario': mov.precio_unitario,
            'total': mov.total,
            'motivo': mov.motivo,
            'referencia': mov.referencia,
            'responsable': mov.responsable,
            'observaciones': mov.observaciones,
            'saldo': saldo_running,
            'usuario': mov.usuario.username if mov.usuario else 'N/A'
        })
    
    # Calcular totales
    total_entradas = producto.calcular_entradas()
    total_salidas = producto.calcular_salidas()
    saldo_final = producto.calcular_saldo_final()
    
    return render_template('kardex_producto.html',
                         producto=producto,
                         kardex=kardex,
                         total_entradas=total_entradas,
                         total_salidas=total_salidas,
                         saldo_final=saldo_final)

# ===== RUTAS PARA SISTEMA DE NOTIFICACIONES =====

@app.route('/api/notificaciones')
@login_required
def api_notificaciones():
    """API para obtener notificaciones"""
    no_leidas = request.args.get('no_leidas', 'false').lower() == 'true'
    print(f"📡 API notificaciones llamada - no_leidas: {no_leidas}")
    return obtener_notificaciones_api(no_leidas)

@app.route('/api/notificaciones/<int:notificacion_id>/leida', methods=['POST'])
@login_required
def api_marcar_notificacion_leida(notificacion_id):
    """API para marcar una notificación como leída"""
    return marcar_notificacion_leida_api(notificacion_id)

@app.route('/api/notificaciones/marcar-todas-leidas', methods=['POST'])
@login_required
def api_marcar_todas_leidas():
    """API para marcar todas las notificaciones como leídas"""
    try:
        # Marcar todas como leídas en la base de datos
        notificaciones = Notificacion.query.filter_by(leida=False).all()
        for notificacion in notificaciones:
            notificacion.leida = True
        db.session.commit()
        print(f"✅ {len(notificaciones)} notificaciones marcadas como leídas en BD")
        return jsonify({'success': True, 'message': 'Todas las notificaciones marcadas como leídas'})
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error marcando todas como leídas: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/notificaciones/<int:notificacion_id>/eliminar', methods=['DELETE'])
@login_required
def api_eliminar_notificacion(notificacion_id):
    """API para eliminar una notificación específica"""
    try:
        # Eliminar de la base de datos
        notificacion = Notificacion.query.get(notificacion_id)
        if notificacion:
            db.session.delete(notificacion)
            db.session.commit()
            print(f"🗑️ Notificación {notificacion_id} eliminada de la BD")
            return jsonify({'success': True, 'message': 'Notificación eliminada'})
        else:
            return jsonify({'success': False, 'message': 'Notificación no encontrada'}), 404
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error eliminando notificación: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/notificaciones/limpiar', methods=['POST'])
@login_required
def api_limpiar_notificaciones():
    """API para limpiar todas las notificaciones"""
    try:
        # Eliminar todas las notificaciones de la base de datos
        count = Notificacion.query.count()
        Notificacion.query.delete()
        db.session.commit()
        print(f"🗑️ {count} notificaciones eliminadas de la BD")
        return jsonify({'success': True, 'message': f'{count} notificaciones eliminadas'})
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error limpiando notificaciones: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/notificaciones/crear', methods=['POST'])
@login_required
def api_crear_notificacion():
    """API para crear una notificación manual"""
    try:
        data = request.get_json()
        titulo = data.get('titulo', 'Notificación')
        mensaje = data.get('mensaje', '')
        tipo = data.get('tipo', 'info')
        tipo_sonido = data.get('tipo_sonido', 'alerta')
        
        notificacion_id = notificacion_manager.agregar_notificacion(
            titulo=titulo,
            mensaje=mensaje,
            tipo=tipo,
            tipo_sonido=tipo_sonido
        )
        
        return jsonify({
            'success': True,
            'message': 'Notificación creada',
            'notificacion_id': notificacion_id
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/sounds/<path:filename>')
def servir_sonidos(filename):
    """Servir archivos de sonido"""
    try:
        return send_file(f'sounds/{filename}', mimetype='audio/wav')
    except FileNotFoundError:
        return jsonify({'error': 'Archivo de sonido no encontrado'}), 404

@app.route('/test-notificacion')
@login_required
def test_notificacion():
    """Ruta de prueba para notificaciones"""
    try:
        print("🧪 TEST: Creando notificación de prueba...")
        notif_id = notificar_asistencia_entrada("Usuario Prueba", "12:00")
        print(f"✅ TEST: Notificación creada con ID: {notif_id}")
        flash('Notificación de prueba creada', 'success')
        return redirect(url_for('dashboard'))
    except Exception as e:
        print(f"❌ TEST: Error creando notificación: {e}")
        flash(f'Error en prueba: {str(e)}', 'error')
        return redirect(url_for('dashboard'))

@app.route('/limpiar-bd')
@login_required
def limpiar_bd_ruta():
    """Ruta temporal para limpiar base de datos"""
    if not current_user.is_admin:
        flash('No tienes permisos para esta acción', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        from sqlalchemy import text
        
        # Lista de tablas a eliminar
        tablas_a_eliminar = [
            'asistencia', 'visitante', 'notificacion',
            'categoria_inventario', 'producto', 'movimiento_inventario', 'contrato_generado'
        ]
        
        mensajes = []
        mensajes.append("🧹 LIMPIANDO BASE DE DATOS")
        mensajes.append("📋 Manteniendo: user, empleado, contrato")
        
        # Eliminar tablas
        for tabla in tablas_a_eliminar:
            try:
                with db.engine.connect() as conn:
                    result = conn.execute(text(f"""
                        SELECT EXISTS (
                            SELECT FROM information_schema.tables 
                            WHERE table_name = '{tabla}'
                        );
                    """))
                    existe = result.fetchone()[0]
                    
                    if existe:
                        conn.execute(text(f"DROP TABLE IF EXISTS {tabla} CASCADE;"))
                        conn.commit()
                        mensajes.append(f"✅ Tabla {tabla} eliminada")
                    else:
                        mensajes.append(f"⚠️ Tabla {tabla} no existe")
            except Exception as e:
                mensajes.append(f"❌ Error con {tabla}: {str(e)}")
        
        # Regenerar secuencias
        mensajes.append("🔄 Regenerando secuencias...")
        secuencias = [('user', 'id'), ('empleado', 'id'), ('contrato', 'id')]
        
        for tabla, columna in secuencias:
            try:
                with db.engine.connect() as conn:
                    result = conn.execute(text(f"SELECT COALESCE(MAX({columna}), 0) FROM {tabla};"))
                    max_id = result.fetchone()[0]
                    
                    if max_id > 0:
                        conn.execute(text(f"ALTER SEQUENCE {tabla}_{columna}_seq RESTART WITH {max_id + 1};"))
                        conn.commit()
                        mensajes.append(f"✅ Secuencia {tabla}_{columna}_seq reiniciada en {max_id + 1}")
                    else:
                        conn.execute(text(f"ALTER SEQUENCE {tabla}_{columna}_seq RESTART WITH 1;"))
                        conn.commit()
                        mensajes.append(f"✅ Secuencia {tabla}_{columna}_seq reiniciada en 1")
            except Exception as e:
                mensajes.append(f"❌ Error con secuencia {tabla}_{columna}_seq: {str(e)}")
        
        mensajes.append("🎉 LIMPIEZA COMPLETADA")
        
        # Mostrar resultado
        resultado = "<br>".join(mensajes)
        return f"""
        <html>
        <head><title>Limpieza BD</title></head>
        <body>
            <h1>Limpieza de Base de Datos</h1>
            <pre>{resultado}</pre>
            <br><a href="/">Volver al Dashboard</a>
        </body>
        </html>
        """
        
    except Exception as e:
        return f"""
        <html>
        <head><title>Error</title></head>
        <body>
            <h1>Error en limpieza</h1>
            <p>Error: {str(e)}</p>
            <br><a href="/">Volver al Dashboard</a>
        </body>
        </html>
        """

@app.route('/limpiar-todo-bd')
@login_required
def limpiar_todo_bd_ruta():
    """Ruta para limpiar TODA la base de datos excepto usuarios"""
    if not current_user.is_admin:
        flash('No tienes permisos para esta acción', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        from sqlalchemy import text
        
        # Lista de tablas a eliminar (TODO excepto user)
        tablas_a_eliminar = [
            'asistencia', 'visitante', 'notificacion',
            'categoria_inventario', 'producto', 'movimiento_inventario', 
            'contrato_generado', 'contrato', 'empleado'
        ]
        
        mensajes = []
        mensajes.append("🧹 LIMPIANDO TODA LA BASE DE DATOS")
        mensajes.append("📋 Manteniendo SOLO: user (usuarios)")
        mensajes.append("🗑️ Eliminando TODO: empleados, contratos, asistencias, visitantes, inventario, notificaciones")
        
        # Eliminar tablas
        for tabla in tablas_a_eliminar:
            try:
                with db.engine.connect() as conn:
                    result = conn.execute(text(f"""
                        SELECT EXISTS (
                            SELECT FROM information_schema.tables 
                            WHERE table_name = '{tabla}'
                        );
                    """))
                    existe = result.fetchone()[0]
                    
                    if existe:
                        conn.execute(text(f"DROP TABLE IF EXISTS {tabla} CASCADE;"))
                        conn.commit()
                        mensajes.append(f"✅ Tabla {tabla} eliminada")
                    else:
                        mensajes.append(f"⚠️ Tabla {tabla} no existe")
            except Exception as e:
                mensajes.append(f"❌ Error con {tabla}: {str(e)}")
        
        # Regenerar secuencia de user
        mensajes.append("🔄 Regenerando secuencias...")
        try:
            with db.engine.connect() as conn:
                result = conn.execute(text("SELECT COALESCE(MAX(id), 0) FROM \"user\";"))
                max_id = result.fetchone()[0]
                
                if max_id > 0:
                    conn.execute(text(f"ALTER SEQUENCE user_id_seq RESTART WITH {max_id + 1};"))
                    conn.commit()
                    mensajes.append(f"✅ Secuencia user_id_seq reiniciada en {max_id + 1}")
                else:
                    conn.execute(text("ALTER SEQUENCE user_id_seq RESTART WITH 1;"))
                    conn.commit()
                    mensajes.append(f"✅ Secuencia user_id_seq reiniciada en 1")
        except Exception as e:
            mensajes.append(f"❌ Error con secuencia user_id_seq: {str(e)}")
        
        mensajes.append("🎉 LIMPIEZA COMPLETA EXITOSA")
        mensajes.append("⚠️ IMPORTANTE: Solo usuarios administradores quedaron")
        
        # Mostrar resultado
        resultado = "<br>".join(mensajes)
        return f"""
        <html>
        <head><title>Limpieza Completa BD</title></head>
        <body style="font-family: monospace; background: #f5f5f5; padding: 20px;">
            <h1>🧹 Limpieza Completa de Base de Datos</h1>
            <div style="background: white; padding: 20px; border-radius: 5px; border-left: 4px solid #dc3545;">
                <pre style="white-space: pre-wrap; word-wrap: break-word;">{resultado}</pre>
            </div>
            <br><a href="/" style="background: #007bff; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px;">Volver al Dashboard</a>
        </body>
        </html>
        """
        
    except Exception as e:
        return f"""
        <html>
        <head><title>Error</title></head>
        <body style="font-family: monospace; background: #f5f5f5; padding: 20px;">
            <h1>❌ Error en limpieza completa</h1>
            <div style="background: white; padding: 20px; border-radius: 5px; border-left: 4px solid #dc3545;">
                <p>Error: {str(e)}</p>
            </div>
            <br><a href="/" style="background: #007bff; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px;">Volver al Dashboard</a>
        </body>
        </html>
        """

if __name__ == '__main__':
    try:
        print("🚀 Iniciando aplicación...")
        print("🔄 Llamando a init_db()...")
        init_db()
        print("✅ init_db() completado exitosamente")
        port = int(os.environ.get('PORT', 5000))
        print(f"🌐 Servidor iniciado en puerto {port}")
        app.run(host='0.0.0.0', port=port, debug=False)
    except Exception as e:
        print(f"❌ Error al iniciar la aplicación: {str(e)}")
        import traceback
        traceback.print_exc()
        raise

# Configuración para gunicorn en producción
if __name__ != '__main__':
    # Inicializar la base de datos cuando se ejecuta con gunicorn
    try:
        print("🚀 Inicializando aplicación con gunicorn...")
        init_db()
        print("✅ Aplicación lista para gunicorn")
    except Exception as e:
        print(f"❌ Error al inicializar con gunicorn: {str(e)}")
        import traceback
        traceback.print_exc()
