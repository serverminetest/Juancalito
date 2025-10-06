#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para importar datos de inventarios desde archivos Excel
"""

import os
import sys
from sqlalchemy import create_engine, text
from openpyxl import load_workbook
from decimal import Decimal
import re

def get_database_url():
    """Obtener URL de la base de datos"""
    database_url = os.environ.get('DATABASE_PUBLIC_URL') or os.environ.get('DATABASE_URL')
    if not database_url:
        print("❌ Error: DATABASE_URL no está configurada")
        return None
    return database_url

def limpiar_texto(texto):
    """Limpiar y normalizar texto"""
    if not texto:
        return ""
    return str(texto).strip().upper()

def limpiar_numero(numero):
    """Limpiar y convertir número"""
    if not numero:
        return 0
    try:
        # Remover caracteres no numéricos excepto punto y coma
        numero_limpio = re.sub(r'[^\d.,]', '', str(numero))
        # Reemplazar coma por punto
        numero_limpio = numero_limpio.replace(',', '.')
        return float(numero_limpio)
    except:
        return 0

def crear_categorias_base(conn):
    """Crear categorías base de inventarios"""
    print("📝 Creando categorías base...")
    
    categorias = [
        {
            'nombre': 'ALMACEN GENERAL',
            'descripcion': 'Productos de almacén general'
        },
        {
            'nombre': 'QUIMICOS',
            'descripcion': 'Productos químicos y agroquímicos'
        },
        {
            'nombre': 'POSCOSECHA',
            'descripcion': 'Productos de poscosecha y empaque'
        }
    ]
    
    for categoria in categorias:
        try:
            # Verificar si ya existe
            result = conn.execute(text("""
                SELECT id FROM categoria_inventario 
                WHERE nombre = :nombre
            """), {'nombre': categoria['nombre']})
            
            if result.fetchone():
                print(f"✅ Categoría {categoria['nombre']} ya existe")
            else:
                # Crear categoría
                conn.execute(text("""
                    INSERT INTO categoria_inventario (nombre, descripcion, activa, created_at)
                    VALUES (:nombre, :descripcion, true, CURRENT_TIMESTAMP)
                """), categoria)
                print(f"✅ Categoría {categoria['nombre']} creada")
        except Exception as e:
            print(f"❌ Error creando categoría {categoria['nombre']}: {str(e)}")
    
    conn.commit()

def importar_almacen_general(conn):
    """Importar datos del archivo de almacén general"""
    print("\n📦 Importando ALMACÉN GENERAL...")
    print("=" * 50)
    
    try:
        # Obtener ID de categoría
        result = conn.execute(text("""
            SELECT id FROM categoria_inventario WHERE nombre = 'ALMACEN GENERAL'
        """))
        categoria_id = result.fetchone()[0]
        
        # Cargar archivo Excel
        archivo = "INVENTARIO ALMACEN SEPTIEMBRE-  2025 .xlsx"
        if not os.path.exists(archivo):
            print(f"❌ Archivo {archivo} no encontrado")
            return
        
        wb = load_workbook(archivo)
        ws = wb.active
        
        print(f"📊 Procesando {ws.max_row} filas...")
        
        productos_importados = 0
        productos_duplicados = 0
        
        for row in range(2, ws.max_row + 1):  # Saltar encabezado
            try:
                # Leer datos de la fila
                producto = limpiar_texto(ws[f'B{row}'].value)  # Columna B: PRODUCTO
                saldo = limpiar_numero(ws[f'C{row}'].value)    # Columna C: SALDO
                fecha = ws[f'D{row}'].value                    # Columna D: FECHA
                factura = limpiar_texto(ws[f'E{row}'].value)   # Columna E: N. FACTURA
                proveedor = limpiar_texto(ws[f'F{row}'].value) # Columna F: PROVE
                cantidad = limpiar_numero(ws[f'G{row}'].value) # Columna G: CANT
                valor_und = limpiar_numero(ws[f'H{row}'].value) # Columna H: VALOR UND
                valor_total = limpiar_numero(ws[f'I{row}'].value) # Columna I: VALOR TOTAL
                
                if not producto or producto == "":
                    continue
                
                # Generar código único
                codigo = f"ALM-{row-1:04d}"
                
                # Verificar si ya existe
                result = conn.execute(text("""
                    SELECT id FROM producto WHERE codigo = :codigo
                """), {'codigo': codigo})
                
                if result.fetchone():
                    productos_duplicados += 1
                    continue
                
                # Insertar producto
                conn.execute(text("""
                    INSERT INTO producto (
                        codigo, nombre, descripcion, categoria_id, unidad_medida,
                        precio_unitario, stock_actual, proveedor, activo, created_at
                    ) VALUES (
                        :codigo, :nombre, :descripcion, :categoria_id, :unidad_medida,
                        :precio_unitario, :stock_actual, :proveedor, true, CURRENT_TIMESTAMP
                    )
                """), {
                    'codigo': codigo,
                    'nombre': producto,
                    'descripcion': f'Importado desde {archivo}',
                    'categoria_id': categoria_id,
                    'unidad_medida': 'UNIDAD',
                    'precio_unitario': valor_und,
                    'stock_actual': int(saldo),
                    'proveedor': proveedor
                })
                
                productos_importados += 1
                
                if productos_importados % 50 == 0:
                    print(f"📊 Procesados: {productos_importados} productos...")
                
            except Exception as e:
                print(f"⚠️ Error en fila {row}: {str(e)}")
                continue
        
        conn.commit()
        print(f"✅ ALMACÉN GENERAL: {productos_importados} productos importados")
        if productos_duplicados > 0:
            print(f"⚠️ {productos_duplicados} productos duplicados omitidos")
        
    except Exception as e:
        print(f"❌ Error importando almacén general: {str(e)}")

def importar_quimicos(conn):
    """Importar datos del archivo de químicos"""
    print("\n🧪 Importando QUÍMICOS...")
    print("=" * 50)
    
    try:
        # Obtener ID de categoría
        result = conn.execute(text("""
            SELECT id FROM categoria_inventario WHERE nombre = 'QUIMICOS'
        """))
        categoria_id = result.fetchone()[0]
        
        # Cargar archivo Excel
        archivo = "INV QUIMICOS SEPTIEMBRE - 2025 (1) .xlsx"
        if not os.path.exists(archivo):
            print(f"❌ Archivo {archivo} no encontrado")
            return
        
        wb = load_workbook(archivo)
        ws = wb.active
        
        print(f"📊 Procesando {ws.max_row} filas...")
        
        productos_importados = 0
        productos_duplicados = 0
        
        for row in range(2, ws.max_row + 1):  # Saltar encabezado
            try:
                # Leer datos de la fila
                clase = limpiar_texto(ws[f'B{row}'].value)      # Columna B: CLASE
                producto = limpiar_texto(ws[f'C{row}'].value)   # Columna C: PRODUCTO
                saldo = limpiar_numero(ws[f'D{row}'].value)     # Columna D: SALDO REAL
                fecha = ws[f'E{row}'].value                     # Columna E: FECHA
                factura = limpiar_texto(ws[f'F{row}'].value)    # Columna F: FACTURA
                proveedor = limpiar_texto(ws[f'G{row}'].value)  # Columna G: PROVE
                cantidad = limpiar_numero(ws[f'H{row}'].value)  # Columna H: CANT
                valor_und = limpiar_numero(ws[f'I{row}'].value) # Columna I: VALOR C/U
                valor_total = limpiar_numero(ws[f'J{row}'].value) # Columna J: TOTAL
                
                if not producto or producto == "":
                    continue
                
                # Generar código único
                codigo = f"QUI-{row-1:04d}"
                
                # Verificar si ya existe
                result = conn.execute(text("""
                    SELECT id FROM producto WHERE codigo = :codigo
                """), {'codigo': codigo})
                
                if result.fetchone():
                    productos_duplicados += 1
                    continue
                
                # Insertar producto
                conn.execute(text("""
                    INSERT INTO producto (
                        codigo, nombre, descripcion, categoria_id, unidad_medida,
                        precio_unitario, stock_actual, proveedor, activo, created_at
                    ) VALUES (
                        :codigo, :nombre, :descripcion, :categoria_id, :unidad_medida,
                        :precio_unitario, :stock_actual, :proveedor, true, CURRENT_TIMESTAMP
                    )
                """), {
                    'codigo': codigo,
                    'nombre': producto,
                    'descripcion': f'Clase: {clase} - Importado desde {archivo}',
                    'categoria_id': categoria_id,
                    'unidad_medida': 'UNIDAD',
                    'precio_unitario': valor_und,
                    'stock_actual': int(saldo),
                    'proveedor': proveedor
                })
                
                productos_importados += 1
                
                if productos_importados % 50 == 0:
                    print(f"📊 Procesados: {productos_importados} productos...")
                
            except Exception as e:
                print(f"⚠️ Error en fila {row}: {str(e)}")
                continue
        
        conn.commit()
        print(f"✅ QUÍMICOS: {productos_importados} productos importados")
        if productos_duplicados > 0:
            print(f"⚠️ {productos_duplicados} productos duplicados omitidos")
        
    except Exception as e:
        print(f"❌ Error importando químicos: {str(e)}")

def importar_poscosecha(conn):
    """Importar datos del archivo de poscosecha"""
    print("\n🌱 Importando POSCOSECHA...")
    print("=" * 50)
    
    try:
        # Obtener ID de categoría
        result = conn.execute(text("""
            SELECT id FROM categoria_inventario WHERE nombre = 'POSCOSECHA'
        """))
        categoria_id = result.fetchone()[0]
        
        # Cargar archivo Excel
        archivo = "SALDOS POSCOSECHA  SEPTIEMBRE - 2025.xlsx"
        if not os.path.exists(archivo):
            print(f"❌ Archivo {archivo} no encontrado")
            return
        
        wb = load_workbook(archivo)
        ws = wb.active
        
        print(f"📊 Procesando {ws.max_row} filas...")
        
        productos_importados = 0
        productos_duplicados = 0
        
        for row in range(2, ws.max_row + 1):  # Saltar encabezado
            try:
                # Leer datos de la fila
                producto = limpiar_texto(ws[f'A{row}'].value)   # Columna A: PRODUCTO
                saldo = limpiar_numero(ws[f'B{row}'].value)     # Columna B: SALDO
                fecha = ws[f'C{row}'].value                     # Columna C: FECHA
                factura = limpiar_texto(ws[f'D{row}'].value)    # Columna D: N. FACTURA
                proveedor = limpiar_texto(ws[f'E{row}'].value)  # Columna E: PROVE
                cantidad = limpiar_numero(ws[f'F{row}'].value)  # Columna F: CANT
                valor_und = limpiar_numero(ws[f'G{row}'].value) # Columna G: VALOR UND
                valor_total = limpiar_numero(ws[f'H{row}'].value) # Columna H: VALOR TOTAL
                
                if not producto or producto == "":
                    continue
                
                # Generar código único
                codigo = f"POS-{row-1:04d}"
                
                # Verificar si ya existe
                result = conn.execute(text("""
                    SELECT id FROM producto WHERE codigo = :codigo
                """), {'codigo': codigo})
                
                if result.fetchone():
                    productos_duplicados += 1
                    continue
                
                # Insertar producto
                conn.execute(text("""
                    INSERT INTO producto (
                        codigo, nombre, descripcion, categoria_id, unidad_medida,
                        precio_unitario, stock_actual, proveedor, activo, created_at
                    ) VALUES (
                        :codigo, :nombre, :descripcion, :categoria_id, :unidad_medida,
                        :precio_unitario, :stock_actual, :proveedor, true, CURRENT_TIMESTAMP
                    )
                """), {
                    'codigo': codigo,
                    'nombre': producto,
                    'descripcion': f'Importado desde {archivo}',
                    'categoria_id': categoria_id,
                    'unidad_medida': 'UNIDAD',
                    'precio_unitario': valor_und,
                    'stock_actual': int(saldo),
                    'proveedor': proveedor
                })
                
                productos_importados += 1
                
                if productos_importados % 50 == 0:
                    print(f"📊 Procesados: {productos_importados} productos...")
                
            except Exception as e:
                print(f"⚠️ Error en fila {row}: {str(e)}")
                continue
        
        conn.commit()
        print(f"✅ POSCOSECHA: {productos_importados} productos importados")
        if productos_duplicados > 0:
            print(f"⚠️ {productos_duplicados} productos duplicados omitidos")
        
    except Exception as e:
        print(f"❌ Error importando poscosecha: {str(e)}")

def main():
    """Función principal"""
    print("🚀 IMPORTACIÓN DE INVENTARIOS DESDE EXCEL")
    print("=" * 60)
    
    # Obtener URL de la base de datos
    database_url = get_database_url()
    if not database_url:
        return False
    
    try:
        # Crear conexión
        engine = create_engine(database_url)
        
        with engine.connect() as conn:
            # Verificar conexión
            result = conn.execute(text("SELECT version();"))
            version = result.fetchone()[0]
            print(f"✅ Conectado a PostgreSQL: {version[:50]}...")
            
            # Crear categorías base
            crear_categorias_base(conn)
            
            # Importar datos de cada archivo
            importar_almacen_general(conn)
            importar_quimicos(conn)
            importar_poscosecha(conn)
            
            # Estadísticas finales
            print("\n📊 ESTADÍSTICAS FINALES:")
            print("=" * 60)
            
            result = conn.execute(text("""
                SELECT c.nombre, COUNT(p.id) as total_productos
                FROM categoria_inventario c
                LEFT JOIN producto p ON c.id = p.categoria_id
                GROUP BY c.id, c.nombre
                ORDER BY c.nombre;
            """))
            
            for row in result.fetchall():
                print(f"📦 {row[0]}: {row[1]} productos")
            
            result = conn.execute(text("SELECT COUNT(*) FROM producto;"))
            total_productos = result.fetchone()[0]
            print(f"\n🎉 TOTAL: {total_productos} productos importados")
            
            print("\n✅ ¡Importación completada exitosamente!")
            return True
            
    except Exception as e:
        print(f"❌ Error durante la importación: {str(e)}")
        return False

if __name__ == "__main__":
    success = main()
    if success:
        print("\n🚀 El sistema de inventarios está listo con datos reales!")
        sys.exit(0)
    else:
        print("\n❌ Importación falló - Revisar logs para más detalles")
        sys.exit(1)
