#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para leer y analizar los archivos de inventario Excel
"""

import openpyxl
import os
from datetime import datetime

def leer_archivo_excel(nombre_archivo):
    """Lee un archivo Excel y extrae información básica"""
    try:
        print(f"\n📊 Analizando: {nombre_archivo}")
        print("=" * 50)
        
        # Verificar si el archivo existe
        if not os.path.exists(nombre_archivo):
            print(f"❌ El archivo {nombre_archivo} no existe")
            return None
        
        # Cargar el archivo Excel
        workbook = openpyxl.load_workbook(nombre_archivo)
        
        print(f"✅ Archivo cargado exitosamente")
        print(f"📋 Hojas disponibles: {workbook.sheetnames}")
        
        # Analizar cada hoja
        for sheet_name in workbook.sheetnames:
            print(f"\n📄 Hoja: {sheet_name}")
            worksheet = workbook[sheet_name]
            
            # Obtener dimensiones
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            print(f"   📏 Dimensiones: {max_row} filas x {max_col} columnas")
            
            # Leer las primeras filas para entender la estructura
            print(f"   📝 Primeras 5 filas:")
            for row in range(1, min(6, max_row + 1)):
                row_data = []
                for col in range(1, min(11, max_col + 1)):  # Solo primeras 10 columnas
                    cell = worksheet.cell(row=row, column=col)
                    value = cell.value
                    if value is not None:
                        row_data.append(str(value)[:30])  # Limitar a 30 caracteres
                    else:
                        row_data.append("")
                
                if any(row_data):  # Solo mostrar filas que no estén vacías
                    print(f"      Fila {row}: {' | '.join(row_data)}")
            
            # Buscar patrones comunes en inventarios
            print(f"   🔍 Análisis de contenido:")
            
            # Buscar columnas que podrían ser códigos, nombres, cantidades, etc.
            header_row = 1
            headers = []
            for col in range(1, min(11, max_col + 1)):
                cell = worksheet.cell(row=header_row, column=col)
                if cell.value:
                    headers.append(str(cell.value).strip())
            
            if headers:
                print(f"      Encabezados encontrados: {headers}")
            
            # Buscar datos numéricos (posibles cantidades/precios)
            numericos = 0
            for row in range(2, min(20, max_row + 1)):  # Revisar primeras 20 filas
                for col in range(1, min(11, max_col + 1)):
                    cell = worksheet.cell(row=row, column=col)
                    if isinstance(cell.value, (int, float)):
                        numericos += 1
            
            print(f"      Celdas numéricas encontradas: {numericos}")
        
        workbook.close()
        return True
        
    except Exception as e:
        print(f"❌ Error al leer {nombre_archivo}: {str(e)}")
        return None

def main():
    """Función principal"""
    print("🚀 ANALIZADOR DE INVENTARIOS EXCEL")
    print("=" * 50)
    
    # Lista de archivos a analizar
    archivos = [
        "INVENTARIO ALMACEN SEPTIEMBRE-  2025 .xlsx",
        "INV QUIMICOS SEPTIEMBRE - 2025 (1) .xlsx", 
        "SALDOS POSCOSECHA  SEPTIEMBRE - 2025.xlsx"
    ]
    
    resultados = {}
    
    for archivo in archivos:
        print(f"\n🔍 Procesando: {archivo}")
        resultado = leer_archivo_excel(archivo)
        resultados[archivo] = resultado
    
    # Resumen final
    print(f"\n📊 RESUMEN FINAL")
    print("=" * 50)
    for archivo, resultado in resultados.items():
        status = "✅ Exitoso" if resultado else "❌ Error"
        print(f"{archivo}: {status}")
    
    print(f"\n💡 PRÓXIMOS PASOS:")
    print("1. Revisar la estructura de cada archivo")
    print("2. Identificar las columnas principales")
    print("3. Crear el sistema de inventarios basado en esta estructura")
    print("4. Implementar importación automática de datos")

if __name__ == "__main__":
    main()
